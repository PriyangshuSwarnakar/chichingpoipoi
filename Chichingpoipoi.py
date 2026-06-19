"""
Chichingpoipoi v2.0
Text Normalization and QC Tool
Created with Love and Designed with Care · All Rights Reserved © Priyangshu Swarnakar
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import re
import json
import csv
import unicodedata
import sys
import os
from pathlib import Path
from datetime import datetime


def resource_path(relative_path):
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, relative_path)


C = {
    "bg":         "#FFF8F0",
    "bg2":        "#FFF0E0",
    "accent":     "#FF9A1F",
    "accent_h":   "#E8870A",
    "accent_lt":  "#FFD199",
    "border":     "#FFCB8A",
    "text":       "#3B2A1A",
    "text_muted": "#8B6A4A",
    "white":      "#FFFFFF",
    "red":        "#C0392B",
    "green":      "#27AE60",
    "row_even":   "#FFF8F0",
    "row_odd":    "#FFF0E0",
}

FONT       = ("Segoe UI", 10)
FONT_BOLD  = ("Segoe UI", 10, "bold")
FONT_SMALL = ("Segoe UI", 9)
FONT_TITLE = ("Segoe UI", 13, "bold")


ONES = ["", "one", "two", "three", "four", "five", "six", "seven", "eight",
        "nine", "ten", "eleven", "twelve", "thirteen", "fourteen", "fifteen",
        "sixteen", "seventeen", "eighteen", "nineteen"]
TENS = ["", "", "twenty", "thirty", "forty", "fifty",
        "sixty", "seventy", "eighty", "ninety"]
ORDINAL_MAP = {
    "one":"first","two":"second","three":"third","four":"fourth","five":"fifth",
    "six":"sixth","seven":"seventh","eight":"eighth","nine":"ninth","ten":"tenth",
    "eleven":"eleventh","twelve":"twelfth","thirteen":"thirteenth",
    "fourteen":"fourteenth","fifteen":"fifteenth","sixteen":"sixteenth",
    "seventeen":"seventeenth","eighteen":"eighteenth","nineteen":"nineteenth",
    "twenty":"twentieth","thirty":"thirtieth","forty":"fortieth","fifty":"fiftieth",
    "sixty":"sixtieth","seventy":"seventieth","eighty":"eightieth","ninety":"ninetieth",
    "hundred":"hundredth","thousand":"thousandth","million":"millionth","billion":"billionth",
}

def _say_below_1000(n):
    if n == 0: return ""
    if n < 20: return ONES[n]
    elif n < 100:
        t, o = divmod(n, 10)
        return TENS[t] + ("-" + ONES[o] if o else "")
    else:
        h, r = divmod(n, 100)
        rest = (" and " + _say_below_1000(r)) if r else ""
        return ONES[h] + " hundred" + rest

def integer_to_words(n):
    if n < 0: return "minus " + integer_to_words(-n)
    if n == 0: return "zero"
    if n >= 1_000_000_000_000:
        return digit_by_digit(str(n))
    parts = []
    billions, n  = divmod(n, 1_000_000_000)
    millions, n  = divmod(n, 1_000_000)
    thousands, n = divmod(n, 1_000)
    if billions:  parts.append(_say_below_1000(billions) + " billion")
    if millions:  parts.append(_say_below_1000(millions) + " million")
    if thousands: parts.append(_say_below_1000(thousands) + " thousand")
    if n:
        if parts and n < 100:
            parts.append("and " + _say_below_1000(n))
        else:
            parts.append(_say_below_1000(n))
    return " ".join(parts)

INDIAN_PLACE_WORDS = [
    (10_000_000_00, "kharab"),   # 1,000,00,00,000
    (1_000_000_00,  "arab"),      # 100,00,00,000
    (10_000_000,    "crore"),     # 1,00,00,000
    (100_000,       "lakh"),      # 1,00,000
    (1_000,         "thousand"),  # 1,000
]

def integer_to_words_indian(n):
    """
    Convert an integer to words using the Indian numbering system
    (lakh / crore / arab groupings instead of million / billion).
    654321    -> "six lakh fifty-four thousand three hundred and twenty-one"
    100000    -> "one lakh"
    10000000  -> "one crore"
    1234567   -> "twelve lakh thirty-four thousand five hundred and sixty-seven"
    """
    if n < 0:
        return "minus " + integer_to_words_indian(-n)
    if n == 0:
        return "zero"
    if n >= 10_000_000_00_000:  # beyond kharab grouping — fall back to digits
        return digit_by_digit(str(n))
    parts = []
    remaining = n
    for value, word in INDIAN_PLACE_WORDS:
        if remaining >= value:
            count, remaining = divmod(remaining, value)
            parts.append(_say_below_1000(count) + " " + word)
    if remaining:
        if parts and remaining < 100:
            parts.append("and " + _say_below_1000(remaining))
        else:
            parts.append(_say_below_1000(remaining))
    return " ".join(parts)

INDIAN_NUMBERING_CONTEXT_WORDS = {
    "lakh","lakhs","lac","lacs","crore","crores","arab","kharab",
}

def _detect_indian_numbering(text, match_start, match_end, locale="US"):
    """
    Check whether the sentence containing this number mentions
    lakh/crore — if so, large integers should use Indian grouping.
    Looks at the whole line (not just a small window) since the
    "lakh"/"crore" hint word can appear anywhere in the sentence.

    locale: "India" | "US" | "UK"
      - "India": Indian numbering is the default for this line unless
        the text contains an explicit symbol/code override (handled
        by callers via _CURRENCY_ROWS / CURRENCY_CODE_MAP), but a
        lakh/crore word always confirms Indian numbering regardless
        of locale.
      - "US"/"UK": Indian numbering only applies when a lakh/crore
        word is present.
    """
    line_lower = text.lower()
    if any(w in line_lower for w in INDIAN_NUMBERING_CONTEXT_WORDS):
        return True
    return locale == "India"


def to_ordinal(word):
    parts = word.split()
    if not parts: return word
    last = parts[-1]
    if "-" in last:
        sub = last.split("-")
        sub[-1] = ORDINAL_MAP.get(sub[-1], sub[-1] + "th")
        parts[-1] = "-".join(sub)
    else:
        parts[-1] = ORDINAL_MAP.get(last, last + "th")
    return " ".join(parts)

def is_year_like(n, context_hint=None):
    """
    True if a number should use the "paired" year reading
    (e.g. 2026 -> twenty twenty six, 1500 -> fifteen hundred).

    context_hint: "year" | "quantity" | None
      - "year"     : force year-pair reading for any n in 1000-2099
      - "quantity" : never use year-pair reading
      - None       : use the default heuristic below

    Default heuristic (context_hint=None):
      - Non-round numbers in 1100-2099 (e.g. 2026, 1234) default to
        year-pair reading — this is the overwhelmingly common reading
        for 4-digit numbers in this range that aren't round hundreds.
      - Round hundreds/thousands (1000, 1100, 1500, 2000, 2100, ...)
        default to standard integer reading ("two thousand",
        "one thousand five hundred") UNLESS context_hint == "year",
        in which case they use the year-pair reading
        ("twenty hundred", "fifteen hundred").
    """
    if context_hint == "quantity":
        return False
    if not (1000 <= n <= 2099):
        return False
    if n % 100 == 0:
        return context_hint == "year"
    return 1100 <= n <= 2099

def year_to_words(n):
    high, low = divmod(n, 100)
    h = integer_to_words(high)
    if low == 0:
        return h + " hundred"
    if low < 10:
        return h + " oh " + ONES[low]
    l = integer_to_words(low)
    return h + " " + l

def number_token_to_words(token, context_hint=None, indian=False):
    """
    Convert a numeric token to words. Returns (words, type_str|None).
    context_hint: "year" | "quantity" | None — see is_year_like().
    indian: if True, large integers (and the integer part of decimals)
            use the Indian numbering system (lakh/crore) instead of
            million/billion.
    """
    ord_m = re.fullmatch(r'(-?\d+)(st|nd|rd|th)', token, re.I)
    if ord_m:
        n = int(ord_m.group(1))
        base = integer_to_words_indian(abs(n)) if indian else integer_to_words(abs(n))
        result = to_ordinal(base)
        return ("minus " + result if n < 0 else result), "ordinal"
    dec_m = re.fullmatch(r'-?\d+\.\d+', token)
    if dec_m:
        neg = token.startswith("-")
        parts = token.lstrip("-").split(".")
        left  = integer_to_words_indian(int(parts[0])) if indian else integer_to_words(int(parts[0]))
        raw_dec = parts[1]
        stripped = raw_dec.rstrip("0")
        trailing_zeros = len(raw_dec) - len(stripped)
        if stripped == "":
            if trailing_zeros == 1:
                dec_part = "0"      # 123.0 → "point zero"
            else:
                dec_part = ""       # 123.0000000 → integer only
        elif trailing_zeros <= 1:
            dec_part = raw_dec      # no trailing zeros, or exactly 1 → keep as-is
        else:
            dec_part = stripped     # 2+ trailing zeros → strip all
        if not dec_part:
            return ("minus " if neg else "") + left, "decimal"
        def _digit_word(d):
            n = int(d)
            return "zero" if n == 0 else (ONES[n] if n < len(ONES) else integer_to_words(n))
        right = " ".join(_digit_word(d) for d in dec_part)
        return ("minus " if neg else "") + left + " point " + right, "decimal"
    int_m = re.fullmatch(r'-?\d+', token)
    if int_m:
        n = int(token)
        if is_year_like(abs(n), context_hint):
            result = ("minus " if n < 0 else "") + year_to_words(abs(n))
        elif indian:
            result = integer_to_words_indian(n)
        else:
            result = integer_to_words(n)
        return result, "integer"
    return token, None


def time_to_words(hh, mm, period=None):
    """
    Convert HH:MM [AM/PM] to spoken form.

    Rule (all formats):
      - If minutes == 00: "X hundred"  (e.g. 07:00 → seven hundred, 17:00 → seventeen hundred)
      - If minutes 01-09: "X oh N"     (e.g. 12:02 → twelve oh two)
      - If minutes 10-59: "X [minutes]" via integer_to_words
                           (e.g. 16:45 → sixteen forty five, 17:30 → seventeen thirty)

    For 12-hour times with an AM/PM marker:
      - Same minute rule applies, then append "A M" or "P M".
      - Leading-zero hours like 07:00 AM → "seven A M" (drop the zero prefix).
      - 12:02 PM → "twelve oh two P M"

    The old "hundred" convention for h>=13 is replaced by this unified rule:
    minutes==0 always gives "X hundred" regardless of the hour.
    """
    h = int(hh)
    m = int(mm)

    h_word = integer_to_words(h) if h > 0 else "zero"

    if m == 0:
        m_word = "hundred"
    elif m < 10:
        m_word = "oh " + ONES[m]
    else:
        m_word = integer_to_words(m)

    result = h_word + " " + m_word

    if period:
        p = period.upper().replace("AM", "A M").replace("PM", "P M")
        result += " " + p

    return result.strip()

FLIGHT_PATTERN = re.compile(r'\b([A-Z]{2,3})(\d{1,4})\b')

FLIGHT_TRIGGER_WORDS = {
    "flight", "flights", "airline", "airlines", "service",
    "operated", "aboard", "board", "boarding",
}

def apply_flight_numbers(text):
    """
    Expand flight numbers to spoken form.
    VA8000 → V A zero zero zero zero
    QF4186 → Q F four one eight six
    Only fires when preceded by a flight trigger word OR when the token
    matches the strict 2-3 letter + 1-4 digit pattern.
    """
    changes = []

    def _spell_flight(letters, digits):
        letter_part = " ".join(list(letters.upper()))
        digit_part  = " ".join(
            "zero" if d == "0" else ONES[int(d)]
            for d in digits
        )
        return letter_part + " " + digit_part

    def _repl(m):
        letters = m.group(1)
        digits  = m.group(2)
        result  = _spell_flight(letters, digits)
        changes.append((m.group(0), result))
        return result

    trigger_pat = (
        r'(?i)(?:\b(?:' + '|'.join(FLIGHT_TRIGGER_WORDS) + r')'
        r'(?:\s+\w+){0,4}\s+)'
        r'([A-Z]{2,3})(\d{1,4})\b'
    )
    def _trigger_repl(m):
        letters = m.group(1)
        digits  = m.group(2)
        code_start = m.start(1) - m.start()
        result = m.group(0)[:code_start] + _spell_flight(letters, digits)
        changes.append((m.group(1)+m.group(2), _spell_flight(letters, digits)))
        return result

    text = re.sub(trigger_pat, _trigger_repl, text)
    return text, changes


PRODUCT_CODE_PATTERN = re.compile(r'\b([A-Z]{1,3})-?(\d{1,6})\b')

PASSPORT_CODE_PATTERN = re.compile(r'\b([A-Z])(\d{7,9})\b')

def apply_product_codes(text):
    """
    Expand standalone alphanumeric product/model codes.
    RX-4150 → R X four one five zero
    PX2100  → P X two one zero zero
    A320    → A three two zero
    A1234567 → A one two three four five six seven  (passport-style)
    Conservative: requires 1-3 uppercase letters directly followed by
    (optionally via hyphen) 2-6 digits, as a whole word (or, for
    passport-style codes, exactly 1 letter + 7-9 digits with no
    separator). Skips tokens that are valid units, abbreviations, or
    already-known dictionary entries.
    """
    changes = []

    def _repl(m):
        letters = m.group(1)
        digits  = m.group(2)
        letter_part = " ".join(list(letters))
        digit_part  = " ".join("zero" if d == "0" else ONES[int(d)] for d in digits)
        result = letter_part + " " + digit_part
        changes.append((m.group(0), result))
        return result

    text = re.sub(PRODUCT_CODE_PATTERN, _repl, text)
    text = re.sub(PASSPORT_CODE_PATTERN, _repl, text)
    return text, changes


CODE_TRIGGER_WORDS = {
    "pin","pins","otp","otps","cvv","cvc","cid","mpin","atm",
    "code","codes","id","ids","uid","guid","uuid","token","key","keycode",
    "serial","ref","reference","tracking","trackingid","awb","waybill",
    "consignment","shipment","parcel","courier",
    "booking","ticket","order","invoice","receipt","transaction","txn",
    "confirmation","voucher","coupon","claim","policy","case","ticketno",
    "account","acct","iban","swift","ifsc","routing","sortcode",
    "registration","reg","vin","chassis","license","licence","plate",
    "passport","aadhaar","aadhar","pan","ssn","nin","npi",
    "ip","ipv4","ipv6","subnet","port","ssid","dns",
    "asin","partno","partnumber","sn",
    "passcode","password","passphrase","accesscode","verification",
    "verificationcode","activationcode","securitycode",
    "combination","combo","locker","safe",
    "postcode","zipcode","zip","pincode",
    "no","no.","nos","number","num","numbers","#","ext","extension",
}

def digit_by_digit(num_str):
    """Spell each digit of a number string individually. 1234 → one two three four."""
    result = []
    for ch in num_str:
        if ch.isdigit():
            result.append(ONES[int(ch)] if int(ch) > 0 else "zero")
        elif ch == ".":
            result.append("point")
    return " ".join(result)

CODE_QUANTITY_FOLLOWERS = {
    "lines","line","items","item","copies","copy","times","time",
    "people","persons","units","pieces","entries","entry","records","record",
    "rows","row","columns","column","pages","page","files","file",
    "errors","error","warnings","warning","instances","instance",
    "occurrences","occurrence","attempts","attempt","tries","try",
    "percent","dollars","rupees","pounds","euros","cents","yen","won",
    "votes","vote","members","member","employees","employee",
    "students","student","customers","customer","users","user",
    "requests","request","responses","response","sessions","session",
    "downloads","download","views","view","clicks","click",
    "messages","message","emails","email","calls","call",
    "days","day","hours","hour","minutes","minute","seconds","second",
    "weeks","week","months","month","years","year",
}

def _next_word_is_unit_or_quantity(text, pos):
    """
    Check whether the word immediately following position `pos` (the end
    of a matched number) is a unit word or a quantity-follower word.
    If so, the number is a measurement/count, not a code/identifier.
    """
    rest = text[pos:].lstrip()
    m = re.match(r"[A-Za-z][A-Za-z²³/]*", rest)
    if not m:
        return False
    word = m.group(0)
    word_lower = word.lower().rstrip(".,;:!?")
    if word_lower in CODE_QUANTITY_FOLLOWERS:
        return True
    if word_lower in SINGULAR_PLURAL_UNIT_WORDS:
        return True
    if word_lower in WORD_UNIT_MAP:
        return True
    if word in UNIT_EXACT or word_lower in {k.lower() for k in UNIT_EXACT}:
        return True
    return False


SECURITY_TRIGGER_WORDS = {
    "security","question","answer","secret","hint","detail",
    "passcode","password","passphrase","pw","accesscode",
    "verificationcode","activationcode","securitycode",
    "combination","combo","locker","safe",
    "pin","otp","key","code","passcode",
}

ALL_GROUPING_TRIGGERS = CODE_TRIGGER_WORDS  # all triggers get steps 2+3

def _group_security_code_steps23(digits):
    """
    Apply only steps 2 and 3 of the security code grouping chain.
    Returns (spoken, False) if a match is found, or (None, None) if not.
    Step 2: all digits identical → double/triple notation
    Step 3: equal-length sub-groups each all-identical → double/triple per group
            (only fires when the number divides evenly into groups of 2 or 3)
    """
    n = len(digits)

    def _double_triple(d):
        """Speak a run of identical digits as double/triple X."""
        c = d[0]
        w = ONES[int(c)] if int(c) > 0 else "zero"
        size = len(d)
        if size == 2:
            return f"double {w}"
        if size == 3:
            return f"triple {w}"
        parts = []
        remaining = size
        while remaining > 0:
            if remaining == 4:
                parts.append(f"double {w}")
                parts.append(f"double {w}")
                remaining = 0
            elif remaining >= 3:
                parts.append(f"triple {w}")
                remaining -= 3
            else:
                parts.append(f"double {w}")
                remaining -= 2
        return " ".join(parts)

    if len(set(digits)) == 1:
        return _double_triple(digits), False

    for grp_size in (3, 2):
        if n % grp_size == 0:
            groups = [digits[i:i+grp_size] for i in range(0, n, grp_size)]
            if all(len(set(g)) == 1 for g in groups):
                spoken = " ".join(_double_triple(g) for g in groups)
                return spoken, False

    return None, None


def _group_security_code(digits):
    """
    Full grouping chain for security trigger words (steps 4-6).
    Steps 2+3 are pre-handled by _group_security_code_steps23.

    Step 4: ALL chunks identical AND no leading zero → speak each chunk as number.
            e.g. 181818 → eighteen eighteen eighteen (3 identical chunks of 2)
    Step 5: Sequential "round" pairs (multiples of 10, or obvious arithmetic
            progression with step ≤ 10) → speak as pairs.
            e.g. 102030 → ten twenty thirty
            Strict: only fires for pairs divisible by 10 or equal-step <= 10.
    Step 6: digit-by-digit (fallback, ambiguous=True)
    """
    n = len(digits)

    for chunk_size in (2, 3):
        if n % chunk_size == 0:
            chunks = [digits[i:i+chunk_size] for i in range(0, n, chunk_size)]
            if len(set(chunks)) == 1 and not chunks[0].startswith("0"):
                spoken = " ".join(integer_to_words(int(c)) for c in chunks)
                return spoken, False

    if n % 2 == 0 and n >= 4:
        pairs = [digits[i:i+2] for i in range(0, n, 2)]
        vals  = [int(p) for p in pairs]
        if (all(not p.startswith("0") for p in pairs)
                and all(v % 10 == 0 for v in vals)):
            diffs = [vals[i+1]-vals[i] for i in range(len(vals)-1)]
            if len(set(diffs)) == 1:   # constant arithmetic step
                spoken = " ".join(integer_to_words(v) for v in vals)
                return spoken, False

    return digit_by_digit(digits), True


def apply_code_numbers(text):
    """
    Spell numbers digit-by-digit when they represent a code/ID/PIN rather
    than a quantity. For security-type trigger words (question, answer,
    password, combination, etc.) smart grouping logic is applied instead
    of plain digit-by-digit.

    Returns (text, changes, ambiguous_flags) where ambiguous_flags is a
    list of (original_num_str, spoken_str) tuples needing human review.
    """
    changes = []
    ambiguous_flags = []
    all_triggers = CODE_TRIGGER_WORDS | SECURITY_TRIGGER_WORDS
    trigger_alt = '|'.join(re.escape(w) for w in sorted(all_triggers, key=len, reverse=True))

    pattern = (
        r'(?i)\b(' + trigger_alt + r')\b[:\-]?'
        r'((?:\s+[A-Za-z]+){0,5}?\s+)'
        r'(-?\d{3,18})(?!\.\d)\b'
    )

    def _repl(m):
        trigger   = m.group(1)
        connector = m.group(2) or ""
        num_str   = m.group(3)
        num_end   = m.end(3)

        if _next_word_is_unit_or_quantity(text, num_end):
            return m.group(0)

        connector_words = re.findall(r'[A-Za-z]+', connector)
        if any(w.lower() in YEAR_CONTEXT_WORDS for w in connector_words):
            return m.group(0)

        is_negative = num_str.startswith("-")
        digits = num_str.lstrip("-")
        trig_lower = trigger.lower()
        is_security = trig_lower in SECURITY_TRIGGER_WORDS

        if digits.startswith("0"):
            spoken = digit_by_digit(digits)
            is_ambiguous = False
        else:
            spoken_23, ambig_23 = _group_security_code_steps23(digits)
            if spoken_23 is not None:
                spoken, is_ambiguous = spoken_23, False
            elif is_security:
                spoken, is_ambiguous = _group_security_code(digits)
            else:
                spoken = digit_by_digit(digits)
                is_ambiguous = False

        if is_negative:
            spoken = "minus " + spoken
        if is_ambiguous:
            ambiguous_flags.append((num_str, spoken))

        result = trigger + connector + spoken
        changes.append((m.group(0).strip(), result))
        return result

    text = re.sub(pattern, _repl, text)
    return text, changes, ambiguous_flags

def apply_partial_expansion(text, tokens_config):
    """
    Apply partial-expansion rules to text.

    tokens_config: dict of {TOKEN: {"expand": str, "letters": str, "threshold_pct": int}}
      expand       = full spoken form (e.g. "gigabytes")
      letters      = letter-by-letter form (e.g. "G B")
      threshold_pct = % of occurrences to expand (first N); rest → letters

    e.g. with threshold=30% and 10 occurrences: first 3 → "gigabytes", rest → "G B"
    """
    changes = []
    for token, cfg in tokens_config.items():
        expand_form  = cfg.get("expand", "")
        letters_form = cfg.get("letters", " ".join(list(token)))
        threshold    = cfg.get("threshold_pct", 50)
        if not expand_form:
            continue
        positions = [m for m in re.finditer(r'\b' + re.escape(token) + r'\b', text)]
        if not positions:
            continue
        n_expand = max(1, round(len(positions) * threshold / 100))
        for idx, m in enumerate(reversed(positions)):
            real_idx = len(positions) - 1 - idx
            replacement = expand_form if real_idx < n_expand else letters_form
            changes.append((m.group(0), replacement))
            text = text[:m.start()] + replacement + text[m.end():]
    return text, changes


def apply_dotted_quad(text):
    """
    Convert dotted-quad IP addresses to spoken form.
    Each octet is spoken digit-by-digit, dots spoken as "dot".
      192.168.1.1   → one nine two dot one six eight dot one dot one
      10.0.0.1      → one zero dot zero dot zero dot one
      255.255.255.0 → two five five dot two five five dot two five five dot zero
    Only matches 4-octet dotted notation where each octet is 0-255.
    Does NOT match version numbers (1.2.3.4 preceded by a word like "v"),
    nor decimal chains that are clearly not IPs.
    """
    changes = []

    def _octet_words(oct_str):
        return " ".join(
            "zero" if d == "0" else ONES[int(d)] if int(d) < 10
            else integer_to_words(int(d))
            for d in oct_str
        ) if len(oct_str) == 1 else " ".join(
            "zero" if d == "0" else ONES[int(d)]
            for d in oct_str
        )

    def _quad_repl(m):
        octets = [m.group(i) for i in range(1, 5)]
        for o in octets:
            if int(o) > 255:
                return m.group(0)
        start = m.start()
        pre = text[max(0, start-3):start].strip()
        if re.search(r'[vV]\d*$|ver$|version$', pre):
            return m.group(0)
        spoken = " dot ".join(_octet_words(o) for o in octets)
        changes.append((m.group(0), spoken))
        return spoken

    text = re.sub(
        r'(?<!\d)(\d{1,3})\.(\d{1,3})\.(\d{1,3})\.(\d{1,3})(?!\d)',
        _quad_repl, text)
    return text, changes


    """Convert number ranges like 12-15 → twelve to fifteen."""
    changes = []
    def _range_repl(m):
        lo, hi = m.group(1), m.group(2)
        lo_w = integer_to_words(int(lo))
        hi_w = integer_to_words(int(hi))
        result = f"{lo_w} to {hi_w}"
        changes.append((m.group(0), result))
        return result
    text = re.sub(r'(?<![\d\w])(\d{1,4})-(\d{1,4})(?![\d\w])', _range_repl, text)
    return text, changes


def apply_k_suffix(text):
    """Convert number+k suffix meaning thousand: 42.13k → forty two point one three thousand."""
    changes = []
    def _k_repl(m):
        num_str = m.group(1)
        num_w, _ = number_token_to_words(num_str, "quantity")
        result = f"{num_w} thousand"
        changes.append((m.group(0), result))
        return result
    text = re.sub(r'(\d+(?:\.\d+)?)k\b', _k_repl, text, flags=re.IGNORECASE)
    return text, changes
def apply_dot_time(text):
    """
    Convert dot-notation times common in Australian English.
    12.15         → twelve fifteen  (only when followed by time words, or standalone at end of clause)
    15.45 hours   → fifteen forty five hours
    Does NOT fire when followed by a unit word (km, kg, carats etc.)
    """
    changes = []

    _DOT_TIME_CONTEXT = {
        "at","from","until","till","by","after","before","around","about",
        "time","clock","scheduled","departs","arrives","leaves","opens",
        "closes","starts","ends","begins","meeting","class","show","flight",
        "train","bus","ferry","appointment","session","shift","deadline",
    }

    def _dot_time_repl(m):
        hh = m.group(1)
        mm = m.group(2)
        suffix = (m.group(3) or "").strip()
        h = int(hh)
        mn = int(mm)
        if mn >= 60 or h > 23:
            return m.group(0)

        start = m.start()
        if start > 0 and text[start-1] in "$£€¥₹":
            return m.group(0)

        end = m.end()
        rest = text[end:].lstrip()
        next_word = re.match(r'[A-Za-z/%°][A-Za-z0-9/°²³μ]*', rest)
        has_time_suffix = False
        if next_word:
            nw = next_word.group(0)
            if nw.lower() in {"hours", "hour", "hrs", "hr", "sharp", "am", "pm"}:
                has_time_suffix = True
            else:
                return m.group(0)   # followed by a non-time word → decimal

        if not has_time_suffix:
            before = text[max(0, start-60):start]
            before_words = re.findall(r'[A-Za-z]+', before)[-6:]
            if not any(w.lower() in _DOT_TIME_CONTEXT for w in before_words):
                return m.group(0)

        spoken = time_to_words(hh.zfill(2), mm.zfill(2))
        result = spoken + (" " + suffix if suffix else "")
        changes.append((m.group(0).strip(), result))
        return " " + result + " "

    pattern = (
        r'(?<![\d.])(\d{1,2})\.(\d{2})' 
        r'(\s+(?:hours?|hrs?|sharp|[AaPp][Mm]))?'
        r'(?=[\s,;.!?]|$)'
    )
    text = re.sub(pattern, _dot_time_repl, text)
    text = re.sub(r' {2,}', ' ', text).strip()
    return text, changes
def apply_time(text):
    """
    Convert time expressions before number engine runs.
    Patterns:
      HH:MM AM/PM    →  spoken with period
      HH:MM hours    →  spoken + hours
      HH:MM          →  spoken (bare)
    """
    changes = []

    def _repl(m):
        hh     = m.group(1)
        mm     = m.group(2)
        period = m.group(3)   # AM/PM or None
        suffix = m.group(4)   # "hours" / "hrs" or None

        spoken = time_to_words(hh, mm, period)

        if suffix:
            if m_word_is_hundred(mm):
                result = spoken + " hours"
            else:
                result = spoken + " hours"
        else:
            result = spoken

        changes.append((m.group(0).strip(), result))
        return " " + result + " "

    def m_word_is_hundred(mm):
        return int(mm) == 0

    pattern = (
        r'\b(\d{1,2}):(\d{2})'          # HH:MM
        r'\s*([AaPp][Mm])?'              # optional AM/PM
        r'\s*\b(hours?|hrs?)?\b'         # optional hours/hrs
    )
    text = re.sub(pattern, _repl, text)
    text = re.sub(r' {2,}', ' ', text).strip()
    return text, changes


INVARIANT = {
    "hertz","lux","siemens","henry","farad","ohm","tesla","weber",
    "pascal","celsius","fahrenheit","kelvin","candela","mole","radian",
    "steradian","becquerel","gray","sievert","katal","decibel",
    "rpm","baud","ppm","ppb","ppt","percent",
    "bar","millibar","atmosphere",
    "gigatonne","megatonne",
}

def _u(singular, plural=None, invariant=False):
    """Helper to build unit entry."""
    if invariant or singular in INVARIANT:
        return (singular, singular)   # same form for singular and plural
    if plural is None:
        return (singular, singular + "s")
    return (singular, plural)

UNIT_EXACT = {
    "m":   _u("metre"),          # matched only via digit+unit pattern
    "g":   _u("gram"),           # matched only via digit+unit pattern
    "s":   _u("second"),         # matched only via digit+unit pattern
    "A":    _u("ampere"),         # matched only via digit+unit pattern
    "amp":  _u("ampere"),
    "amps": _u("ampere", "amperes"),
    "mol": _u("mole"),
    "cd":  _u("candela"),
    "km":  _u("kilometre"),
    "cm":  _u("centimetre"),
    "mm":  _u("millimetre"),
    "nm":  _u("nanometre"),
    "μm":  _u("micrometre"),
    "Å":   _u("angstrom"),
    "mi":  _u("mile"),
    "ft":  _u("foot", "feet"),
    "yd":  _u("yard"),
    "nmi": _u("nautical mile"),
    "kg":  _u("kilogram"),
    "mg":  _u("milligram"),
    "μg":  _u("microgram"),
    "µg":  _u("microgram"),    # U+00B5 micro sign variant
    "ng":  _u("nanogram"),
    "pg":  _u("picogram"),
    "t":   _u("tonne"),          # matched only via digit+unit pattern
    "lb":  _u("pound"),
    "oz":  _u("ounce"),
    "st":  _u("stone"),
    "L":   _u("litre"),
    "l":   _u("litre"),
    "mL":  _u("millilitre"),
    "ml":  _u("millilitre"),
    "μL":  _u("microlitre"),
    "dL":  _u("decilitre"),
    "cL":  _u("centilitre"),
    "kL":  _u("kilolitre"),
    "gal": _u("gallon"),
    "pt":  _u("pint"),
    "qt":  _u("quart"),
    "fl oz": _u("fluid ounce"),
    "tsp": _u("teaspoon"),
    "tbsp":_u("tablespoon"),
    "cup": _u("cup"),
    "ms":  _u("millisecond"),
    "μs":  _u("microsecond"),
    "ns":  _u("nanosecond"),
    "ps":  _u("picosecond"),
    "min": _u("minute"),
    "h":   _u("hour"),
    "hr":  _u("hour"),
    "wk":  _u("week"),
    "mo":  _u("month"),
    "yr":  _u("year"),
    "Hz":  _u("hertz", invariant=True),
    "kHz": _u("kilohertz", invariant=True),
    "MHz": _u("megahertz", invariant=True),
    "GHz": _u("gigahertz", invariant=True),
    "THz": _u("terahertz", invariant=True),
    "N":   _u("newton"),         # matched only via digit+unit pattern
    "kN":  _u("kilonewton"),
    "MN":  _u("meganewton"),
    "kgf": _u("kilogram-force"),
    "lbf": _u("pound-force"),
    "dyn": _u("dyne"),
    "Pa":  _u("pascal", invariant=True),
    "kPa": _u("kilopascal", invariant=True),
    "MPa": _u("megapascal", invariant=True),
    "GPa": _u("gigapascal", invariant=True),
    "hPa": _u("hectopascal", invariant=True),
    "bar": _u("bar", invariant=True),
    "mbar":_u("millibar", invariant=True),
    "atm": _u("atmosphere"),
    "psi": _u("pounds per square inch", invariant=True),
    "mmHg":_u("millimetre of mercury", invariant=True),
    "inHg":_u("inch of mercury", invariant=True),
    "J":   _u("joule"),          # matched only via digit+unit pattern
    "kJ":  _u("kilojoule"),
    "MJ":  _u("megajoule"),
    "GJ":  _u("gigajoule"),
    "cal": _u("calorie"),
    "kcal":_u("kilocalorie"),
    "eV":  _u("electronvolt"),
    "keV": _u("kiloelectronvolt"),
    "MeV": _u("megaelectronvolt"),
    "GeV": _u("gigaelectronvolt"),
    "Btu": _u("British thermal unit"),
    "W":   _u("watt"),           # matched only via digit+unit pattern
    "kW":  _u("kilowatt"),
    "MW":  _u("megawatt"),
    "GW":  _u("gigawatt"),
    "mW":  _u("milliwatt"),
    "μW":  _u("microwatt"),
    "hp":  _u("horsepower"),
    "bhp": _u("brake horsepower"),
    "PS":  _u("metric horsepower"),
    "V":   _u("volt"),  # only matched via digit+unit pattern
    "kV":  _u("kilovolt"),
    "mV":  _u("millivolt"),
    "μV":  _u("microvolt"),
    "mA":  _u("milliampere"),
    "μA":  _u("microampere"),
    "kA":  _u("kiloampere"),
    "Ω":   _u("ohm", invariant=True),
    "kΩ":  _u("kilohm", invariant=True),
    "MΩ":  _u("megohm", invariant=True),
    "mΩ":  _u("milliohm", invariant=True),
    "mS":  _u("millisiemens", invariant=True),
    "mF":  _u("millifarad", invariant=True),
    "μF":  _u("microfarad", invariant=True),
    "nF":  _u("nanofarad", invariant=True),
    "pF":  _u("picofarad", invariant=True),
    "mH":  _u("millihenry", invariant=True),
    "μH":  _u("microhenry", invariant=True),
    "Wb":  _u("weber", invariant=True),
    "mT":  _u("millitesla", invariant=True),
    "μT":  _u("microtesla", invariant=True),
    "Gs":  _u("gauss"),
    "mC":  _u("millicoulomb"),
    "μC":  _u("microcoulomb"),
    "Ah":  _u("ampere hour"),
    "mAh": _u("milliampere hour"),
    "kWh": _u("kilowatt hour"),
    "MWh": _u("megawatt hour"),
    "GWh": _u("gigawatt hour"),
    "lx":  _u("lux", invariant=True),
    "lm":  _u("lumen"),
    "cd":  _u("candela"),
    "Bq":  _u("becquerel", invariant=True),
    "kBq": _u("kilobecquerel", invariant=True),
    "MBq": _u("megabecquerel", invariant=True),
    "GBq": _u("gigabecquerel", invariant=True),
    "Gy":  _u("gray", invariant=True),
    "Sv":  _u("sievert", invariant=True),
    "mSv": _u("millisievert", invariant=True),
    "μSv": _u("microsievert", invariant=True),
    "Ci":  _u("curie"),
    "mCi": _u("millicurie"),
    "B":   _u("byte"),
    "KB":  _u("kilobyte"),
    "MB":  _u("megabyte"),
    "GB":  _u("gigabyte"),
    "TB":  _u("terabyte"),
    "PB":  _u("petabyte"),
    "EB":  _u("exabyte"),
    "Kb":  _u("kilobit"),
    "Mb":  _u("megabit"),
    "Gb":  _u("gigabit"),
    "Tb":  _u("terabit"),
    "kbps":_u("kilobit per second", invariant=True),
    "Tbps":_u("terabit per second", invariant=True),
    "KBps":_u("kilobyte per second", invariant=True),
    "MBps":_u("megabyte per second", invariant=True),
    "Gt":  _u("gigatonne"),
    "Mt":  _u("megatonne"),
    "m3/s":  _u("cubic metre per second", invariant=True),
    "m³/s":  _u("cubic metre per second", invariant=True),
    "m3/h":  _u("cubic metre per hour", invariant=True),
    "m³/h":  _u("cubic metre per hour", invariant=True),
    "m3/hr": _u("cubic metre per hour", invariant=True),
    "m³/hr": _u("cubic metre per hour", invariant=True),
    "m3/min":_u("cubic metre per minute", invariant=True),
    "m³/min":_u("cubic metre per minute", invariant=True),
    "cm3/s": _u("cubic centimetre per second", invariant=True),
    "cm³/s": _u("cubic centimetre per second", invariant=True),
    "ft3/s": _u("cubic foot per second", invariant=True),
    "ft³/s": _u("cubic foot per second", invariant=True),
    "L/min":_u("litre per minute", invariant=True),
    "L/hr": _u("litre per hour", invariant=True),
    "mL/s": _u("millilitre per second", invariant=True),
    "mL/min":_u("millilitre per minute", invariant=True),
    "Mbps":_u("megabit per second", invariant=True),
    "Gbps":_u("gigabit per second", invariant=True),
    "L/s":  _u("litre per second", invariant=True),
    "km/h":  _u("kilometre per hour", invariant=True),
    "km/L":  _u("kilometre per litre", invariant=True),
    "km/l":  _u("kilometre per litre", invariant=True),
    "L/100km": _u("litre per hundred kilometres", invariant=True),
    "L/1000km": _u("litre per thousand kilometres", invariant=True),
    "km/hr": _u("kilometre per hour", invariant=True),
    "m/s":   _u("metre per second", invariant=True),
    "mph":   _u("mile per hour", invariant=True),
    "knot":  _u("knot"),
    "kt":    _u("knot"),
    "kn":    _u("knot"),
    "fathom":_u("fathom"),
    "ftm":   _u("fathom"),
    "carat": _u("carat"),
    "ct":    _u("carat"),
    "amu":   _u("atomic mass unit"),
    "u":     _u("atomic mass unit"),
    "Da":    _u("dalton"),
    "Torr":  _u("torr", invariant=True),
    "mmol":  _u("millimole"),
    "μmol":  _u("micromole"),
    "ppm":   _u("part per million", invariant=True),
    "ppb":   _u("part per billion", invariant=True),
    "ppt":   _u("part per trillion", invariant=True),
    "mol/L": _u("mole per litre", invariant=True),
    "mmol/L":_u("millimole per litre", invariant=True),
    "μmol/L":_u("micromole per litre", invariant=True),
    "mg/L":  _u("milligram per litre", invariant=True),
    "μg/L":  _u("microgram per litre", invariant=True),
    "g/L":   _u("gram per litre", invariant=True),
    "mg/dL": _u("milligram per decilitre", invariant=True),
    "μg/mL": _u("microgram per millilitre", invariant=True),
    "μg/m³": _u("microgram per cubic metre", invariant=True),
    "µg/m³": _u("microgram per cubic metre", invariant=True),
    "ng/mL": _u("nanogram per millilitre", invariant=True),
    "g/mol": _u("gram per mole", invariant=True),
    "g/cm³": _u("gram per cubic centimetre", invariant=True),
    "g/cm3": _u("gram per cubic centimetre", invariant=True),
    "kg/m³": _u("kilogram per cubic metre", invariant=True),
    "kg/m3": _u("kilogram per cubic metre", invariant=True),
    "g/ml": _u("gram per millilitre", invariant=True),
    "rad":  _u("radian", invariant=True),
    "mrad": _u("milliradian", invariant=True),
    "sr":   _u("steradian", invariant=True),
    "°":    _u("degree"),   # bare degree handled by temperature logic
    "′":    _u("arcminute"),
    "″":    _u("arcsecond"),
    "rpm":  _u("revolution per minute", invariant=True),
    "RPM":  _u("revolution per minute", invariant=True),
    "bpm":  _u("beat per minute", invariant=True),
    "beats/minute": _u("beat per minute", invariant=True),
    "beats/min":    _u("beat per minute", invariant=True),
    "steps/min":    _u("step per minute", invariant=True),
    "cal/min":      _u("calorie per minute", invariant=True),
    "kcal/min":     _u("kilocalorie per minute", invariant=True),
    "words/min":    _u("word per minute", invariant=True),
    "wpm":          _u("word per minute", invariant=True),

    "BPM":  _u("beat per minute", invariant=True),
    "rps":  _u("revolution per second", invariant=True),
    "m²":  _u("square metre", invariant=True),
    "km²": _u("square kilometre", invariant=True),
    "cm²": _u("square centimetre", invariant=True),
    "mm²": _u("square millimetre", invariant=True),
    "ha":  _u("hectare"),
    "ac":  _u("acre"),
    "ft²": _u("square foot", "square feet"),
    "in²": _u("square inch", "square inches"),
    "mi²": _u("square mile"),
    "m³":  _u("cubic metre", invariant=True),
    "km³": _u("cubic kilometre", invariant=True),
    "cm³": _u("cubic centimetre", invariant=True),
    "mm³": _u("cubic millimetre", invariant=True),
    "ft³": _u("cubic foot", "cubic feet"),
    "in³": _u("cubic inch", "cubic inches"),
    "cc":  _u("cubic centimetre", invariant=True),
    "Pa·s":_u("pascal second", invariant=True),
    "cP":  _u("centipoise"),
    "St":  _u("stokes"),
    "cSt": _u("centistokes"),
    "N·m": _u("newton metre", invariant=True),
    "J/K": _u("joule per kelvin", invariant=True),
    "W/m²":_u("watt per square metre", invariant=True),
    "dB":  _u("decibel"),
    "MP":  _u("megapixel"),
    "Mp":  _u("megapixel"),
    "px":  _u("pixel"),
    "dpi": _u("dot per inch", invariant=True),
    "ppi": _u("pixel per inch", invariant=True),
    "dBA": _u("A-weighted decibel"),
    "Np":  _u("neper"),
    "kat": _u("katal", invariant=True),
    "IU":  _u("international unit"),
    "mM":  _u("millimolar", invariant=True),
    "μM":  _u("micromolar", invariant=True),
    "nM":  _u("nanomolar",  invariant=True),
    "wt%": _u("weight percent", invariant=True),
    "vol%":_u("volume percent", invariant=True),
    "pKa": _u("pKa", invariant=True),
}

_UNIT_KEYS_SORTED = sorted(UNIT_EXACT.keys(), key=len, reverse=True)

_DOT_TIME_UNIT_STARTERS = set(UNIT_EXACT.keys()) | {
    "metres", "meter", "meters", "metre", "kilograms", "kilogram",
    "grams", "gram", "litres", "litre", "liters", "liter",
    "kilometres", "kilometre", "kilometers", "kilometer",
    "centimetres", "centimetre", "millimetres", "millimetre",
    "seconds", "second", "minutes", "minute",
    "carats", "carat", "knots", "knot", "fathoms", "fathom",
    "percent", "tonnes", "tonne", "pounds", "pound",
    "kilopascals", "kilopascal", "watts", "kilowatts",
    "volts", "amperes", "ampere", "ohms",
    "megapixels", "megapixel", "pixels", "pixel",
    "gigabytes", "megabytes", "kilobytes",
    "runs", "points", "ha", "hectares",
    "MP", "Mbps", "Gbps", "kbps",
}


TEMP_UNITS = {
    "C": "celsius",
    "F": "fahrenheit",
    "K": "kelvin",
}

def _get_unit(tok):
    """Return (singular, plural) for a unit token, or None."""
    return UNIT_EXACT.get(tok)

def _pluralise(singular, plural, count_word, num_str=None):
    """
    Return singular or plural form.
    Rule: numeric value exactly 1 → singular; everything else (0, 0.x, 2+) → plural.
    num_str: original numeric string for exact float comparison.
    count_word: fallback when num_str unavailable.
    """
    if plural == singular:   # invariant (hertz, pascal etc.)
        return singular
    if num_str is not None:
        try:
            val = float(num_str.replace(",", ""))
            return singular if val == 1.0 else plural
        except ValueError:
            pass
    if count_word in ("one", "minus one"):
        return singular
    return plural


def apply_temperature(text):
    """
    Convert temperature expressions before main unit pass.
    Patterns:
      -12°C    → minus twelve degrees celsius
      25°C     → twenty five degrees celsius
      98.6°F   → ninety eight point six degrees fahrenheit
      300K     → three hundred kelvin
      °C       → degrees celsius (bare, no number)
    """
    changes = []

    def _temp_repl(m):
        num_str  = m.group(1)   # e.g. "-12" or "25" or None
        deg_sym  = m.group(2)   # "°" or None
        unit_str = m.group(3)   # "C", "F", "K"

        unit_word = TEMP_UNITS.get(unit_str.upper(), unit_str)
        is_kelvin = unit_str.upper() == "K"

        if num_str:
            num_w, _ = number_token_to_words(num_str, "quantity")
            if is_kelvin:
                result = f"{num_w} {unit_word}"
            else:
                result = f"{num_w} degrees {unit_word}"
        else:
            if is_kelvin:
                result = unit_word
            else:
                result = f"degrees {unit_word}"

        changes.append((m.group(0).strip(), result))
        return " " + result + " "

    pattern = (
        r'(-?\d+(?:\.\d+)?)?'   # optional number (captures minus too)
        r'\s*(°)'               # degree symbol (required for C/F)
        r'([CFK])\b'            # unit letter
    )
    text = re.sub(pattern, _temp_repl, text)

    def _kelvin_repl(m):
        num_str = m.group(1)
        num_w, _ = number_token_to_words(num_str, "quantity")
        result = f"{num_w} kelvin"
        changes.append((m.group(0).strip(), result))
        return " " + result + " "

    text = re.sub(r'(-?\d+(?:\.\d+)?)\s*K\b', _kelvin_repl, text)
    text = re.sub(r' {2,}', ' ', text).strip()
    return text, changes


def apply_dimensions(text):
    """
    Convert dimension separators: 12.45 m x 3 m → twelve point four five metres by three metres.
    The "x" must be surrounded by spaces and between numbers/units.
    """
    changes = []
    def _dim_repl(m):
        changes.append((m.group(0), m.group(0).replace(" x ", " by ").replace(" X ", " by ")))
        return m.group(0).replace(" x ", " by ").replace(" X ", " by ")

    text = re.sub(
        r'(?<=\w)\s+[xX]\s+(?=\d)',
        lambda m: " by ",
        text
    )
    if changes:
        pass  # changes tracked implicitly; log below
    return text, []   # changes logged in caller via report


def apply_dimensions_tracked(text):
    """Same as apply_dimensions but returns proper change list."""
    changes = []
    def _repl(m):
        result = " by "
        changes.append((m.group(0).strip(), "by"))
        return result
    text = re.sub(r'(?<=\w)\s+[xX]\s+(?=\d)', _repl, text)

    def _mag_repl(m):
        changes.append((m.group(0), m.group(1) + " x"))
        return m.group(1) + " x"
    text = re.sub(r'(?<![A-Za-z])(\d+(?:\.\d+)?)[xX](?![A-Za-z0-9])', _mag_repl, text)
    return text, changes
def apply_rate_qualifiers(text):
    """
    Expand abbreviated rate qualifiers before punctuation stripping.
    p.a. → per annum
    p.m. → per month
    p.w. → per week
    p.d. → per day
    p.h. → per hour
    Also handles: pa, pm, pw, pd, ph when adjacent to % or rate context.
    """
    changes = []
    RATE_MAP = {
        r'\bp\.a\.\b':  "per annum",
        r'\bp\.m\.\b':  "per month",
        r'\bp\.w\.\b':  "per week",
        r'\bp\.d\.\b':  "per day",
        r'\bp\.h\.\b':  "per hour",
        r'\bp\.a':        "per annum",   # without trailing dot
        r'(?<=[%\d])\s+p\.a\b': " per annum",
        r'(?<=[%\d])\s+p\.m\b': " per month",
        r'(?<=[%\d])\s+p\.w\b': " per week",
        r'(?<=[%\d])\s+p\.h\b': " per hour",
    }
    for pattern, replacement in RATE_MAP.items():
        def _repl(m, r=replacement):
            changes.append((m.group(0).strip(), r.strip()))
            return " " + r + " "
        text = re.sub(pattern, _repl, text, flags=re.IGNORECASE)
    text = re.sub(r' {2,}', ' ', text).strip()
    return text, changes
def apply_fractions(text):
    """
    Convert fraction patterns to spoken form.
    1/100    → one by one hundred
    120/80.5 → one hundred and twenty by eighty point five
    1.5/2    → one point five by two
    But NOT: km/h (unit compounds — letters present).
    """
    changes = []
    def _frac_repl(m):
        num = m.group(1)
        den = m.group(2)
        num_w, _ = number_token_to_words(num, "quantity")
        den_w, _ = number_token_to_words(den, "quantity")
        result = f"{num_w} by {den_w}"
        changes.append((m.group(0), result))
        return result
    text = re.sub(
        r'(?<![A-Za-z\d])(\d+(?:\.\d+)?)\s*/\s*(\d+(?:\.\d+)?)(?![A-Za-z\d])',
        _frac_repl, text)
    return text, changes

def apply_ranges(text):
    """Convert number ranges like 12-15 → twelve to fifteen."""
    changes = []
    def _range_repl(m):
        lo, hi = m.group(1), m.group(2)
        lo_w = integer_to_words(int(lo))
        hi_w = integer_to_words(int(hi))
        result = f"{lo_w} to {hi_w}"
        changes.append((m.group(0), result))
        return result
    text = re.sub(r'(?<![\d\w])(\d{1,4})-(\d{1,4})(?![\d\w])', _range_repl, text)
    return text, changes

def apply_ratios(text):
    """
    Convert ratio expressions (N:N or N:N.N) to spoken form.
      1:16.789   → one is to sixteen point seven eight nine
      3:4        → three is to four
    Does NOT fire on valid HH:MM time patterns (LHS 0-23, RHS exactly
    two digits 00-59).
    Trailing-zero rule applies to decimal sides.
    """
    changes = []

    def _ratio_repl(m):
        lhs = m.group(1)
        rhs = m.group(2)
        if (re.fullmatch(r'\d{1,2}', lhs) and re.fullmatch(r'\d{2}', rhs)
                and int(lhs) <= 23 and int(rhs) <= 59):
            return m.group(0)
        lhs_w, _ = number_token_to_words(lhs, "quantity")
        rhs_w, _ = number_token_to_words(rhs, "quantity")
        result = f"{lhs_w} is to {rhs_w}"
        changes.append((m.group(0), result))
        return result

    text = re.sub(
        r'(?<!\d)(\d+(?:\.\d+)?):(\d+(?:\.\d+)?)(?!\d)',
        _ratio_repl, text)
    return text, changes

def apply_complex_rates(text):
    """
    Handle rate expressions where a number appears in the denominator.
    e.g. L/1000 km → litre per thousand kilometres
         mg/100 mL → milligram per hundred millilitres
         kcal/100 g → kilocalorie per hundred grams
    """
    changes = []

    def _complex_rate_repl(m):
        unit_top = m.group(1)
        denom_num = m.group(2)
        unit_bot = m.group(3)
        top_entry = UNIT_EXACT.get(unit_top)
        bot_entry = UNIT_EXACT.get(unit_bot)
        if top_entry is None or bot_entry is None:
            return m.group(0)
        top_w = top_entry[0]   # singular
        bot_w = bot_entry[0]
        num_w = integer_to_words(int(denom_num))
        result = f"{top_w} per {num_w} {bot_w}s"
        changes.append((m.group(0).strip(), result))
        return " " + result + " "

    unit_pat = '|'.join(re.escape(k) for k in _UNIT_KEYS_SORTED if len(k) >= 1)
    pattern = (
        r'('+ unit_pat + r')'
        r'\s*/\s*'
        r'(\d+)'
        r'\s+'
        r'('+ unit_pat + r')'
        r'(?![A-Za-z])'
    )
    text = re.sub(pattern, _complex_rate_repl, text)
    text = re.sub(r' {2,}', ' ', text).strip()
    return text, changes
def normalise_ascii_powers(text):
    """
    Convert ASCII power notation to unicode superscripts before unit pass.
    m3 → m³,  m2 → m²,  cm3 → cm³,  km2 → km²
    Only fires when digit 2 or 3 immediately follows letters (unit pattern).
    Does NOT fire on standalone numbers or number+unit like "500 m3" handled elsewhere.
    """
    text = re.sub(r'(?<=[A-Za-zμ])([23])(?![\d])',
                  lambda m: "²" if m.group(1) == "2" else "³",
                  text)
    return text
def apply_units(text, locale="US"):
    """
    Convert unit expressions. Strict word-boundary matching to avoid
    matching unit abbreviations inside ordinary words.

    Order:
      0. Superscript/power units WITH a denominator (m/s² → m/s + square → "per square second"
         handled specially) and plain superscripts (km², m³) — run FIRST so compound
         Pass 1 doesn't consume the base unit before the superscript is seen.
      1. Pre-built compound keys (km/h, mAh, kWh etc.) longest first
      2. number immediately before unit token (50kg, 100mL)
    """
    changes = []
    _indian = _detect_indian_numbering(text, 0, len(text), locale)

    def _ntw(num_str, hint="quantity"):
        """Wrapper: calls number_token_to_words with correct indian flag."""
        return number_token_to_words(num_str, hint, indian=_indian)

    PRE  = r'(?<![A-Za-z])'   # not preceded by a letter
    POST = r'(?![A-Za-z])'    # not followed by a letter

    def _pluralise_word(word, num_str):
        """Pluralise a spoken unit phrase (possibly containing 'per') based on value."""
        try:
            val = float(num_str.replace(",", "")) if num_str else None
        except ValueError:
            val = None
        if val is None or val == 1.0:
            return word
        if " per " in word:
            head, tail = word.split(" per ", 1)
            base_head = head.rstrip("s") if head.endswith("s") else head
            if base_head.lower() not in INVARIANT and not head.endswith("s"):
                head = head + "s"
            return head + " per " + tail
        base = word.rstrip("s") if word.endswith("s") else word
        if base.lower() in INVARIANT:
            return word
        if not word.endswith("s"):
            return word + "s"
        return word

    _sup_compound_keys = frozenset(
        k for k in UNIT_EXACT if "/" in k and any(c in k for c in "²³"))

    def _rate_power_repl(m):
        num_str  = m.group(1)
        unit_top = m.group(2)
        unit_bot = m.group(3)
        sup_char = m.group(4)
        full_key = f"{unit_top}/{unit_bot}{sup_char}"
        if full_key in _sup_compound_keys:
            return m.group(0)
        top_entry = UNIT_EXACT.get(unit_top)
        bot_entry = UNIT_EXACT.get(unit_bot)
        if top_entry is None or bot_entry is None:
            return m.group(0)
        modifier = "square" if sup_char in ("²", "\u00b2") else "cubic"
        top_sing = top_entry[0]
        bot_sing = bot_entry[0]
        base = f"{top_sing} per {modifier} {bot_sing}"
        if num_str:
            num_w, _ = _ntw(num_str)
            unit_w = _pluralise_word(base, num_str)
            result = f"{num_w} {unit_w}"
        else:
            result = base
        changes.append((m.group(0).strip(), result))
        return " " + result + " "

    text = re.sub(
        r'(-?\d+(?:\.\d+)?)?\s*([A-Za-zμµ]+)/([A-Za-zμµ]+)([²³])',
        _rate_power_repl, text)
    text = re.sub(r' {2,}', ' ', text).strip()

    def _compound_sup_repl(m):
        num_str = m.group(1)
        key     = m.group(2)
        entry   = UNIT_EXACT.get(key)
        if entry is None:
            return m.group(0)
        sing, plur = entry
        if num_str:
            num_w, _ = _ntw(num_str)
            unit_w   = _pluralise_word(sing, num_str)
            result   = f"{num_w} {unit_w}"
        else:
            result = sing
        changes.append((m.group(0).strip(), result))
        return " " + result + " "

    _sup_slash_keys = sorted(
        (k for k in UNIT_EXACT if ("/" in k) and any(c in k for c in "²³")),
        key=len, reverse=True)
    if _sup_slash_keys:
        _sup_slash_alt = '|'.join(re.escape(k) for k in _sup_slash_keys)
        text = re.sub(
            r'(-?\d+(?:\.\d+)?)?\s*(' + _sup_slash_alt + r')(?![A-Za-zµμ\d])',
            _compound_sup_repl, text)
    text = re.sub(r' {2,}', ' ', text).strip()

    def _sup_repl(m):
        num_str   = m.group(1)
        unit_part = m.group(2)
        sup_char  = m.group(3)
        modifier = "square" if sup_char in ("²", "\u00b2") else "cubic"
        entry = UNIT_EXACT.get(unit_part)
        if entry is None:
            return m.group(0)
        sing, plur = entry
        base = f"{modifier} {sing}"
        if num_str:
            num_w, _ = _ntw(num_str)
            unit_w = _pluralise_word(base, num_str)
            result = f"{num_w} {unit_w}"
        else:
            result = base
        changes.append((m.group(0).strip(), result))
        return " " + result + " "

    text = re.sub(r'(-?\d+(?:\.\d+)?)?\s*([A-Za-zμµ]+)([²³])', _sup_repl, text)
    text = re.sub(r' {2,}', ' ', text).strip()

    _BARE_UNIT_WORDS = {
        "metre","metres","meter","meters","yard","yards","foot","feet",
        "mile","miles","kilometre","kilometres","kilometer","kilometers",
        "litre","litres","liter","liters","barrel","barrels","gallon","gallons",
        "breath","breaths","beat","beats","step","steps","word","words",
        "calorie","calories","cycle","cycles","rotation","rotations",
        "revolution","revolutions","pulse","pulses","packet","packets",
        "minute","min","second","sec","hour","hr","day","week","month","year",
        "person","people","unit","units","item","items","count","counts",
    }
    def _slash_per_repl(m):
        num_str  = m.group(1)
        left     = m.group(2)
        right    = m.group(3)
        l_lower  = left.rstrip("s").lower() if left.lower() not in _BARE_UNIT_WORDS else left.lower()
        r_lower  = right.rstrip("s").lower()
        if left.lower() not in _BARE_UNIT_WORDS and r_lower not in _BARE_UNIT_WORDS:
            return m.group(0)
        _time_expand = {"min":"minute","sec":"second","hr":"hour"}
        right_w = _time_expand.get(right.lower(), right)
        result = (f"{num_str} {left} per {right_w}" if num_str
                  else f"{left} per {right_w}")
        changes.append((m.group(0).strip(), result))
        return " " + result + " "

    text = re.sub(
        r'(?<![/\d])(-?\d+(?:\.\d+)?)?\s*([A-Za-z]+)/([A-Za-z]+)(?![²³/\d])',
        _slash_per_repl, text)
    text = re.sub(r' {2,}', ' ', text).strip()


    for key in _UNIT_KEYS_SORTED:
        if len(key) < 2:
            continue  # single-char bare units handled in pass 2
        entry = UNIT_EXACT[key]
        escaped = re.escape(key)
        pat = PRE + r'(-?\d+(?:\.\d+)?)?\s*' + escaped + POST

        def _repl(m, e=entry, k=key):
            num_str = m.group(1)
            sing, plur = e
            if num_str:
                num_w, _ = _ntw(num_str)
                if sing == plur:
                    unit_w = _pluralise_word(sing, num_str)
                else:
                    unit_w = _pluralise(sing, plur, num_w, num_str)
                result = f"{num_w} {unit_w}"
            else:
                result = sing
            changes.append((m.group(0).strip(), result))
            return " " + result + " "

        text = re.sub(pat, _repl, text)

    text = re.sub(r' {2,}', ' ', text).strip()

    DECADE_PAT = re.compile(r'^\d{1,3}0$')  # matches 10,20,...,80,90,100,...,1900 etc.

    for key in _UNIT_KEYS_SORTED:
        entry = UNIT_EXACT[key]
        escaped = re.escape(key)
        pat = r'(\d+(?:\.\d+)?)\s*' + escaped + POST

        def _num_unit_repl(m, e=entry, k=key):
            num_str = m.group(1)
            if (len(k) == 1 and k.lower() == "s"
                    and "." not in num_str
                    and DECADE_PAT.match(num_str)):
                return m.group(0)  # leave untouched — handled by number engine
            sing, plur = e
            num_w, _ = _ntw(num_str)
            if sing == plur:
                unit_w = _pluralise_word(sing, num_str)
            else:
                unit_w = _pluralise(sing, plur, num_w, num_str)
            result = f"{num_w} {unit_w}"
            changes.append((m.group(0).strip(), result))
            return " " + result + " "

        text = re.sub(pat, _num_unit_repl, text)

    text = re.sub(r' {2,}', ' ', text).strip()
    return text, changes


CURRENCY_CODE_MAP = {
    "INR": ("I N R", "Indian Rupees"),
    "USD": ("U S D", "US Dollars"),
    "GBP": ("G B P", "British Pounds"),
    "EUR": ("E U R", "Euros"),
    "JPY": ("J P Y", "Japanese Yen"),
    "CNY": ("C N Y", "Chinese Yuan"),
    "AUD": ("A U D", "Australian Dollars"),
    "CAD": ("C A D", "Canadian Dollars"),
    "CHF": ("C H F", "Swiss Francs"),
    "NZD": ("N Z D", "New Zealand Dollars"),
    "SGD": ("S G D", "Singapore Dollars"),
    "HKD": ("H K D", "Hong Kong Dollars"),
    "ZAR": ("Z A R", "South African Rand"),
    "AED": ("A E D", "UAE Dirhams"),
    "SAR": ("S A R", "Saudi Riyals"),
    "KRW": ("K R W", "South Korean Won"),
    "RUB": ("R U B", "Russian Roubles"),
    "BRL": ("B R L", "Brazilian Real"),
    "MXN": ("M X N", "Mexican Pesos"),
    "SEK": ("S E K", "Swedish Krona"),
    "NOK": ("N O K", "Norwegian Krone"),
    "DKK": ("D K K", "Danish Krone"),
    "THB": ("T H B", "Thai Baht"),
    "PHP": ("P H P", "Philippine Pesos"),
    "IDR": ("I D R", "Indonesian Rupiah"),
    "MYR": ("M Y R", "Malaysian Ringgit"),
    "VND": ("V N D", "Vietnamese Dong"),
    "PKR": ("P K R", "Pakistani Rupees"),
    "BDT": ("B D T", "Bangladeshi Taka"),
    "LKR": ("L K R", "Sri Lankan Rupees"),
    "NPR": ("N P R", "Nepalese Rupees"),
}

def apply_currency_codes(text, expand=False):
    """
    Convert 3-letter currency codes following a number.
    "500 INR"  → "500 Indian Rupees"   (expand=True)
    "500 INR"  → "500 I N R"           (expand=False)
    Only fires when the code immediately follows a number (with optional space).
    """
    changes = []
    code_alt = '|'.join(CURRENCY_CODE_MAP.keys())
    pattern = r'(?<=\d)\s*\b(' + code_alt + r')\b'

    def _repl(m):
        code = m.group(1).upper()
        letter_form, expanded_form = CURRENCY_CODE_MAP[code]
        replacement = expanded_form if expand else letter_form
        changes.append((m.group(0).strip(), replacement))
        return " " + replacement

    text = re.sub(pattern, _repl, text)
    return text, changes


_SCALE_SUFFIX = r'(?:\s*(million|billion|trillion|lakh|crore|arab))?'

_CURRENCY_ROWS = [
    (r'\$\s*(-?\d[\d,]*(?:\.\d+)?)' + _SCALE_SUFFIX, "dollars", False),
    (r'£\s*(-?\d[\d,]*(?:\.\d+)?)' + _SCALE_SUFFIX,  "pounds",  False),
    (r'€\s*(-?\d[\d,]*(?:\.\d+)?)' + _SCALE_SUFFIX,  "euros",   False),
    (r'¥\s*(-?\d[\d,]*(?:\.\d+)?)' + _SCALE_SUFFIX,  "yen",     False),
    (r'₹\s*(-?\d[\d,]*(?:\.\d+)?)' + _SCALE_SUFFIX,  "rupees",  True),
]

def _make_currency_repl(currency_word, indian_eligible, indian_flag):
    def _repl(m):
        num_str   = m.group(1).replace(",", "")
        scale     = m.group(2)  # e.g. "million", "lakh", or None
        use_indian = indian_flag and indian_eligible
        word, _ = number_token_to_words(num_str, "quantity", indian=use_indian)
        if scale:
            return word + " " + scale + " " + currency_word
        return word + " " + currency_word
    return _repl

SYMBOL_MAP = [
    (r'#([A-Za-z][A-Za-z0-9]*)',     lambda m: "hashtag " + re.sub(r'([a-z])([A-Z])', r'\1 \2', m.group(1))),
    (r'@([A-Za-z][A-Za-z0-9_]*)',    lambda m: "at " + m.group(1)),
    (r'(-?\d+(?:\.\d+)?)\s*%',       lambda m: number_token_to_words(m.group(1), "quantity")[0] + " percent"),
    (r'(?<!\w)%(?!\w)',               " percent "),
    (r'\s*&\s*', " and "),
    (r'\+',  " plus "),
    (r'=',   " equals "),
    (r'~',   " approximately "),
    (r'<',   " less than "),
    (r'>',   " greater than "),
    (r'\\',  " backslash "),
    (r'\|',  " pipe "),
    (r'\^',  " caret "),
    (r'Ω',   " ohm "),
    (r'©',   " copyright "),
    (r'®',   " registered "),
    (r'™',   " trademark "),
    (r'§',   " section "),
    (r'¶',   " paragraph "),
    (r'µ',   " micro "),   # loose μ not caught by unit table
    (r'±',   " plus or minus "),
    (r'≠',   " not equal to "),
    (r'≤',   " less than or equal to "),
    (r'≥',   " greater than or equal to "),
    (r'²',   ""),   # any remaining superscripts after unit pass
    (r'³',   ""),
    (r'°',   " degrees "),  # any remaining bare degree signs
    (r'…',   "..."),
    (r'–',   "-"),
    (r'—',   "-"),
]


COMMON_WORDS = {
    "a","an","the","and","or","but","if","in","on","at","to","for","of","up",
    "as","by","be","is","it","its","my","me","we","us","he","she","they","them",
    "his","her","our","you","your","who","not","no","so","do","go","am","are",
    "was","were","has","had","have","can","may","will","shall","did","get","got",
    "let","put","set","try","use","see","say","said","come","came","make","made",
    "take","took","give","gave","know","knew","think","thought","tell","told",
    "find","found","call","keep","left","mean","hold","turn","ask","need","feel",
    "seem","look","show","back","just","also","than","then","when","where","how",
    "all","any","some","few","more","most","both","each","with","from","into",
    "out","off","over","under","again","after","before","between","through",
    "during","while","about","against","among","around","without","within",
    "long","down","much","too","very","well","still","even","never","always",
    "often","now","here","there","this","that","these","those","what","which",
    "new","old","big","small","high","low","right","left","good","bad","next",
    "last","own","other","same","real","sure","clear","free","full","open","past",
    "able","like","great","little","large","early","first","second","third",
}



SINGULAR_PLURAL_UNIT_WORDS = {
    "seconds":"second","minutes":"minute","hours":"hour","days":"day",
    "weeks":"week","months":"month","years":"years","metres":"metre",
    "meters":"meter","kilometres":"kilometre","kilometers":"kilometer",
    "centimetres":"centimetre","centimeters":"centimeter",
    "millimetres":"millimetre","millimeters":"millimeter",
    "nanometres":"nanometre","grams":"gram","kilograms":"kilogram",
    "milligrams":"milligram","micrograms":"microgram","tonnes":"tonne",
    "litres":"litre","liters":"liter","millilitres":"millilitre",
    "milliliters":"milliliter","watts":"watt","kilowatts":"kilowatt",
    "megawatts":"megawatt","amperes":"ampere","amps":"amp","volts":"volt",
    "ohms":"ohm","newtons":"newton","joules":"joule","kilojoules":"kilojoule",
    "pascals":"pascal","kilopascals":"kilopascal","decibels":"decibel",
    "megapixels":"megapixel","pixels":"pixel","bytes":"byte",
    "kilobytes":"kilobyte","megabytes":"megabyte","gigabytes":"gigabyte",
    "terabytes":"terabyte","fathoms":"fathom","carats":"carat","knots":"knot",
    "pounds":"pound","ounces":"ounce","feet":"foot","inches":"inch",
    "yards":"yard","miles":"mile","gallons":"gallon","pints":"pint",
    "quarts":"quart","cups":"cup","stones":"stone","atmospheres":"atmosphere",
    "millibars":"millibar","calories":"calorie","kilocalories":"kilocalorie",
    "degrees":"degree","radians":"radian","revolutions":"revolution",
    "hertz":"hertz",  # invariant — listed for completeness
}

def apply_source_grammar_fixes(text):
    """
    Correct singular/plural agreement errors in the SOURCE text itself,
    e.g. "0.000001 seconds" → "0.000001 second" (value's absolute number
    is 1 or less than 1 → singular unit word).
    Returns (corrected_text, list_of_corrections) where each correction is
    a dict {original, corrected}.
    """
    corrections = []

    def _repl(m):
        num_str  = m.group(1)
        unit_word = m.group(2)
        try:
            val = float(num_str.replace(",", ""))
        except ValueError:
            return m.group(0)
        singular = SINGULAR_PLURAL_UNIT_WORDS.get(unit_word.lower())
        if singular is None or singular == unit_word.lower():
            return m.group(0)  # invariant or unknown — no change
        if abs(val) == 1.0:
            if unit_word[0].isupper():
                replacement = singular.capitalize()
            else:
                replacement = singular
            corrections.append((m.group(0), f"{num_str} {replacement}"))
            return f"{num_str} {replacement}"
        return m.group(0)

    plural_alt = '|'.join(re.escape(w) for w in
                           sorted(SINGULAR_PLURAL_UNIT_WORDS.keys(), key=len, reverse=True))
    pattern = r'(-?\d+(?:\.\d+)?)\s+(' + plural_alt + r')\b'
    text = re.sub(pattern, _repl, text, flags=re.IGNORECASE)
    return text, corrections


WORD_UNIT_MAP = {
    "gramme": "gram", "grammes": "grams",
    "kilogramme": "kilogram", "kilogrammes": "kilograms",
    "tonne": "tonne", "tonnes": "tonnes",
    "stonme": "stone",
    "litre": "litre", "litres": "litres",
    "millilitre": "millilitre", "millilitres": "millilitres",
    "centilitre": "centilitre",
    "metre": "metre", "metres": "metres",
    "kilometre": "kilometre", "kilometres": "kilometres",
    "centimetre": "centimetre", "centimetres": "centimetres",
    "millimetre": "millimetre", "millimetres": "millimetres",
    "kelvin": "kelvin",
    "celsius": "celsius",
    "fahrenheit": "fahrenheit",
    "kilopascal": "kilopascal", "kilopascals": "kilopascal",
    "kilojoule": "kilojoule", "kilojoules": "kilojoules",
    "kilowatt": "kilowatt", "kilowatts": "kilowatts",
    "megawatt": "megawatt", "megawatts": "megawatts",
    "milligram": "milligram", "milligrams": "milligrams",
    "microgram": "microgram", "micrograms": "micrograms",
    "nanometre": "nanometre", "nanometres": "nanometres",
    "fathom": "fathom", "fathoms": "fathoms",
    "carat": "carat", "carats": "carats",
    "knot": "knot", "knots": "knots",
    "hertz": "hertz",
    "kilohertz": "kilohertz",
    "megahertz": "megahertz",
    "gigahertz": "gigahertz",
    "watt": "watt", "watts": "watts",
    "ampere": "ampere", "amperes": "amperes",
    "volt": "volt", "volts": "volts",
    "ohm": "ohm", "ohms": "ohm",
    "newton": "newton", "newtons": "newtons",
    "joule": "joule", "joules": "joules",
    "pascal": "pascal",
    "hectopascal": "hectopascal",
    "millibar": "millibar", "millibars": "millibars",
    "atmosphere": "atmosphere", "atmospheres": "atmospheres",
    "decibel": "decibel", "decibels": "decibels",
    "megapixel": "megapixel", "megapixels": "megapixels",
    "pixel": "pixel", "pixels": "pixels",
    "byte": "byte", "bytes": "bytes",
    "kilobyte": "kilobyte", "kilobytes": "kilobytes",
    "megabyte": "megabyte", "megabytes": "megabytes",
    "gigabyte": "gigabyte", "gigabytes": "gigabytes",
    "terabyte": "terabyte", "terabytes": "terabytes",
    "second": "second", "seconds": "seconds",
    "minute": "minute", "minutes": "minutes",
    "hour": "hour", "hours": "hours",
    "day": "day", "days": "days",
    "week": "week", "weeks": "weeks",
    "month": "month", "months": "months",
    "year": "year", "years": "years",
}


PRONOUNCEABLE = {
    "NASA","NATO","FEMA","LASER","RADAR","SONAR","SCUBA","AIDS","OPEC","ASAP",
    "AWOL","AWACS","INTERPOL","UNESCO","UNICEF","NAFTA","SENSEX","NIFTY","SEBI",
    "ISRO","DRDO","IELTS","TOEFL","GIF","ZIP","PIN","SIM","ROM","RAM","PDF",
    "HTML","HTTP","HTTPS","XML","JSON","SQL","API","GUI","CPU","GPU","USB",
    "LAN","WAN","WIFI","HDMI","JPEG","MPEG","ASCII","BIOS","CAPTCHA","CODEC",
    "CERN","WHO","UNESCO","IAEA","OPEC","FIFA","UEFA","NASCAR","DARPA",
    "AM","PM",
}

MIXED_CASE_ABBR = {
    "ph":   "P H",
    "mrna": "M R N A",
    "rrna": "R R N A",
    "trna": "T R N A",
    "cdna": "C D N A",
    "pcr":  "P C R",
    "kda":  "K D A",
    "mda":  "M D A",
    "pka":  "P K A",
    "pkb":  "P K B",
    "pkc":  "P K C",
    "pko":  "P K O",
    "nfkb": "N F K B",
    "atp":  "A T P",
    "adp":  "A D P",
    "nadh": "N A D H",
    "nadph":"N A D P H",
    "fad":  "F A D",
    "coq":  "C O Q",
    "ipa":  "I P A",
    "thc":  "T H C",
    "cbd":  "C B D",
    "hiv":  "H I V",
    "aids": "AIDS",    # pronounceable — keep
    "dna":  "D N A",
    "rna":  "R N A",
}

DESIGNATIONS = {
    "DR":"doctor","PROF":"professor","REV":"reverend","FR":"father",
    "SR":"sister","BR":"brother","CAPT":"captain","MAJ":"major",
    "COL":"colonel","GEN":"general","SGT":"sergeant","CPL":"corporal",
    "LT":"lieutenant","ADM":"admiral","CMDR":"commander",
    "MR":"mister","MRS":"missus","MS":"miss","MESSRS":"misters",
    "HON":"honourable","PHD":"P H D",
    "JR":"junior","SNR":"senior","EST":"established","DEPT":"department",
    "GOVT":"government","INTL":"international","NATL":"national",
    "ASSOC":"associate","ASST":"assistant","MGMT":"management",
    "CORP":"corporation","INC":"incorporated","LTD":"limited","PLC":"P L C",
    "AVE":"avenue","BLVD":"boulevard","APT":"apartment","HQ":"H Q",
    "JAN":"january","FEB":"february","MAR":"march","APR":"april",
    "AUG":"august","SEP":"september","SEPT":"september","OCT":"october",
    "NOV":"november","DEC":"december","MON":"monday","TUE":"tuesday",
    "WED":"wednesday","THU":"thursday","FRI":"friday","SAT":"saturday",
    "SUN":"sunday",
    "NO":"number","NOS":"numbers","PG":"page","PGS":"pages","CH":"chapter",
    "VOL":"volume","ED":"edition","FIG":"figure","EQ":"equation",
    "REF":"reference","REFS":"references","EX":"example","EXS":"examples",
    "QTY":"quantity","MAX":"maximum","AVG":"average","APPROX":"approximately",
    "TEMP":"temperature","INFO":"information","INTRO":"introduction",
    "MISC":"miscellaneous","RE":"regarding","SUBJ":"subject",
    "VS":"versus","ETC":"et cetera","EG":"for example","IE":"that is",
    "NB":"note","PS":"postscript","VIZ":"namely","CA":"circa","CF":"compare",
    "QED":"quod erat demonstrandum",
}

SPELLED_OUT = {
    "FBI","CIA","NSA","IRS","BBC","CNN","NBC","CBS","ABC","ITV","MTV",
    "ESPN","HBO","NFL","NBA","NHL","MLB","MLS","FIFA","UEFA","IPL","ICC",
    "BJP","AAP","TV","PC","OK","UK","US","UN","EU","AU",
    "RIP","FYI","BTW","IDK","IMO","IMHO","TBH","NGL",
    "RSVP","ETA","ETD","ATM","POS",
    "HR","PR","AI","ML","NLP","IOT","AR","VR","XR",
    "CEO","CFO","COO","CTO","VP","GM",
    "ID","IP","OS","UI","UX","QA","QC","DB","CD","DVD",
    "OTP","SMS","MMS","CVV","DOB","DOJ","FIR","PIN",
    "ISO","IEC","IEEE","ASTM","ANSI","DIN","BS","JIS",
}

def expand_abbreviation(token, original_token):
    """Returns (expanded, was_expanded, method)."""
    if original_token.lower() in COMMON_WORDS:
        return original_token, False, None

    is_source_allcaps = (original_token == original_token.upper()
                         and len(original_token) >= 2
                         and original_token.isalpha())

    clean = token.replace("'", "")
    upper = clean.upper()
    lower = clean.lower()

    if lower in MIXED_CASE_ABBR and not original_token.isupper():
        return MIXED_CASE_ABBR[lower], True, "mixed_case"

    if upper in PRONOUNCEABLE:
        return upper, False, "pronounceable"
    if upper in DESIGNATIONS:
        return DESIGNATIONS[upper], True, "designation"
    if upper in SPELLED_OUT:
        return " ".join(list(upper)), True, "spelled_out"

    if is_source_allcaps and re.fullmatch(r'[A-Z]{2,6}', upper):
        return " ".join(list(upper)), True, "heuristic_spelled"

    return original_token, False, None


def apply_symbols(text, locale="US"):
    changes = []

    indian_flag = _detect_indian_numbering(text, 0, len(text), locale)
    currency_map = [
        (pattern, _make_currency_repl(word, eligible, indian_flag))
        for pattern, word, eligible in _CURRENCY_ROWS
    ]

    for pattern, repl in currency_map + SYMBOL_MAP:
        def _repl(m, r=repl):
            result = r(m) if callable(r) else r
            if result.strip():
                changes.append((m.group(0), result.strip()))
            return result
        text = re.sub(pattern, _repl, text)
    return text, changes


YEAR_CONTEXT_WORDS = {
    "in","since","during","until","till","before","after","by","from",
    "year","years","yr","yrs","century","centuries","decade","decades",
    "era","ad","bc","ce","bce","ago","anniversary",
    "born","died","founded","established","built","published","released",
    "launched","opened","closed","began","ended","started","finished",
    "copyright","circa","ca","dated","return","back",
    "history","historically","timeline","throwback","retro","vintage",
    "january","february","march","april","may","june","july","august",
    "september","october","november","december",
    "jan","feb","mar","apr","jun","jul","aug","sep","sept","oct","nov","dec",
    "monday","tuesday","wednesday","thursday","friday","saturday","sunday",
    "spring","summer","autumn","fall","winter","semester","quarter",
}

QUANTITY_CONTEXT_WORDS = {
    "dollars","dollar","rupees","rupee","pounds","pound","euros","euro",
    "cents","cent","yen","won","francs","rand","pesos",
    "people","persons","items","units","pieces","copies","times",
    "votes","vote","points","point","members","member","employees","employee",
    "students","student","customers","customer","users","user",
    "attendees","attendee","residents","resident","participants","participant",
    "spectators","spectator","viewers","viewer","visitors","visitor",
    "subscribers","subscriber","followers","follower","listeners","listener",
    "cost","costs","price","prices","worth","total","totals","sum",
    "amount","amounts","quantity","quantities","count","counts",
    "population","populations","capacity","capacities","output","outputs",
    "approximately","about","around","over","under","nearly","almost",
    "exactly","only","just","more","less","than","least","most","roughly",
    "increased","decreased","grew","shrank","rose","fell","gained","lost",
    "totaling","totalling","amounting","sold","bought","spent",
    "earned","paid","received","collected","raised","donated","invested",
    "drew","gathered","attracted","hosted","reached","hit","exceeded",
}

def _detect_number_context(text, match_start, match_end):
    """
    Inspect words immediately before and after a matched number to decide
    whether it should be read as a year, a quantity, or left to the default
    heuristic (None).

    Looks at up to 3 words before and 2 words after the number.
    Year context takes priority if both signals are somehow present,
    since year-context words are more specific/intentional.
    """
    before = text[max(0, match_start-40):match_start]
    after  = text[match_end:match_end+40]

    before_words = re.findall(r"[A-Za-z]+\.?", before)[-3:]
    after_words  = re.findall(r"[A-Za-z]+\.?", after)[:2]

    surrounding = [w.lower().rstrip(".") for w in before_words + after_words]

    if any(w in YEAR_CONTEXT_WORDS for w in surrounding):
        return "year"
    if any(w in QUANTITY_CONTEXT_WORDS for w in surrounding):
        return "quantity"
    return None

def apply_numbers(text, locale="US"):
    changes = []

    def _ordinal_decade_word(n):
        """Return the '-ies' decade word for a multiple of 10 (10-90)."""
        base = integer_to_words(n)   # "ten","twenty",..."ninety"
        if base.endswith("y"):
            return base[:-1] + "ies"   # twenty -> twenties
        return base + "s"               # ten -> tens

    def _decade_repl(m):
        num_str = m.group(1)
        n = int(num_str)
        if n % 10 != 0:
            return m.group(0)

        if n % 100 == 0 and n >= 100:
            high = n // 100
            word = integer_to_words(high) + " hundreds"
        elif 10 <= n <= 90:
            word = _ordinal_decade_word(n)
        elif n >= 100:
            prefix_n = n // 100
            decade_digits = n % 100   # e.g. 20
            if decade_digits == 0:
                word = integer_to_words(prefix_n) + " hundreds"
            else:
                word = integer_to_words(prefix_n) + " " + _ordinal_decade_word(decade_digits)
        else:
            return m.group(0)

        changes.append((m.group(0), word))
        return word

    text = re.sub(r'\b(\d{1,3}0)s\b', _decade_repl, text)  # 10s..90s, 100s..9990s

    use_indian = _detect_indian_numbering(text, 0, len(text), locale)

    def _num_repl(m):
        start = m.start()
        if start > 0 and text[start-1].isalpha():
            return m.group(0)
        context_hint = _detect_number_context(text, m.start(), m.end())
        word, typ = number_token_to_words(m.group(0), context_hint, indian=use_indian)
        if typ:
            changes.append((m.group(0), word))
        return word
    text = re.sub(r'-?\d+(?:\.\d+)?(?:st|nd|rd|th)?', _num_repl, text)
    text = re.sub(r'\b(zero|one|two|three|four|five|six|seven|eight|nine|ten|'
                  r'eleven|twelve|thirteen|fourteen|fifteen|sixteen|seventeen|'
                  r'eighteen|nineteen|twenty|thirty|forty|fifty|sixty|seventy|'
                  r'eighty|ninety|hundred|thousand|million|billion)(am|pm)\b',
                  r'\1 \2', text, flags=re.IGNORECASE)
    return text, changes

def apply_diacritics(text):
    changes = []
    result = []
    for ch in text:
        norm = unicodedata.normalize("NFD", ch)
        stripped = "".join(c for c in norm if unicodedata.category(c) != "Mn")
        if stripped != ch:
            changes.append((ch, stripped))
        result.append(stripped)
    return "".join(result), changes

def apply_abbreviations(text, custom_dict):
    changes = []
    unverified = []
    tokens = text.split(" ")
    out = []
    for tok in tokens:
        lead  = re.match(r'^([^A-Za-z0-9\']*)', tok)
        trail = re.search(r'([^A-Za-z0-9\'"]*?)$', tok)
        prefix = lead.group(1) if lead else ""
        suffix = trail.group(1) if trail else ""
        core = tok[len(prefix):len(tok)-len(suffix)] if suffix else tok[len(prefix):]
        upper_core = core.upper()

        if upper_core in custom_dict:
            expanded = custom_dict[upper_core]
            changes.append((core, expanded, "custom_dict"))
            out.append(prefix + expanded + suffix)
        else:
            expanded, was, method = expand_abbreviation(core, core)
            if was:
                changes.append((core, expanded, method))
            elif (method is None
                  and core == core.upper()
                  and core.isalpha()
                  and 2 <= len(core) <= 6
                  and core.lower() not in COMMON_WORDS):
                unverified.append(core)
            out.append(prefix + expanded + suffix)
    return " ".join(out), changes, unverified

def strip_punctuation(text):
    return re.sub(r"[^\w\s']", " ", text)

def fix_spacing(text):
    changes = []
    stripped = text.strip()
    if stripped != text:
        snippet = text[:40].replace("\n", " ")
        changes.append(("leading/trailing spaces", snippet))
    def _fix(m):
        start = max(0, m.start() - 20)
        end   = min(len(text), m.end() + 20)
        snippet = "…" + text[start:end].replace("\n", " ") + "…"
        changes.append(("extra space", snippet))
        return " "
    result = re.sub(r' {2,}', _fix, stripped)
    return result, changes

def apply_case(text, mode):
    if mode == "UPPERCASE":    return text.upper()
    elif mode == "lowercase":  return text.lower()
    elif mode == "Sentence case":
        s = text.lower()
        return s[0].upper() + s[1:] if s else s
    elif mode == "Title Case": return text.title()
    else:                      return text

def normalize_text(text, options, custom_dict, chem_decisions=None):
    """
    chem_decisions: dict of {formula: "expand"|"spell"|"leave"} or None.
    If provided, chemicals are substituted before all other passes.
    """
    report = {
        "corrections":[],
        "spacing":[],"time":[],"temperature":[],"units":[],"numbers":[],
        "symbols":[],"abbreviations":[],"diacritics":[],"unverified":[],
        "chemicals":[],"ambiguous":[],"spelling":[],
    }
    lines = text.splitlines()
    out_lines = []

    for lineno, line in enumerate(lines, 1):

        line, gfixes = apply_source_grammar_fixes(line)
        for orig, conv in gfixes:
            report["corrections"].append({"line": lineno, "original": orig, "corrected": conv})

        if chem_decisions:
            line, chem_ch = apply_chemicals(line, chem_decisions)
            for orig, conv in chem_ch:
                report["chemicals"].append({"line": lineno, "original": orig, "converted": conv})

        if options.get("numbers", True):
            line, quad = apply_dotted_quad(line)
            for orig, conv in quad:
                report["numbers"].append({"line": lineno, "original": orig, "converted": conv})

        if options.get("numbers", True):
            line, rat = apply_ratios(line)
            for orig, conv in rat:
                report["numbers"].append({"line": lineno, "original": orig, "converted": conv})

        px_config = options.get("partial_expansion", {})
        if px_config:
            line, px_ch = apply_partial_expansion(line, px_config)
            for orig, conv in px_ch:
                report["numbers"].append({"line": lineno, "original": orig, "converted": conv})
        line, cod, amb = apply_code_numbers(line)
        for orig, conv in cod:
            report["numbers"].append({"line": lineno, "original": orig, "converted": conv})
        for orig, spoken in amb:
            report["ambiguous"].append({"line": lineno, "original": orig, "spoken": spoken,
                "comment": f"Ambiguous security code detected: '{orig}' — verify intended reading"})

        line, flt = apply_flight_numbers(line)
        for orig, conv in flt:
            report["numbers"].append({"line": lineno, "original": orig, "converted": conv})

        if options.get("numbers", True):
            line, pcode = apply_product_codes(line)
            for orig, conv in pcode:
                report["numbers"].append({"line": lineno, "original": orig, "converted": conv})

        if options.get("symbols", True):
            line, curr = apply_currency_codes(line, expand=options.get("expand_currency", False))
            for orig, conv in curr:
                report["symbols"].append({"line": lineno, "original": orig, "converted": conv})

        if options.get("symbols", True) or options.get("punctuation", True):
            line, rq = apply_rate_qualifiers(line)
            for orig, conv in rq:
                report["symbols"].append({"line": lineno, "original": orig, "converted": conv})

        if options.get("numbers", True):
            line, dim = apply_dimensions_tracked(line)
            for orig, conv in dim:
                report["numbers"].append({"line": lineno, "original": orig, "converted": conv})

        if options.get("numbers", True):
            line, rng = apply_ranges(line)
            for orig, conv in rng:
                report["numbers"].append({"line": lineno, "original": orig, "converted": conv})

        if options.get("numbers", True):
            line, ksuf = apply_k_suffix(line)
            for orig, conv in ksuf:
                report["numbers"].append({"line": lineno, "original": orig, "converted": conv})

        if options.get("numbers", True):
            line, frac = apply_fractions(line)
            for orig, conv in frac:
                report["numbers"].append({"line": lineno, "original": orig, "converted": conv})

        if options.get("numbers", True):
            line, dottm = apply_dot_time(line)
            for orig, conv in dottm:
                report["time"].append({"line": lineno, "original": orig, "converted": conv})

        if options.get("spacing", True):
            line, sp = fix_spacing(line)
            for typ, snip in sp:
                report["spacing"].append({"line":lineno,"type":typ,"snippet":snip})

        if options.get("numbers", True):
            line, tm = apply_time(line)
            for orig, conv in tm:
                report["time"].append({"line":lineno,"original":orig,"converted":conv})

        if options.get("symbols", True) or options.get("numbers", True):
            line, temp = apply_temperature(line)
            for orig, conv in temp:
                report["temperature"].append({"line":lineno,"original":orig,"converted":conv})

        if options.get("numbers", True) or options.get("symbols", True):
            line, cxr = apply_complex_rates(line)
            for orig, conv in cxr:
                report["units"].append({"line":lineno,"original":orig,"converted":conv})

        if options.get("numbers", True) or options.get("symbols", True):
            line, unt = apply_units(line, locale=options.get("locale", "US"))
            for orig, conv in unt:
                report["units"].append({"line":lineno,"original":orig,"converted":conv})

        if options.get("symbols", True):
            line, sym = apply_symbols(line, locale=options.get("locale", "US"))
            for orig, conv in sym:
                report["symbols"].append({"line":lineno,"original":orig,"converted":conv})

        if options.get("numbers", True):
            line, num = apply_numbers(line, locale=options.get("locale", "US"))
            for orig, conv in num:
                report["numbers"].append({"line":lineno,"original":orig,"converted":conv})

        if options.get("diacritics", True):
            line, dia = apply_diacritics(line)
            for orig, conv in dia:
                report["diacritics"].append({"line":lineno,"original":orig,"converted":conv})

        if options.get("abbreviations", True):
            line, abbr, unver = apply_abbreviations(line, custom_dict)
            for orig, conv, method in abbr:
                report["abbreviations"].append({
                    "line":lineno,"original":orig,"converted":conv,"method":method})
            for tok in unver:
                report["unverified"].append({"line":lineno,"token":tok})

        if options.get("punctuation", True):
            line = strip_punctuation(line)
            line = re.sub(r' {2,}', ' ', line).strip()

        out_lines.append(line)

    result = "\n".join(out_lines)
    result = apply_case(result, options.get("case", "UPPERCASE"))

    if options.get("spell_check"):
        report["spelling"] = check_spelling(text)

    return result, report



def detect_format_errors(text):
    """
    Flag patterns in the input that look like formatting mistakes rather
    than content that needs normalization — things that would likely
    indicate a malformed source file or an OCR/encoding glitch.

    Returns a list of issue dicts with keys: line, token, issue_type,
    suggestion, snippet.
    """
    issues = []
    for lineno, line in enumerate(text.splitlines(), 1):
        def _snippet(pos, length=30):
            s = max(0, pos - length)
            e = min(len(line), pos + length)
            return "…" + line[s:e] + "…"

        for open_c, close_c, name in [("(",")","parenthesis"),
                                        ("[","]","square bracket"),
                                        ("{","}","curly brace")]:
            if line.count(open_c) != line.count(close_c):
                issues.append({"line":lineno,"token":open_c+close_c,
                    "issue_type":f"Unbalanced {name}",
                    "suggestion":"check matching pairs","snippet":line[:50]})
        for q in ['"', "'"]:
            if line.count(q) % 2 != 0:
                issues.append({"line":lineno,"token":q,
                    "issue_type":"Unbalanced quote",
                    "suggestion":"check matching pairs","snippet":line[:50]})

        for m in re.finditer(r'\b\d[\d.,]*\b', line):
            tok = m.group(0)
            if tok.count(".") > 1:
                issues.append({"line":lineno,"token":tok,
                    "issue_type":"Malformed number (multiple decimal points)",
                    "suggestion":"verify source value","snippet":_snippet(m.start())})
            elif tok.endswith(".") or tok.endswith(","):
                issues.append({"line":lineno,"token":tok,
                    "issue_type":"Malformed number (trailing separator)",
                    "suggestion":"verify source value","snippet":_snippet(m.start())})

        for m in re.finditer(r'([!?,;:]){2,}', line):
            issues.append({"line":lineno,"token":m.group(0),
                "issue_type":"Repeated punctuation",
                "suggestion":m.group(1),"snippet":_snippet(m.start())})

        for m in re.finditer(r"\b[A-Za-z]+\b", line):
            tok = m.group(0)
            if len(tok) > 2 and tok not in (tok.upper(), tok.lower(), tok.title(), tok.capitalize()):
                if not re.fullmatch(r'[a-z]+(?:[A-Z][a-z]*)+', tok):
                    issues.append({"line":lineno,"token":tok,
                        "issue_type":"Irregular capitalisation",
                        "suggestion":"check intended casing","snippet":_snippet(m.start())})

        for i, ch in enumerate(line):
            if unicodedata.category(ch) == "Cc" and ch != "\t":
                issues.append({"line":lineno,"token":repr(ch),
                    "issue_type":"Control character",
                    "suggestion":"remove","snippet":_snippet(i)})

        if "\t" in line and " " in line:
            issues.append({"line":lineno,"token":repr("\\t"),
                "issue_type":"Mixed tabs and spaces",
                "suggestion":"use consistent whitespace","snippet":line[:50]})

    return issues


def detect_abbreviations_in_text(text):
    """
    Return a sorted list of unique abbreviation-like tokens found in text
    that are not already in standard dictionaries. Each entry is a dict:
      {"token": str, "suggestion": str, "lines": [int, ...]}
    Detects:
      - All-caps tokens of 2-6 letters (e.g. GDP, NASA, OTP)
      - Tokens in SPELLED_OUT or PRONOUNCEABLE sets
      - Mixed-case known abbreviations (e.g. kWh, MHz)
    Skips tokens already handled by units, chemicals, or common words.
    """
    found = {}      # token -> suggestion
    lines_by_tok = {}  # token -> set of line numbers

    for lineno, line in enumerate(text.splitlines(), 1):
        for m in re.finditer(r'\b([A-Z]{2,6})\b', line):
            tok = m.group(1)
            if tok in COMMON_WORDS or tok.lower() in COMMON_WORDS:
                continue
            if tok in UNIT_EXACT or tok in CHEM_DICT:
                continue
            suggestion = ""
            if tok in SPELLED_OUT:
                suggestion = " ".join(list(tok))
            elif tok in PRONOUNCEABLE:
                suggestion = tok.capitalize() if len(tok) > 4 else tok
            found[tok] = suggestion
            lines_by_tok.setdefault(tok, set()).add(lineno)
        for tok in MIXED_CASE_ABBR:
            if tok in line:
                found[tok] = " ".join(list(tok.upper()))
                lines_by_tok.setdefault(tok, set()).add(lineno)

    return [{"token": k, "suggestion": v, "lines": sorted(lines_by_tok[k])}
            for k, v in sorted(found.items())]


    """
    Flag patterns in the input that look like formatting mistakes rather
    than content that needs normalization — things that would likely
    indicate a malformed source file or an OCR/encoding glitch.

    Returns a list of issue dicts with keys: line, token, issue_type,
    suggestion, snippet.
    """
    issues = []
    for lineno, line in enumerate(text.splitlines(), 1):
        def _snippet(pos, length=30):
            s = max(0, pos - length)
            e = min(len(line), pos + length)
            return "…" + line[s:e] + "…"

        for open_c, close_c, name in [("(",")","parenthesis"),
                                        ("[","]","square bracket"),
                                        ("{","}","curly brace")]:
            if line.count(open_c) != line.count(close_c):
                issues.append({"line":lineno,"token":open_c+close_c,
                    "issue_type":f"Unbalanced {name}",
                    "suggestion":"check matching pairs","snippet":line[:50]})
        for q in ['"', "'"]:
            if line.count(q) % 2 != 0:
                issues.append({"line":lineno,"token":q,
                    "issue_type":"Unbalanced quote",
                    "suggestion":"check matching pairs","snippet":line[:50]})

        for m in re.finditer(r'\b\d[\d.,]*\b', line):
            tok = m.group(0)
            if tok.count(".") > 1:
                issues.append({"line":lineno,"token":tok,
                    "issue_type":"Malformed number (multiple decimal points)",
                    "suggestion":"verify source value","snippet":_snippet(m.start())})
            elif tok.endswith(".") or tok.endswith(","):
                issues.append({"line":lineno,"token":tok,
                    "issue_type":"Malformed number (trailing separator)",
                    "suggestion":"verify source value","snippet":_snippet(m.start())})

        for m in re.finditer(r'([!?,;:]){2,}', line):
            issues.append({"line":lineno,"token":m.group(0),
                "issue_type":"Repeated punctuation",
                "suggestion":m.group(1),"snippet":_snippet(m.start())})

        for m in re.finditer(r"\b[A-Za-z]+\b", line):
            tok = m.group(0)
            if len(tok) > 2 and tok not in (tok.upper(), tok.lower(), tok.title(), tok.capitalize()):
                if not re.fullmatch(r'[a-z]+(?:[A-Z][a-z]*)+', tok):
                    issues.append({"line":lineno,"token":tok,
                        "issue_type":"Irregular capitalisation",
                        "suggestion":"check intended casing","snippet":_snippet(m.start())})

        for i, ch in enumerate(line):
            if unicodedata.category(ch) == "Cc" and ch != "\t":
                issues.append({"line":lineno,"token":repr(ch),
                    "issue_type":"Control character",
                    "suggestion":"remove","snippet":_snippet(i)})

        if "\t" in line and " " in line:
            issues.append({"line":lineno,"token":repr("\\t"),
                "issue_type":"Mixed tabs and spaces",
                "suggestion":"use consistent whitespace","snippet":line[:50]})

    return issues


def check_spelling(text):
    """
    Flag suspected misspellings using pyspellchecker.
    Returns list of {"line": int, "token": str, "suggestion": str}.
    Gracefully returns [] if pyspellchecker is not installed.
    Skips all-caps tokens, likely proper nouns, and known common words.
    """
    try:
        from spellchecker import SpellChecker
    except ImportError:
        return []
    spell  = SpellChecker()
    issues = []
    for lineno, line in enumerate(text.splitlines(), 1):
        words = re.findall(r'\b[A-Za-z]{3,}\b', line)
        for word in words:
            if word.isupper():
                continue
            if word[0].isupper():
                continue   # skip likely proper nouns / sentence-start
            lower = word.lower()
            if lower in COMMON_WORDS:
                continue
            if spell.unknown([lower]):
                suggestion = spell.correction(lower) or ""
                if suggestion and suggestion != lower:
                    issues.append({
                        "line":       lineno,
                        "token":      word,
                        "suggestion": suggestion,
                    })
    return issues


def detect_consistency_issues(text):
    """
    Cross-line/cross-row consistency checks for QC. Flags:
      - The same abbreviation token spelled out two different ways across lines
      - Mixed casing conventions across lines (some UPPERCASE, some not)
      - Mixed number-locale conventions (lakh/crore wording mixed with
        million/billion wording within the same file)
    Returns list of {"line": int, "token": str, "issue_type": str, "suggestion": str, "snippet": str}.
    """
    issues = []
    lines = text.splitlines()

    token_forms = {}   # base token (no spaces) -> set of forms seen
    for lineno, line in enumerate(lines, 1):
        for m in re.finditer(r'\b([A-Z](?:\s[A-Z]){1,5})\b', line):
            spelled = m.group(1)
            solid = spelled.replace(" ", "")
            token_forms.setdefault(solid, set()).add(("spelled", lineno))
        for m in re.finditer(r'\b([A-Z]{2,6})\b', line):
            solid = m.group(1)
            token_forms.setdefault(solid, set()).add(("solid", lineno))

    for solid, forms in token_forms.items():
        kinds = {k for k, _ in forms}
        if len(kinds) > 1:
            lines_involved = sorted({ln for _, ln in forms})
            issues.append({
                "line": lines_involved[0], "token": solid,
                "issue_type": "Inconsistent abbreviation form",
                "suggestion": f"appears both spelled-out and solid across lines {lines_involved}",
                "snippet": "",
            })

    case_counts = {"upper": 0, "mixed": 0}
    for line in lines:
        letters = re.sub(r'[^A-Za-z]', '', line)
        if not letters:
            continue
        if letters == letters.upper():
            case_counts["upper"] += 1
        else:
            case_counts["mixed"] += 1
    if case_counts["upper"] > 0 and case_counts["mixed"] > 0:
        issues.append({
            "line": 1, "token": "(document-wide)",
            "issue_type": "Mixed casing convention",
            "suggestion": f"{case_counts['upper']} fully-uppercase line(s) vs "
                          f"{case_counts['mixed']} mixed-case line(s) — verify consistency",
            "snippet": "",
        })

    has_indian  = any(w in line.lower() for line in lines
                       for w in ("lakh","crore","arab"))
    has_western = any(w in line.lower() for line in lines
                       for w in ("million","billion","trillion"))
    if has_indian and has_western:
        issues.append({
            "line": 1, "token": "(document-wide)",
            "issue_type": "Mixed number-locale wording",
            "suggestion": "both lakh/crore and million/billion wording found — verify locale setting",
            "snippet": "",
        })

    return issues


def qc_check_text(text, options):
    issues = []
    expected_case = options.get("check_case") and options.get("expected_case")

    if options.get("format_errors"):
        issues.extend(detect_format_errors(text))

    if options.get("consistency", True):
        issues.extend(detect_consistency_issues(text))

    if options.get("spell_check"):
        for sp in check_spelling(text):
            issues.append({"line": sp["line"], "token": sp["token"],
                "issue_type": "Possible spelling issue",
                "suggestion": sp["suggestion"], "snippet": ""})

    for lineno, line in enumerate(text.splitlines(), 1):
        def _snippet(pos, length=30):
            s = max(0, pos - length)
            e = min(len(line), pos + length)
            return "…" + line[s:e] + "…"

        if options.get("numbers"):
            for m in re.finditer(r'\b\d+(?:st|nd|rd|th)?\b', line):
                issues.append({"line":lineno,"token":m.group(0),
                    "issue_type":"Lingering number",
                    "suggestion":number_token_to_words(m.group(0))[0],
                    "snippet":_snippet(m.start())})

        if options.get("symbols"):
            for sym in ['&','$','#','%','@','£','€','₹','©','®','™','°']:
                for m in re.finditer(re.escape(sym), line):
                    issues.append({"line":lineno,"token":sym,
                        "issue_type":"Lingering symbol",
                        "suggestion":"expand symbol","snippet":_snippet(m.start())})

        if options.get("diacritics"):
            for i, ch in enumerate(line):
                norm = unicodedata.normalize("NFD", ch)
                stripped = "".join(c for c in norm if unicodedata.category(c) != "Mn")
                if stripped != ch:
                    issues.append({"line":lineno,"token":ch,
                        "issue_type":"Lingering diacritic",
                        "suggestion":stripped,"snippet":_snippet(i)})

        if options.get("spacing"):
            for m in re.finditer(r' {2,}', line):
                issues.append({"line":lineno,"token":repr(m.group(0)),
                    "issue_type":"Extra space","suggestion":"single space",
                    "snippet":_snippet(m.start())})
            if line != line.strip():
                issues.append({"line":lineno,"token":repr(line[:10]),
                    "issue_type":"Leading/trailing space",
                    "suggestion":"strip line","snippet":line[:50]})

        if options.get("abbreviations"):
            for m in re.finditer(r'\b[A-Z]{2,6}\b', line):
                tok = m.group(0)
                if tok not in PRONOUNCEABLE and tok.lower() not in COMMON_WORDS:
                    exp, was, _ = expand_abbreviation(tok, tok)
                    if was:
                        issues.append({"line":lineno,"token":tok,
                            "issue_type":"Possible unexpanded abbreviation",
                            "suggestion":exp,"snippet":_snippet(m.start())})

        if expected_case:
            mode = options.get("expected_case")
            for w in re.findall(r"[A-Za-z']+", line):
                if mode == "UPPERCASE" and w != w.upper():
                    issues.append({"line":lineno,"token":w,
                        "issue_type":"Case mismatch (expected UPPERCASE)",
                        "suggestion":w.upper(),"snippet":line[:50]})
                elif mode == "lowercase" and w != w.lower():
                    issues.append({"line":lineno,"token":w,
                        "issue_type":"Case mismatch (expected lowercase)",
                        "suggestion":w.lower(),"snippet":line[:50]})
    return issues

def build_qc_report(issues, filenames=None):
    from collections import Counter
    lines = []
    lines += ["="*64, "  NORMALIZER — QC REPORT",
              f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
    if filenames:
        lines.append(f"  Files checked: {len(filenames)}")
        for f in filenames: lines.append(f"    · {f}")
    lines += ["="*64, ""]

    tc = Counter(i["issue_type"] for i in issues)
    lines.append("TOTALS")
    if not issues: lines.append("  No issues found.")
    for typ, count in sorted(tc.items()):
        lines.append(f"  {typ:<44} {count}")
    lines.append("")

    lines += ["-"*64, "  ISSUES DETAIL", "-"*64]
    if not issues:
        lines.append("  (none)")
    else:
        hdr = f"  {'Line':<6}  {'Token':<20}  {'Issue':<38}  {'Suggestion':<25}  Snippet"
        lines += [hdr, "  " + "-"*(len(hdr)-2)]
        for iss in issues:
            fn = iss.get("file","")
            lines.append(
                f"  {str(iss['line']):<6}  {str(iss['token']):<20}  "
                f"{str(iss['issue_type']):<38}  {str(iss['suggestion']):<25}  "
                f"{'['+fn+'] ' if fn else ''}{iss['snippet']}")
    lines.append("")

    lines += ["="*64, "  Created and Designed with love",
              "  All Rights Reserved © Priyangshu Swarnakar", "="*64]
    return "\n".join(lines)

def build_norm_report(report, batch_files=None):
    lines = []
    lines += ["="*64, "  NORMALIZER — SUMMARY REPORT",
              f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
    if batch_files:
        lines.append(f"  Files processed: {len(batch_files)}")
        for f in batch_files: lines.append(f"    · {f}")
    lines += ["="*64, "", "TOTALS"]

    for k, label in [
        ("corrections","Source text corrections"),
        ("chemicals","Chemical formulas processed"),
        ("spacing","Spacing fixes"),("time","Time expressions converted"),
        ("temperature","Temperature expressions converted"),
        ("units","Unit expressions converted"),("symbols","Symbols converted"),
        ("numbers","Numbers converted"),("abbreviations","Abbreviations expanded"),
        ("diacritics","Diacritics removed"),("unverified","Unverified tokens")]:
        lines.append(f"  {label:<38} {len(report.get(k,[]))}")
    lines.append("")

    def section(out, title, items, cols):
        out += ["-"*64, f"  {title.upper()}", "-"*64]
        if not items:
            out.append("  (none)")
        else:
            hdr = "  " + "  |  ".join(f"{c:<22}" for c in cols)
            out += [hdr, "  " + "-"*(len(hdr)-2)]
            for item in items:
                out.append("  " + "  |  ".join(f"{str(item.get(c,'')):<22}" for c in cols))
        out.append("")

    section(lines, "Source Text Corrections",        report.get("corrections",[]),   ["line","original","corrected"])
    section(lines, "Chemical Formulas",              report.get("chemicals",[]),     ["line","original","converted"])
    section(lines, "Spacing Fixes",                  report.get("spacing",[]),       ["line","type","snippet"])
    section(lines, "Time Expressions",               report.get("time",[]),          ["line","original","converted"])
    section(lines, "Temperature Expressions",        report.get("temperature",[]),   ["line","original","converted"])
    section(lines, "Unit Expressions",               report.get("units",[]),         ["line","original","converted"])
    section(lines, "Symbols Converted",              report.get("symbols",[]),       ["line","original","converted"])
    section(lines, "Numbers Converted",              report.get("numbers",[]),       ["line","original","converted"])
    section(lines, "Abbreviations Expanded",         report.get("abbreviations",[]), ["line","original","converted","method"])
    section(lines, "Diacritics Removed",             report.get("diacritics",[]),    ["line","original","converted"])

    lines += ["-"*64, "  UNVERIFIED TOKENS", "-"*64]
    unver = report.get("unverified",[])
    if not unver:
        lines.append("  (none)")
    else:
        for u in unver:
            lines.append(f"  Line {u['line']:<5}  {u['token']}")
    lines.append("")

    lines += ["="*64, "  Created and Designed with love",
              "  All Rights Reserved © Priyangshu Swarnakar", "="*64]
    return "\n".join(lines)



ELEMENTS = {
    "H":"hydrogen","He":"helium","Li":"lithium","Be":"beryllium","B":"boron",
    "C":"carbon","N":"nitrogen","O":"oxygen","F":"fluorine","Ne":"neon",
    "Na":"sodium","Mg":"magnesium","Al":"aluminium","Si":"silicon","P":"phosphorus",
    "S":"sulphur","Cl":"chlorine","Ar":"argon","K":"potassium","Ca":"calcium",
    "Sc":"scandium","Ti":"titanium","V":"vanadium","Cr":"chromium","Mn":"manganese",
    "Fe":"iron","Co":"cobalt","Ni":"nickel","Cu":"copper","Zn":"zinc",
    "Ga":"gallium","Ge":"germanium","As":"arsenic","Se":"selenium","Br":"bromine",
    "Kr":"krypton","Rb":"rubidium","Sr":"strontium","Y":"yttrium","Zr":"zirconium",
    "Nb":"niobium","Mo":"molybdenum","Tc":"technetium","Ru":"ruthenium",
    "Rh":"rhodium","Pd":"palladium","Ag":"silver","Cd":"cadmium","In":"indium",
    "Sn":"tin","Sb":"antimony","Te":"tellurium","I":"iodine","Xe":"xenon",
    "Cs":"caesium","Ba":"barium","La":"lanthanum","Ce":"cerium","Pr":"praseodymium",
    "Nd":"neodymium","Pm":"promethium","Sm":"samarium","Eu":"europium",
    "Gd":"gadolinium","Tb":"terbium","Dy":"dysprosium","Ho":"holmium",
    "Er":"erbium","Tm":"thulium","Yb":"ytterbium","Lu":"lutetium",
    "Hf":"hafnium","Ta":"tantalum","W":"tungsten","Re":"rhenium","Os":"osmium",
    "Ir":"iridium","Pt":"platinum","Au":"gold","Hg":"mercury","Tl":"thallium",
    "Pb":"lead","Bi":"bismuth","Po":"polonium","At":"astatine","Rn":"radon",
    "Fr":"francium","Ra":"radium","Ac":"actinium","Th":"thorium","Pa":"protactinium",
    "U":"uranium","Np":"neptunium","Pu":"plutonium","Am":"americium","Cm":"curium",
    "Bk":"berkelium","Cf":"californium","Es":"einsteinium","Fm":"fermium",
    "Md":"mendelevium","No":"nobelium","Lr":"lawrencium",
}

CHEM_DICT = {
    "H2O":    "water",
    "H2O2":   "hydrogen peroxide",
    "CO2":    "carbon dioxide",
    "CO":     "carbon monoxide",
    "O2":     "oxygen",
    "O3":     "ozone",
    "N2":     "nitrogen",
    "N2O":    "nitrous oxide",
    "NO":     "nitric oxide",
    "NO2":    "nitrogen dioxide",
    "SO2":    "sulphur dioxide",
    "SO3":    "sulphur trioxide",
    "H2S":    "hydrogen sulphide",
    "NH3":    "ammonia",
    "CH4":    "methane",
    "C2H6":   "ethane",
    "C3H8":   "propane",
    "C4H10":  "butane",
    "C2H4":   "ethylene",
    "C2H2":   "acetylene",
    "C6H6":   "benzene",
    "C6H12O6":"glucose",
    "C12H22O11":"sucrose",
    "C2H5OH": "ethanol",
    "CH3OH":  "methanol",
    "CH3COOH":"acetic acid",
    "HCOOH":  "formic acid",
    "C3H6O":  "acetone",
    "NaCl":   "sodium chloride",
    "KCl":    "potassium chloride",
    "CaCl2":  "calcium chloride",
    "MgCl2":  "magnesium chloride",
    "FeCl3":  "iron three chloride",
    "AlCl3":  "aluminium chloride",
    "NaOH":   "sodium hydroxide",
    "KOH":    "potassium hydroxide",
    "Ca(OH)2":"calcium hydroxide",
    "Mg(OH)2":"magnesium hydroxide",
    "NH4OH":  "ammonium hydroxide",
    "NaHCO3": "sodium bicarbonate",
    "Na2CO3": "sodium carbonate",
    "CaCO3":  "calcium carbonate",
    "K2CO3":  "potassium carbonate",
    "Na2SO4": "sodium sulphate",
    "CaSO4":  "calcium sulphate",
    "MgSO4":  "magnesium sulphate",
    "FeSO4":  "iron two sulphate",
    "Fe2(SO4)3":"iron three sulphate",
    "Na3PO4": "sodium phosphate",
    "Ca3(PO4)2":"calcium phosphate",
    "NaNO3":  "sodium nitrate",
    "KNO3":   "potassium nitrate",
    "NH4NO3": "ammonium nitrate",
    "NH4Cl":  "ammonium chloride",
    "NH4HCO3":"ammonium bicarbonate",
    "HCl":    "hydrochloric acid",
    "H2SO4":  "sulphuric acid",
    "HNO3":   "nitric acid",
    "H3PO4":  "phosphoric acid",
    "HF":     "hydrofluoric acid",
    "HBr":    "hydrobromic acid",
    "HI":     "hydroiodic acid",
    "H2CO3":  "carbonic acid",
    "Fe2O3":  "iron three oxide",
    "Fe3O4":  "iron two three oxide",
    "FeO":    "iron two oxide",
    "CuO":    "copper two oxide",
    "Cu2O":   "copper one oxide",
    "ZnO":    "zinc oxide",
    "MgO":    "magnesium oxide",
    "CaO":    "calcium oxide",
    "Al2O3":  "aluminium oxide",
    "SiO2":   "silicon dioxide",
    "TiO2":   "titanium dioxide",
    "MnO2":   "manganese dioxide",
    "PbO":    "lead two oxide",
    "PbO2":   "lead four oxide",
    "Pb3O4":  "lead three four oxide",
    "Cr2O3":  "chromium three oxide",
    "V2O5":   "vanadium five oxide",
    "WO3":    "tungsten trioxide",
    "SnO2":   "tin dioxide",
    "BaO":    "barium oxide",
    "K2O":    "potassium oxide",
    "Na2O":   "sodium oxide",
    "ATP":    "adenosine triphosphate",
    "ADP":    "adenosine diphosphate",
    "AMP":    "adenosine monophosphate",
    "DNA":    "D N A",
    "RNA":    "R N A",
    "NADH":   "N A D H",
    "NADPH":  "N A D P H",
    "FAD":    "F A D",
    "FADH2":  "F A D H two",
    "KMnO4":  "potassium permanganate",
    "K2Cr2O7":"potassium dichromate",
    "AgNO3":  "silver nitrate",
    "BaSO4":  "barium sulphate",
    "PbSO4":  "lead sulphate",
    "CuSO4":  "copper sulphate",
    "ZnSO4":  "zinc sulphate",
    "FeCl2":  "iron two chloride",
    "SnCl2":  "tin two chloride",
    "SnCl4":  "tin four chloride",
    "SiC":    "silicon carbide",
    "SiN":    "silicon nitride",
    "BN":     "boron nitride",
    "GaAs":   "gallium arsenide",
    "InP":    "indium phosphide",
    "CdTe":   "cadmium telluride",
}

CHEM_DETECT = re.compile(
    r'\b'
    r'[A-Z][a-z]?'                   # first element
    r'(?:\d*[A-Z][a-z]?\d*)*'        # more element+count groups
    r'\d+'                            # must end with or contain a digit
    r'(?:[A-Z][a-z]?\d*)*'           # optional trailing elements
    r'\b'
)

def detect_chemicals(text):
    """
    Return sorted list of {"text": str, "lines": [int, ...]} for unique
    chemical tokens found in text, with every line number each occurs on.
    Includes:
      - Formulas with digits detected by regex (CO2, H2SO4, C6H12O6)
      - Known no-digit formulas from CHEM_DICT (NaCl, KOH, HCl etc.)
    Excludes:
      - Flight numbers (2-3 uppercase letters + 1-4 digits: VA8000, QF4186)
      - Alphanumeric codes that are clearly not chemical formulas
    """
    found = {}   # token -> set of line numbers

    _flight_like = re.compile(r'^[A-Z]{2,3}\d{1,4}$')
    _has_element = re.compile(r'[A-Z][a-z]')

    for lineno, line in enumerate(text.splitlines(), 1):
        for m in CHEM_DETECT.finditer(line):
            tok = m.group(0)
            if _flight_like.match(tok) and not _has_element.search(tok):
                continue
            if re.match(r'^[A-Z]\d+$', tok) and tok not in CHEM_DICT:
                continue
            found.setdefault(tok, set()).add(lineno)

        for w in re.findall(r'\b[A-Za-z][A-Za-z0-9()]*\b', line):
            if w in CHEM_DICT:
                found.setdefault(w, set()).add(lineno)

    return [{"text": tok, "lines": sorted(lines)}
            for tok, lines in sorted(found.items())]


def detect_units_in_text(text):
    """
    Return sorted list of {"text": str, "lines": [int, ...]} for unique
    "<number> <unit>" snippets found in text, where <unit> is a recognised
    key in UNIT_EXACT.
    """
    found = {}
    for lineno, line in enumerate(text.splitlines(), 1):
        for key in _UNIT_KEYS_SORTED:
            if not key:
                continue
            pat = re.compile(
                r'(-?\d[\d,]*(?:\.\d+)?)\s*' + re.escape(key) + r'\b')
            for m in pat.finditer(line):
                snippet = m.group(0).strip()
                found.setdefault(snippet, set()).add(lineno)
    return [{"text": tok, "lines": sorted(lines)}
            for tok, lines in sorted(found.items())]

def detect_currency_in_text(text):
    """
    Return sorted list of {"text": str, "lines": [int, ...]} for unique
    currency snippets found in text: symbol-based ($12, £5.50, ₹1,000)
    and 3-letter code-based (500 INR).
    """
    found = {}
    code_alt = '|'.join(CURRENCY_CODE_MAP.keys())
    for lineno, line in enumerate(text.splitlines(), 1):
        for pattern, _word, _eligible in _CURRENCY_ROWS:
            for m in re.finditer(pattern, line):
                found.setdefault(m.group(0).strip(), set()).add(lineno)
        for m in re.finditer(r'\d[\d,]*(?:\.\d+)?\s*\b(?:' + code_alt + r')\b', line):
            found.setdefault(m.group(0).strip(), set()).add(lineno)
    return [{"text": tok, "lines": sorted(lines)}
            for tok, lines in sorted(found.items())]

def detect_flight_ref_ids(text, exclude=None):
    """
    Return sorted list of {"text": str, "lines": [int, ...]} for unique
    flight-number-like and product/model/passport-code-like tokens.
    `exclude` is an optional iterable of tokens to skip (e.g. tokens
    already identified as chemical formulas).
    """
    exclude = set(exclude or ())
    found = {}
    for lineno, line in enumerate(text.splitlines(), 1):
        for pat in (FLIGHT_PATTERN, PRODUCT_CODE_PATTERN, PASSPORT_CODE_PATTERN):
            for m in pat.finditer(line):
                tok = m.group(0)
                if tok in exclude:
                    continue
                found.setdefault(tok, set()).add(lineno)
    return [{"text": tok, "lines": sorted(lines)}
            for tok, lines in sorted(found.items())]

def detect_code_pins(text):
    """
    Return sorted list of {"text": str, "lines": [int, ...]} for unique
    "<trigger word> ... <digits>" snippets that apply_code_numbers would
    treat as a digit-by-digit code rather than a quantity.
    """
    found = {}
    trigger_alt = '|'.join(re.escape(w) for w in sorted(CODE_TRIGGER_WORDS | SECURITY_TRIGGER_WORDS, key=len, reverse=True))
    pattern = (
        r'(?i)\b(' + trigger_alt + r')\b'
        r'((?:\s+[A-Za-z]+){0,5}?\s+)'
        r'(\d{3,18})(?!\.\d)\b'
    )
    for lineno, line in enumerate(text.splitlines(), 1):
        for m in re.finditer(pattern, line):
            num_end = m.end(3)
            if _next_word_is_unit_or_quantity(line, num_end):
                continue
            found.setdefault(m.group(0).strip(), set()).add(lineno)
    return [{"text": tok, "lines": sorted(lines)}
            for tok, lines in sorted(found.items())]


def spell_out_formula(formula):
    """
    Convert a chemical formula to its spelled-out letter+number form.
    CO2 → C O two
    H2SO4 → H two S O four
    NaCl → Na Cl
    """
    clean = formula.replace("(", "").replace(")", "")

    tokens = re.findall(r'[A-Z][a-z]?|\d+', clean)
    parts = []
    for tok in tokens:
        if tok.isdigit():
            parts.append(integer_to_words(int(tok)))
        elif tok in ELEMENTS:
            parts.append(tok)   # keep symbol as-is; TTS reads each letter
        else:
            parts.append(tok)
    return " ".join(parts)

def apply_chemicals(text, decisions):
    """
    Replace chemical tokens in text according to decisions dict.
    decisions: {formula: "expand" | "spell" | "leave"}
    Returns (processed_text, list_of_changes)
    """
    changes = []
    for formula, decision in decisions.items():
        if decision == "leave":
            continue
        elif decision == "expand" and formula in CHEM_DICT:
            replacement = CHEM_DICT[formula]
        else:
            replacement = spell_out_formula(formula)

        escaped = re.escape(formula)
        def _repl(m, r=replacement, f=formula):
            changes.append((f, r))
            return " " + r + " "
        text = re.sub(r'\b' + escaped + r'\b', _repl, text)

    text = re.sub(r' {2,}', ' ', text).strip()
    return text, changes


class PreloadDialog(tk.Toplevel):
    """
    Dialog shown before normalization when the pre-load scan finds items
    worth reviewing: chemical formulas (interactive — user chooses how each
    is normalised), plus read-only review lists for units, currency,
    flight/reference-style codes, and code/PIN/OTP numbers (informational —
    these are converted automatically by the normal pipeline, this dialog
    simply surfaces them so the user can catch unexpected matches before
    running normalization).
    """
    def __init__(self, parent, formulas=None, units=None, currency=None,
                 flight_ref=None, code_pins=None, abbreviations=None,
                 row_label="Line"):
        super().__init__(parent)
        formulas      = formulas or []      # [{"text":.., "lines":[..]}]
        units         = units or []         # [{"text":.., "lines":[..]}]
        currency      = currency or []      # [{"text":.., "lines":[..]}]
        flight_ref    = flight_ref or []    # [{"text":.., "lines":[..]}]
        code_pins     = code_pins or []     # [{"text":.., "lines":[..]}]
        abbreviations = abbreviations or []  # [{"token":..,"suggestion":..,"lines":[..]}]
        self.row_label = row_label   # "Line" for .txt sources, "Row" for Excel sources

        self.title("Pre-load Detection Review")
        self.configure(bg=C["bg"])
        self.resizable(True, True)
        self.grab_set()
        self.decisions = {}       # formula → "expand"|"spell"|"leave"
        self._abbr_entries = {}   # token → tk.StringVar (user's custom expansion)

        tk.Label(self, text="Pre-load Detection Review",
                 font=FONT_TITLE, bg=C["bg"], fg=C["text"]).pack(padx=20, pady=(14,2))
        tk.Label(self,
                 text="Review detected items before normalization runs. "
                      "Chemical formulas and abbreviations allow custom choices; "
                      "other tabs are informational.",
                 font=FONT_SMALL, bg=C["bg"], fg=C["text_muted"],
                 justify="left", wraplength=660).pack(padx=20, pady=(0,8))

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=16, pady=4)

        def _scrollable_tab(label):
            """Create a scrollable frame tab, return (tab_frame, inner_frame)."""
            tab = tk.Frame(nb, bg=C["bg"])
            nb.add(tab, text=f"  {label}  ")
            outer = tk.Frame(tab, bg=C["border"], bd=1)
            outer.pack(fill="both", expand=True, padx=4, pady=4)
            canvas = tk.Canvas(outer, bg=C["bg"], bd=0, highlightthickness=0, height=360)
            vsb = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
            canvas.configure(yscrollcommand=vsb.set)
            vsb.pack(side="right", fill="y")
            canvas.pack(side="left", fill="both", expand=True)
            inner = tk.Frame(canvas, bg=C["bg"])
            canvas.create_window((0, 0), window=inner, anchor="nw")
            inner.bind("<Configure>",
                lambda e, c=canvas: c.configure(scrollregion=c.bbox("all")))
            return tab, inner

        def _lines_str(lines):
            """Format a list of line numbers compactly, e.g. '3, 7, 12' or '3-9'."""
            if not lines:
                return "—"
            if len(lines) > 6:
                return f"{lines[0]}-{lines[-1]} ({len(lines)} occurrences)"
            return ", ".join(str(n) for n in lines)

        self._vars = {}
        _, chem_inner = _scrollable_tab(f"Chemicals ({len(formulas)})")
        if formulas:
            hrow = tk.Frame(chem_inner, bg=C["bg2"])
            hrow.pack(fill="x", padx=4, pady=(4,0))
            for txt, w in [("Formula",14),(self.row_label,14),("Known expansion",24),
                           ("Expand",8),("Spell out",10),("Leave",8)]:
                tk.Label(hrow, text=txt, font=FONT_BOLD, bg=C["bg2"],
                         fg=C["text"], width=w, anchor="w").pack(side="left", padx=4)
            for i, item in enumerate(formulas):
                formula = item["text"]
                lines   = item["lines"]
                bg = C["row_even"] if i%2==0 else C["row_odd"]
                row = tk.Frame(chem_inner, bg=bg)
                row.pack(fill="x", padx=4, pady=2)
                known = CHEM_DICT.get(formula, "—")
                var = tk.StringVar(value="expand" if formula in CHEM_DICT else "spell")
                self._vars[formula] = var
                tk.Label(row, text=formula, font=FONT_BOLD, bg=bg,
                         fg=C["accent_h"], width=14, anchor="w").pack(side="left", padx=4)
                tk.Label(row, text=_lines_str(lines), font=FONT_SMALL, bg=bg,
                         fg=C["text_muted"], width=14, anchor="w").pack(side="left", padx=4)
                tk.Label(row, text=known, font=FONT_SMALL, bg=bg,
                         fg=C["text_muted"], width=24, anchor="w").pack(side="left", padx=4)
                for val, lbl, w in [("expand","Expand",8),
                                    ("spell","Spell out",10),("leave","Leave",8)]:
                    state = "normal" if (val!="expand" or formula in CHEM_DICT) else "disabled"
                    tk.Radiobutton(row, text="", variable=var, value=val,
                                   bg=bg, fg=C["text"], activebackground=bg,
                                   selectcolor=C["accent_lt"],
                                   state=state, width=w).pack(side="left", padx=2)
            af = tk.Frame(chem_inner, bg=C["bg"])
            af.pack(fill="x", padx=4, pady=(4,8))
            tk.Label(af, text="Apply to all:", font=FONT_SMALL,
                     bg=C["bg"], fg=C["text_muted"]).pack(side="left")
            self._all_var = tk.StringVar(value="none")
            for val, lbl in [("expand","All expand"),("spell","All spell out"),
                              ("leave","All leave"),("none","Individual")]:
                tk.Radiobutton(af, text=lbl, variable=self._all_var, value=val,
                               bg=C["bg"], fg=C["text"], activebackground=C["bg"],
                               selectcolor=C["accent_lt"], font=FONT_SMALL,
                               command=self._apply_all).pack(side="left", padx=6)
        else:
            self._all_var = tk.StringVar(value="none")
            tk.Label(chem_inner, text="No chemical formulas detected.",
                     font=FONT_SMALL, bg=C["bg"], fg=C["text_muted"],
                     anchor="w").pack(padx=8, pady=8)

        _, abbr_inner = _scrollable_tab(f"Abbreviations ({len(abbreviations)})")
        if abbreviations:
            tk.Label(abbr_inner,
                     text="Detected abbreviations. Edit the expansion field and click Save "
                          "to add to your Custom Dictionary. Leave blank to use the default.",
                     font=FONT_SMALL, bg=C["bg"], fg=C["text_muted"],
                     wraplength=620, justify="left").pack(padx=8, pady=(6,4))
            hrow = tk.Frame(abbr_inner, bg=C["bg2"])
            hrow.pack(fill="x", padx=4)
            for txt, w in [("Token",10),(self.row_label,12),("Default expansion",26),
                           ("Custom expansion",24),("",8)]:
                tk.Label(hrow, text=txt, font=FONT_BOLD, bg=C["bg2"],
                         fg=C["text"], width=w, anchor="w").pack(side="left", padx=4)
            for i, item in enumerate(abbreviations):
                tok   = item["token"]
                sugg  = item["suggestion"]
                lines = item.get("lines", [])
                bg    = C["row_even"] if i%2==0 else C["row_odd"]
                row   = tk.Frame(abbr_inner, bg=bg)
                row.pack(fill="x", padx=4, pady=2)
                tk.Label(row, text=tok, font=FONT_BOLD, bg=bg,
                         fg=C["accent_h"], width=10, anchor="w").pack(side="left", padx=4)
                tk.Label(row, text=_lines_str(lines), font=FONT_SMALL, bg=bg,
                         fg=C["text_muted"], width=12, anchor="w").pack(side="left", padx=4)
                tk.Label(row, text=sugg or "—", font=FONT_SMALL, bg=bg,
                         fg=C["text_muted"], width=26, anchor="w").pack(side="left", padx=4)
                evar = tk.StringVar(value=sugg)
                self._abbr_entries[tok] = evar
                ent = tk.Entry(row, textvariable=evar, font=FONT_SMALL,
                               bg=C["bg"], fg=C["text"], relief="flat",
                               highlightbackground=C["border"], highlightthickness=1,
                               width=24)
                ent.pack(side="left", padx=4)
                def _save_abbr(t=tok, v=evar):
                    exp = v.get().strip()
                    if exp:
                        self._abbr_save_callback(t, exp)
                tk.Button(row, text="Save", font=FONT_SMALL,
                          bg=C["accent"], fg=C["white"], relief="flat",
                          activebackground=C["accent_h"], cursor="hand2",
                          command=_save_abbr, padx=6, pady=2).pack(side="left", padx=4)
        else:
            tk.Label(abbr_inner, text="No abbreviations detected.",
                     font=FONT_SMALL, bg=C["bg"], fg=C["text_muted"],
                     anchor="w").pack(padx=8, pady=8)

        def _review_tab(label, items, hint):
            _, inner = _scrollable_tab(f"{label} ({len(items)})")
            tk.Label(inner, text=hint, font=FONT_SMALL, bg=C["bg"],
                     fg=C["text_muted"], anchor="w", justify="left",
                     wraplength=620).pack(fill="x", padx=8, pady=(6,4))
            if items:
                hrow = tk.Frame(inner, bg=C["bg2"])
                hrow.pack(fill="x", padx=4)
                tk.Label(hrow, text="Text", font=FONT_BOLD, bg=C["bg2"],
                         fg=C["text"], width=30, anchor="w").pack(side="left", padx=4)
                tk.Label(hrow, text=self.row_label, font=FONT_BOLD, bg=C["bg2"],
                         fg=C["text"], width=20, anchor="w").pack(side="left", padx=4)
                for i, item in enumerate(items):
                    bg = C["row_even"] if i%2==0 else C["row_odd"]
                    row = tk.Frame(inner, bg=bg)
                    row.pack(fill="x", padx=4, pady=1)
                    tk.Label(row, text=item["text"], font=FONT_SMALL, bg=bg,
                             fg=C["text"], width=30, anchor="w").pack(side="left", padx=4)
                    tk.Label(row, text=_lines_str(item["lines"]), font=FONT_SMALL, bg=bg,
                             fg=C["text_muted"], width=20, anchor="w").pack(side="left", padx=4)
            else:
                tk.Label(inner, text="None detected.", font=FONT_SMALL,
                         bg=C["bg"], fg=C["text_muted"], anchor="w").pack(padx=8, pady=4)

        _review_tab("Units", units,
            "Number + unit expressions that will be expanded to words (e.g. 5 kg → five kilograms).")
        _review_tab("Currency", currency,
            "Currency amounts/codes that will be expanded (e.g. $50 → fifty dollars).")
        _review_tab("Flight / Ref / Codes", flight_ref,
            "Alphanumeric codes that will be spelled letter-by-letter and digit-by-digit.")
        _review_tab("Code / PIN / OTP", code_pins,
            "Numbers after trigger words that will be spelled digit-by-digit.")

        tk.Button(self, text="Confirm and Continue", font=FONT_BOLD,
                  bg=C["accent"], fg=C["white"], relief="flat",
                  activebackground=C["accent_h"], cursor="hand2",
                  command=self._confirm, padx=20, pady=8).pack(pady=12)

    def _abbr_save_callback(self, token, expansion):
        """Save a custom abbreviation expansion to custom_dict.json."""
        try:
            dict_file = Path(os.environ.get("APPDATA", Path.home())) / "Chichingpoipoi" / "custom_dict.json"
            dict_file.parent.mkdir(parents=True, exist_ok=True)
            data = {}
            if dict_file.exists():
                try:
                    data = json.loads(dict_file.read_text(encoding="utf-8"))
                except Exception:
                    pass
            data[token.upper()] = expansion
            dict_file.write_text(json.dumps(data, indent=2, ensure_ascii=False),
                                 encoding="utf-8")
            messagebox.showinfo("Saved",
                f"'{token}' → '{expansion}' saved to Custom Dictionary.",
                parent=self)
        except Exception as e:
            messagebox.showerror("Save error", str(e), parent=self)

    def _apply_all(self):
        choice = self._all_var.get()
        if choice == "none":
            return
        for formula, var in self._vars.items():
            if choice == "expand" and formula not in CHEM_DICT:
                var.set("spell")   # can't expand what we don't know
            else:
                var.set(choice)

    def _confirm(self):
        for formula, var in self._vars.items():
            self.decisions[formula] = var.get()
        self.destroy()


ChemicalDialog = PreloadDialog


def read_excel_column(filepath, sheet_name, column_name):
    """
    Read all data rows (excluding header) for the given column from an
    Excel workbook. Returns (rows, header_row_idx) where rows is a list
    of (excel_row_number, text) tuples, excel_row_number being the
    1-indexed row as it appears in Excel (so row 2 is the first data row
    when row 1 is the header).
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]
    header = [str(c) if c is not None else f"Column {i+1}"
              for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ()))]
    col_idx = header.index(column_name)
    rows = []
    for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        val = row[col_idx] if col_idx < len(row) else None
        rows.append((r, "" if val is None else str(val)))
    wb.close()
    return rows, col_idx, header


def write_excel_normalized(filepath, sheet_name, column_name, normalized_by_row, output_path):
    """
    Write a copy of the workbook with a new "<column>_Normalized" column
    inserted immediately to the right of the source column. All other
    columns and formatting in the row are preserved (Option A).
    normalized_by_row: dict of {excel_row_number: normalized_text}
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath)
    ws = wb[sheet_name]

    header_row = next(ws.iter_rows(min_row=1, max_row=1), ())
    header = [str(c.value) if c.value is not None else f"Column {i+1}"
              for i, c in enumerate(header_row)]
    col_idx = header.index(column_name)   # 0-indexed
    insert_at = col_idx + 2               # openpyxl is 1-indexed; insert right after source col

    ws.insert_cols(insert_at)
    ws.cell(row=1, column=insert_at, value=f"{column_name}_Normalized")

    for r, text in normalized_by_row.items():
        ws.cell(row=r, column=insert_at, value=text)

    from openpyxl.utils import get_column_letter
    src_letter = get_column_letter(col_idx + 1)
    new_letter = get_column_letter(insert_at)
    if src_letter in ws.column_dimensions:
        ws.column_dimensions[new_letter].width = ws.column_dimensions[src_letter].width

    wb.save(output_path)


def flatten_report_changes(report):
    """
    Flatten a normalize_text() report dict into a single list of
    {"line": int, "category": str, "original": str, "converted": str}
    entries suitable for the Excel "Changes" sheet.
    """
    changes = []
    for category in ("corrections", "chemicals", "numbers", "symbols",
                      "time", "temperature", "units", "diacritics", "abbreviations"):
        for entry in report.get(category, []):
            changes.append({
                "line": entry.get("line", "—"),
                "category": category,
                "original": entry.get("original", ""),
                "converted": entry.get("converted", entry.get("corrected", "")),
            })
    return changes


def export_qc_issues_excel(file_reports, output_path, row_label="Line"):
    """
    Generate a two-sheet QC issues report Excel file (separate from the
    normalization report — QC issues use issue_type/snippet shape rather
    than category/converted shape).

    file_reports: list of dicts:
        {
          "filename": str, "needs_qc": bool, "comments": [str,...],
          "qc_issues": [ {"line":int,"token":str,"issue_type":str,
                           "suggestion":str,"snippet":str} ],
        }
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    GREEN  = PatternFill("solid", fgColor="C6EFCE")
    RED    = PatternFill("solid", fgColor="FFC7CE")
    YELLOW = PatternFill("solid", fgColor="FFEB9C")
    HDR_FILL = PatternFill("solid", fgColor="2F4F6F")
    HDR_FONT = Font(color="FFFFFF", bold=True)
    BOLD   = Font(bold=True)
    CENTER = Alignment(horizontal="center", vertical="center")
    WRAP   = Alignment(wrap_text=True, vertical="top")
    thin   = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _hdr_row(ws, cols, row=1):
        for c, title in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=title)
            cell.fill, cell.font, cell.alignment, cell.border = HDR_FILL, HDR_FONT, CENTER, BORDER

    def _set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "All Files"
    ws1.row_dimensions[1].height = 22
    _hdr_row(ws1, ["#", "File Name", "Needs QC", "Issue Summary"])
    _set_col_widths(ws1, [5, 40, 12, 70])

    for i, fr in enumerate(file_reports, 1):
        row = i + 1
        needs = fr["needs_qc"]
        comment_text = "\n".join(fr["comments"]) if fr["comments"] else "All clear"
        ws1.cell(row=row, column=1, value=i).alignment = CENTER
        ws1.cell(row=row, column=2, value=fr["filename"]).alignment = WRAP
        qc_cell = ws1.cell(row=row, column=3, value="YES" if needs else "NO")
        qc_cell.fill = GREEN if not needs else RED
        qc_cell.font = Font(bold=True, color="375623" if not needs else "9C0006")
        qc_cell.alignment = CENTER
        cm_cell = ws1.cell(row=row, column=4, value=comment_text)
        cm_cell.alignment = WRAP
        for c in range(1, 5):
            ws1.cell(row=row, column=c).border = BORDER
        ws1.row_dimensions[row].height = max(15, 15 * (comment_text.count("\n") + 1))
    ws1.freeze_panes = "A2"

    ws2 = wb.create_sheet("Issue Detail")
    ws2.row_dimensions[1].height = 22
    _hdr_row(ws2, ["#", "File Name", row_label, "Issue Type", "Token", "Suggestion", "Snippet"])
    _set_col_widths(ws2, [5, 30, 8, 26, 20, 26, 40])

    detail_row = 2
    num = 0
    for fr in file_reports:
        issues = fr.get("qc_issues", [])
        if not issues:
            continue
        num += 1
        for c in range(1, 8):
            ws2.cell(row=detail_row, column=c).fill = PatternFill("solid", fgColor="D9E1F2")
            ws2.cell(row=detail_row, column=c).border = BORDER
        ws2.cell(row=detail_row, column=1, value=num).alignment = CENTER
        ws2.cell(row=detail_row, column=2, value=fr["filename"]).font = BOLD
        ws2.cell(row=detail_row, column=4, value=f"{len(issues)} issue(s)").font = BOLD
        detail_row += 1
        for iss in sorted(issues, key=lambda x: x.get("line", 0)):
            for c in range(1, 8):
                ws2.cell(row=detail_row, column=c).border = BORDER
            ws2.cell(row=detail_row, column=3, value=iss.get("line","—")).alignment = CENTER
            ws2.cell(row=detail_row, column=4, value=iss.get("issue_type",""))
            ws2.cell(row=detail_row, column=5, value=iss.get("token",""))
            sugg_cell = ws2.cell(row=detail_row, column=6, value=iss.get("suggestion",""))
            sugg_cell.fill = YELLOW
            sugg_cell.alignment = WRAP
            snip_cell = ws2.cell(row=detail_row, column=7, value=iss.get("snippet",""))
            snip_cell.alignment = WRAP
            detail_row += 1

    if detail_row == 2:
        ws2.cell(row=2, column=1, value="No issues found — all clear.").alignment = CENTER
    ws2.freeze_panes = "A2"

    wb.save(output_path)


class ColumnPickerDialog(tk.Toplevel):
    """
    Dialog for picking the source text column (and sheet) from an Excel
    workbook before normalization or QC. Shows a preview of the first
    few rows of the selected column.
    """
    def __init__(self, parent, filepath):
        super().__init__(parent)
        self.filepath = filepath
        self.result = None   # (sheet_name, column_name) or None if cancelled
        self.title("Select Source Column")
        self.configure(bg=C["bg"])
        self.resizable(False, False)
        self.grab_set()

        import openpyxl
        self._wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        self._sheet_names = self._wb.sheetnames

        tk.Label(self, text=f"File: {Path(filepath).name}",
                 font=FONT_BOLD, bg=C["bg"], fg=C["text"]).pack(padx=20, pady=(16,4))

        sel = tk.Frame(self, bg=C["bg"])
        sel.pack(padx=20, pady=8, fill="x")

        tk.Label(sel, text="Sheet:", font=FONT, bg=C["bg"], fg=C["text"]
                 ).grid(row=0, column=0, sticky="w", pady=4)
        self._sheet_var = tk.StringVar(value=self._sheet_names[0])
        sheet_cb = ttk.Combobox(sel, textvariable=self._sheet_var,
                                 values=self._sheet_names, state="readonly",
                                 width=28, font=FONT)
        sheet_cb.grid(row=0, column=1, sticky="w", pady=4, padx=8)
        sheet_cb.bind("<<ComboboxSelected>>", lambda e: self._refresh_columns())

        tk.Label(sel, text="Source column:", font=FONT, bg=C["bg"], fg=C["text"]
                 ).grid(row=1, column=0, sticky="w", pady=4)
        self._col_var = tk.StringVar()
        self._col_cb = ttk.Combobox(sel, textvariable=self._col_var,
                                     state="readonly", width=28, font=FONT)
        self._col_cb.grid(row=1, column=1, sticky="w", pady=4, padx=8)
        self._col_cb.bind("<<ComboboxSelected>>", lambda e: self._refresh_preview())

        tk.Label(self, text="Preview (first 5 rows):", font=FONT_SMALL,
                 bg=C["bg"], fg=C["text_muted"]).pack(padx=20, anchor="w", pady=(8,2))
        self._preview = tk.Text(self, height=6, width=60, font=FONT_SMALL,
                                 bg=C["bg2"], fg=C["text"], relief="flat",
                                 highlightbackground=C["border"], highlightthickness=1,
                                 wrap="word", state="disabled")
        self._preview.pack(padx=20, pady=(0,10))

        btn_row = tk.Frame(self, bg=C["bg"])
        btn_row.pack(pady=(4,16))
        tk.Button(btn_row, text="Cancel", font=FONT, bg=C["bg2"], fg=C["text"],
                  relief="flat", cursor="hand2", command=self._cancel,
                  padx=16, pady=6).pack(side="left", padx=6)
        tk.Button(btn_row, text="Confirm", font=FONT_BOLD, bg=C["accent"],
                  fg=C["white"], relief="flat", activebackground=C["accent_h"],
                  cursor="hand2", command=self._confirm, padx=16, pady=6).pack(side="left", padx=6)

        self._refresh_columns()

    def _get_headers(self):
        ws = self._wb[self._sheet_var.get()]
        row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        return [str(c) if c is not None else f"Column {i+1}" for i, c in enumerate(row1)]

    def _refresh_columns(self):
        headers = self._get_headers()
        self._headers = headers
        self._col_cb.configure(values=headers)
        if headers:
            self._col_var.set(headers[0])
        self._refresh_preview()

    def _refresh_preview(self):
        if not self._col_var.get():
            return
        ws = self._wb[self._sheet_var.get()]
        headers = self._headers
        try:
            col_idx = headers.index(self._col_var.get())
        except ValueError:
            return
        preview_lines = []
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True)):
            if col_idx < len(row):
                val = row[col_idx]
                preview_lines.append(f"Row {i+2}: {val}")
        self._preview.config(state="normal")
        self._preview.delete("1.0", "end")
        self._preview.insert("1.0", "\n".join(preview_lines) or "(no data)")
        self._preview.config(state="disabled")

    def _cancel(self):
        self.result = None
        self._wb.close()
        self.destroy()

    def _confirm(self):
        self.result = (self._sheet_var.get(), self._col_var.get())
        self._wb.close()
        self.destroy()


class MultiFileSelectDialog(tk.Toplevel):
    """
    Dialog showing a checklist of files found in a folder (.txt and .xlsx
    mixed together), letting the user deselect any before processing.
    """
    def __init__(self, parent, files):
        super().__init__(parent)
        self.files = files
        self.result = None   # list of selected Path objects, or None if cancelled
        self.title("Select Files to Process")
        self.configure(bg=C["bg"])
        self.resizable(True, True)
        self.grab_set()

        tk.Label(self, text=f"{len(files)} file(s) found",
                 font=FONT_TITLE, bg=C["bg"], fg=C["text"]).pack(padx=20, pady=(16,4))
        tk.Label(self, text="Uncheck any files you don't want to process.",
                 font=FONT_SMALL, bg=C["bg"], fg=C["text_muted"]).pack(padx=20, pady=(0,8))

        outer = tk.Frame(self, bg=C["border"], bd=1)
        outer.pack(fill="both", expand=True, padx=20, pady=4)
        canvas = tk.Canvas(outer, bg=C["bg"], bd=0, highlightthickness=0,
                            height=320, width=480)
        vsb = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        inner = tk.Frame(canvas, bg=C["bg"])
        canvas.create_window((0,0), window=inner, anchor="nw")
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        self._vars = {}
        for i, fp in enumerate(files):
            bg = C["row_even"] if i%2==0 else C["row_odd"]
            row = tk.Frame(inner, bg=bg)
            row.pack(fill="x", padx=4, pady=1)
            v = tk.BooleanVar(value=True)
            self._vars[fp] = v
            tk.Checkbutton(row, text=f"{fp.name}  ({fp.suffix.lstrip('.').upper()})",
                           variable=v, font=FONT_SMALL, bg=bg, fg=C["text"],
                           activebackground=bg, selectcolor=C["accent_lt"],
                           relief="flat", cursor="hand2", anchor="w"
                           ).pack(fill="x", padx=4)

        sel_row = tk.Frame(self, bg=C["bg"])
        sel_row.pack(fill="x", padx=20, pady=(6,0))
        tk.Button(sel_row, text="Select All", font=FONT_SMALL, bg=C["bg2"],
                  fg=C["text"], relief="flat", cursor="hand2",
                  command=lambda: [v.set(True) for v in self._vars.values()]
                  ).pack(side="left", padx=2)
        tk.Button(sel_row, text="Select None", font=FONT_SMALL, bg=C["bg2"],
                  fg=C["text"], relief="flat", cursor="hand2",
                  command=lambda: [v.set(False) for v in self._vars.values()]
                  ).pack(side="left", padx=2)

        btn_row = tk.Frame(self, bg=C["bg"])
        btn_row.pack(pady=14)
        tk.Button(btn_row, text="Cancel", font=FONT, bg=C["bg2"], fg=C["text"],
                  relief="flat", cursor="hand2", command=self._cancel,
                  padx=16, pady=6).pack(side="left", padx=6)
        tk.Button(btn_row, text="Continue", font=FONT_BOLD, bg=C["accent"],
                  fg=C["white"], relief="flat", activebackground=C["accent_h"],
                  cursor="hand2", command=self._confirm, padx=16, pady=6).pack(side="left", padx=6)

    def _cancel(self):
        self.result = None
        self.destroy()

    def _confirm(self):
        self.result = [fp for fp, v in self._vars.items() if v.get()]
        self.destroy()


class ConflictDialog(tk.Toplevel):
    def __init__(self, parent, conflicts):
        super().__init__(parent)
        self.title("Dictionary Conflicts")
        self.configure(bg=C["bg"])
        self.resizable(False, False)
        self.grab_set()
        self.results = {}

        tk.Label(self, text="Dictionary conflicts detected",
                 font=FONT_TITLE, bg=C["bg"], fg=C["text"]).pack(padx=20, pady=(16,4))
        tk.Label(self, text="Choose which output to use for each conflicting token:",
                 font=FONT_SMALL, bg=C["bg"], fg=C["text_muted"]).pack(padx=20, pady=(0,10))

        frame = tk.Frame(self, bg=C["bg"])
        frame.pack(padx=20, fill="x")

        self._vars = {}
        for i, (token, std, custom) in enumerate(conflicts):
            bg = C["row_even"] if i%2==0 else C["row_odd"]
            row = tk.Frame(frame, bg=bg, pady=6, padx=8)
            row.pack(fill="x", pady=2)
            tk.Label(row, text=f"Token:  {token}", font=FONT_BOLD,
                     bg=bg, fg=C["text"]).grid(row=0, column=0, sticky="w")
            tk.Label(row, text=f"Standard rule → {std}", font=FONT_SMALL,
                     bg=bg, fg=C["text_muted"]).grid(row=1, column=0, sticky="w")
            tk.Label(row, text=f"Custom dict  → {custom}", font=FONT_SMALL,
                     bg=bg, fg=C["accent_h"]).grid(row=2, column=0, sticky="w")
            var = tk.StringVar(value="custom")
            self._vars[token] = var
            rf = tk.Frame(row, bg=bg)
            rf.grid(row=0, column=1, rowspan=3, padx=20)
            for val, lbl in [("standard","Use standard"),("custom","Use custom")]:
                tk.Radiobutton(rf, text=lbl, variable=var, value=val,
                               bg=bg, fg=C["text"], activebackground=bg,
                               selectcolor=C["accent_lt"],
                               font=FONT_SMALL).pack(anchor="w")

        tk.Frame(self, height=1, bg=C["border"]).pack(fill="x", padx=20, pady=8)
        self._apply_all = tk.StringVar(value="none")
        af = tk.Frame(self, bg=C["bg"])
        af.pack(padx=20, fill="x")
        tk.Label(af, text="Apply same choice to all:", font=FONT_SMALL,
                 bg=C["bg"], fg=C["text_muted"]).pack(side="left")
        for val, lbl in [("standard","All standard"),("custom","All custom"),("none","Individual")]:
            tk.Radiobutton(af, text=lbl, variable=self._apply_all, value=val,
                           bg=C["bg"], fg=C["text"], activebackground=C["bg"],
                           selectcolor=C["accent_lt"], font=FONT_SMALL,
                           command=self._apply_all_fn).pack(side="left", padx=6)

        tk.Button(self, text="Confirm", font=FONT_BOLD,
                  bg=C["accent"], fg=C["white"], relief="flat",
                  activebackground=C["accent_h"], cursor="hand2",
                  command=self._confirm, padx=20, pady=8).pack(pady=16)

    def _apply_all_fn(self):
        c = self._apply_all.get()
        if c in ("standard","custom"):
            for v in self._vars.values(): v.set(c)

    def _confirm(self):
        for token, var in self._vars.items():
            self.results[token] = var.get()
        self.destroy()


def export_qc_report_excel(file_reports, output_path, row_label="Line"):
    """
    Generate a three-sheet QC/normalization report Excel file.

    file_reports: list of dicts, one per processed file:
        {
          "filename": str,
          "needs_qc": bool,
          "comments": [str, ...],   # reasons why QC needed, or [] if clean
          "ambiguous": [ {"line": int, "original": str, "spoken": str, "comment": str} ],
          "spelling":  [ {"line": int, "token": str, "suggestion": str} ],
          "changes":   [ {"line": int, "category": str, "original": str, "converted": str} ],
        }
    row_label: "Line" for .txt sources, "Row" for Excel sources (header text only).

    Sheet 1 — All Files: every file with YES/NO Needs QC cell (green/red)
    Sheet 2 — Needs Attention: only files that need QC, with detail rows
    Sheet 3 — Changes: every individual change made, itemized per file/row
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    GREEN  = PatternFill("solid", fgColor="C6EFCE")
    RED    = PatternFill("solid", fgColor="FFC7CE")
    YELLOW = PatternFill("solid", fgColor="FFEB9C")
    HDR_FILL = PatternFill("solid", fgColor="2F4F6F")
    HDR_FONT = Font(color="FFFFFF", bold=True)
    BOLD   = Font(bold=True)
    CENTER = Alignment(horizontal="center", vertical="center")
    WRAP   = Alignment(wrap_text=True, vertical="top")
    thin   = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _hdr_row(ws, cols, row=1):
        for c, title in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=title)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = CENTER
            cell.border = BORDER

    def _set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "All Files"
    ws1.row_dimensions[1].height = 22
    _hdr_row(ws1, ["#", "File Name", "Needs QC", "Comments",
                   "Spelling Issues"])
    _set_col_widths(ws1, [5, 40, 12, 55, 40])

    for i, fr in enumerate(file_reports, 1):
        row = i + 1
        needs = fr["needs_qc"]
        comment_text = "\n".join(fr["comments"]) if fr["comments"] else "All clear"

        ws1.cell(row=row, column=1, value=i).alignment = CENTER
        ws1.cell(row=row, column=2, value=fr["filename"]).alignment = WRAP

        qc_cell = ws1.cell(row=row, column=3,
                           value="YES" if needs else "NO")
        qc_cell.fill  = GREEN if not needs else RED
        qc_cell.font  = Font(bold=True,
                             color="375623" if not needs else "9C0006")
        qc_cell.alignment = CENTER

        cm_cell = ws1.cell(row=row, column=4, value=comment_text)
        cm_cell.alignment = WRAP
        if fr.get("ambiguous"):
            cm_cell.fill = YELLOW

        spell_items = fr.get("spelling", [])
        spell_text  = "\n".join(
            f"Line {s['line']}: '{s['token']}' → '{s['suggestion']}'"
            for s in spell_items) if spell_items else "—"
        sp_cell = ws1.cell(row=row, column=5, value=spell_text)
        sp_cell.alignment = WRAP
        if spell_items:
            sp_cell.fill = YELLOW

        for c in range(1, 6):
            ws1.cell(row=row, column=c).border = BORDER

        ws1.row_dimensions[row].height = max(
            15, 15 * max(comment_text.count("\n") + 1,
                         spell_text.count("\n") + 1))

    ws1.freeze_panes = "A2"

    ws2 = wb.create_sheet("Needs Attention")
    ws2.row_dimensions[1].height = 22
    _hdr_row(ws2, ["#", "File Name", row_label, "Issue Type",
                   "Original", "Spoken / Note"])
    _set_col_widths(ws2, [5, 35, 7, 22, 22, 45])
    detail_row = 2
    att_num = 0
    for fr in file_reports:
        if not fr["needs_qc"]:
            continue
        att_num += 1

        for c in range(1, 7):
            cell = ws2.cell(row=detail_row, column=c)
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
            cell.border = BORDER
        ws2.cell(row=detail_row, column=1, value=att_num).alignment = CENTER
        ws2.cell(row=detail_row, column=2,
                 value=fr["filename"]).font = BOLD
        ws2.cell(row=detail_row, column=3, value="—").alignment = CENTER
        ws2.cell(row=detail_row, column=4, value="File flagged")
        ws2.cell(row=detail_row, column=5,
                 value="\n".join(fr["comments"])).alignment = WRAP
        ws2.row_dimensions[detail_row].height = max(
            15, 15 * (len(fr["comments"]) or 1))
        detail_row += 1

        for amb in fr.get("ambiguous", []):
            for c in range(1, 7):
                ws2.cell(row=detail_row, column=c).border = BORDER
            ws2.cell(row=detail_row, column=1, value="").alignment = CENTER
            ws2.cell(row=detail_row, column=2, value="")
            ws2.cell(row=detail_row, column=3,
                     value=amb["line"]).alignment = CENTER
            ws2.cell(row=detail_row, column=4,
                     value="Ambiguous security code")
            ws2.cell(row=detail_row, column=5,
                     value=amb["original"])
            note_cell = ws2.cell(row=detail_row, column=6,
                                 value=amb["comment"])
            note_cell.fill = YELLOW
            note_cell.alignment = WRAP
            ws2.row_dimensions[detail_row].height = 18
            detail_row += 1

        for sp in fr.get("spelling", []):
            for c in range(1, 7):
                ws2.cell(row=detail_row, column=c).border = BORDER
            ws2.cell(row=detail_row, column=3,
                     value=sp["line"]).alignment = CENTER
            ws2.cell(row=detail_row, column=4,
                     value="Spelling issue")
            ws2.cell(row=detail_row, column=5,
                     value=sp["token"])
            sugg_cell = ws2.cell(row=detail_row, column=6,
                                 value=f"Suggested: {sp['suggestion']}")
            sugg_cell.fill = YELLOW
            sugg_cell.alignment = WRAP
            ws2.row_dimensions[detail_row].height = 18
            detail_row += 1

    if detail_row == 2:
        ws2.cell(row=2, column=1,
                 value="No files need attention — all clear.").alignment = CENTER

    ws2.freeze_panes = "A2"

    ws3 = wb.create_sheet("Changes")
    ws3.row_dimensions[1].height = 22
    _hdr_row(ws3, ["#", "File Name", row_label, "Category", "Original", "Converted To"])
    _set_col_widths(ws3, [5, 35, 8, 18, 35, 35])

    CATEGORY_LABELS = {
        "corrections":   "Grammar Correction",
        "chemicals":     "Chemical Formula",
        "numbers":       "Number",
        "symbols":       "Symbol",
        "time":          "Time",
        "temperature":   "Temperature",
        "units":         "Unit",
        "diacritics":    "Diacritic",
        "abbreviations": "Abbreviation",
    }

    change_row = 2
    change_num = 0
    for fr in file_reports:
        changes = fr.get("changes", [])
        if not changes:
            continue
        for c in range(1, 7):
            cell = ws3.cell(row=change_row, column=c)
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
            cell.border = BORDER
        change_num += 1
        ws3.cell(row=change_row, column=1, value=change_num).alignment = CENTER
        ws3.cell(row=change_row, column=2, value=fr["filename"]).font = BOLD
        ws3.cell(row=change_row, column=4,
                 value=f"{len(changes)} change(s)").font = BOLD
        change_row += 1

        for ch in sorted(changes, key=lambda c: (c.get("line", 0), c.get("category",""))):
            for c in range(1, 7):
                ws3.cell(row=change_row, column=c).border = BORDER
            ws3.cell(row=change_row, column=3,
                     value=ch.get("line", "—")).alignment = CENTER
            ws3.cell(row=change_row, column=4,
                     value=CATEGORY_LABELS.get(ch.get("category",""), ch.get("category","")))
            orig_cell = ws3.cell(row=change_row, column=5, value=ch.get("original",""))
            orig_cell.alignment = WRAP
            conv_cell = ws3.cell(row=change_row, column=6, value=ch.get("converted",""))
            conv_cell.alignment = WRAP
            change_row += 1

    if change_row == 2:
        ws3.cell(row=2, column=1, value="No changes recorded.").alignment = CENTER

    ws3.freeze_panes = "A2"

    wb.save(output_path)


class NormalizerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Chichingpoipoi v2.0")
        try:
            self.iconbitmap(resource_path("chichingpoipoi.ico"))
        except Exception:
            pass
        self.configure(bg=C["bg"])
        self._center_window(980, 800)
        self.minsize(820, 620)

        self.custom_dict: dict = {}
        self.custom_dict_file = Path(os.environ.get("APPDATA", Path.home())) / "Chichingpoipoi" / "custom_dict.json"
        self.custom_dict_file.parent.mkdir(parents=True, exist_ok=True)
        self._load_custom_dict()
        self.batch_files: list = []
        self.last_norm_report: dict = {}
        self.last_qc_issues: list = []
        self._last_unverified: list = []
        self._chem_decisions: dict = {}
        self._preload_scan_done: bool = False
        self._excel_source = None

        self._build_ui()
        self.after(200, self._show_thankyou)

    def _center_window(self, width, height):
        """Centre this Tk window on the current screen."""
        self.update_idletasks()
        screen_w = self.winfo_screenwidth()
        screen_h = self.winfo_screenheight()
        x = (screen_w - width) // 2
        y = (screen_h - height) // 2
        self.geometry(f"{width}x{height}+{x}+{y}")

    def _show_thankyou(self):
        """Show a Thank You splash window centred over the main window,
        with no native title bar (just an OK button to dismiss)."""
        dlg = tk.Toplevel(self)
        dlg.configure(bg=C["accent"])
        dlg.resizable(False, False)
        dlg.overrideredirect(True)   # remove native title bar / min/max/close buttons
        dlg.grab_set()

        dw, dh = 340, 440
        self.update_idletasks()
        mx, my = self.winfo_x(), self.winfo_y()
        mw, mh = self.winfo_width(), self.winfo_height()
        dx = mx + (mw - dw) // 2
        dy = my + (mh - dh) // 2
        dlg.geometry(f"{dw}x{dh}+{dx}+{dy}")

        content = tk.Frame(dlg, bg=C["bg"])
        content.pack(fill="both", expand=True, padx=2, pady=2)

        THANKS_NAMES = [
            "Kiran Mukhiya", "Joe Jacintha", "Liz Grace John",
            "Harika", "Shushanth P.", "Sandesh Koirala", "Yathaarth M.P.",
        ]

        inner = tk.Frame(content, bg=C["bg"])
        inner.pack(expand=True, fill="both", padx=30, pady=22)

        tk.Label(inner, text="Chichingpoipoi", font=("Segoe UI", 16, "bold"),
                 bg=C["bg"], fg=C["accent"], justify="center").pack(pady=(0, 2))
        tk.Label(inner, text="v2.0", font=FONT_SMALL,
                 bg=C["bg"], fg=C["text_muted"], justify="center").pack()

        tk.Frame(inner, height=1, bg=C["border"]).pack(fill="x", pady=14)

        tk.Label(inner, text="Special Thanks To These Guys and Gals",
                 font=FONT_BOLD, bg=C["bg"], fg=C["text"],
                 justify="center").pack()

        tk.Frame(inner, height=6, bg=C["bg"]).pack()

        for name in THANKS_NAMES:
            tk.Label(inner, text=name, font=FONT,
                     bg=C["bg"], fg=C["text_muted"],
                     justify="center").pack(pady=1)

        tk.Frame(inner, height=1, bg=C["border"]).pack(fill="x", pady=14)

        tk.Label(inner,
                 text="Created with Love and Designed with Care\n"
                      "© Priyangshu Swarnakar",
                 font=FONT_SMALL, bg=C["bg"], fg=C["text_muted"],
                 justify="center").pack(pady=(0, 10))

        tk.Button(inner, text="OK", font=FONT_BOLD,
                  bg=C["accent"], fg=C["white"], relief="flat",
                  activebackground=C["accent_h"], cursor="hand2",
                  command=dlg.destroy, padx=24, pady=6).pack(pady=(2, 0))

    def _load_custom_dict(self):
        if self.custom_dict_file.exists():
            try:
                self.custom_dict = json.loads(
                    self.custom_dict_file.read_text(encoding="utf-8"))
            except Exception:
                self.custom_dict = {}

    def _save_custom_dict(self):
        self.custom_dict_file.write_text(
            json.dumps(self.custom_dict, ensure_ascii=False, indent=2),
            encoding="utf-8")

    def _build_ui(self):
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("TNotebook", background=C["bg"], borderwidth=0)
        style.configure("TNotebook.Tab", background=C["bg2"],
                        foreground=C["text_muted"], font=FONT_BOLD, padding=[14,6])
        style.map("TNotebook.Tab",
                  background=[("selected", C["accent"])],
                  foreground=[("selected", C["white"])])
        style.configure("TFrame", background=C["bg"])

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=12, pady=10)

        self.tab_norm   = tk.Frame(nb, bg=C["bg"])
        self.tab_qc     = tk.Frame(nb, bg=C["bg"])
        self.tab_dict   = tk.Frame(nb, bg=C["bg"])
        self.tab_about  = tk.Frame(nb, bg=C["bg"])

        nb.add(self.tab_norm,   text="  Normalizer  ")
        nb.add(self.tab_qc,     text="  QC Tool  ")
        nb.add(self.tab_dict,   text="  Custom Dictionary  ")
        nb.add(self.tab_about,  text="  About  ")

        self._build_normalizer_tab()
        self._build_qc_tab()
        self._build_dict_tab()
        self._build_about_tab()

        nb.bind("<<NotebookTabChanged>>", self._on_tab_changed)

    def _on_tab_changed(self, event):
        if hasattr(self, "_about_recentre"):
            self.after(10, self._about_recentre)

    def _label(self, parent, text, bold=False, muted=False):
        return tk.Label(parent, text=text,
                        font=FONT_BOLD if bold else FONT,
                        bg=C["bg"], fg=C["text_muted"] if muted else C["text"])

    def _btn(self, parent, text, cmd, accent=False):
        bg  = C["accent"] if accent else C["bg2"]
        fg  = C["white"]  if accent else C["text"]
        abg = C["accent_h"] if accent else C["border"]
        return tk.Button(parent, text=text, command=cmd,
                         font=FONT_BOLD if accent else FONT,
                         bg=bg, fg=fg, relief="flat",
                         activebackground=abg, cursor="hand2",
                         padx=12, pady=5)

    def _text_area(self, parent, height=10, readonly=False):
        frame = tk.Frame(parent, bg=C["border"], bd=1)
        t = scrolledtext.ScrolledText(frame, height=height,
                                       font=("Consolas", 10),
                                       bg=C["white"], fg=C["text"],
                                       relief="flat", wrap="word",
                                       insertbackground=C["accent"],
                                       selectbackground=C["accent_lt"])
        if readonly: t.config(state="disabled")
        t.pack(fill="both", expand=True, padx=1, pady=1)
        return frame, t


    def _build_normalizer_tab(self):
        p = self.tab_norm
        p.columnconfigure(0, weight=1)
        p.rowconfigure(1, weight=2)
        p.rowconfigure(7, weight=2)

        top = tk.Frame(p, bg=C["bg"])
        top.grid(row=0, column=0, sticky="ew", pady=(6,2), padx=10)
        self._label(top, "Input Text", bold=True).pack(side="left")
        self._btn(top, "Open File",   self._open_file).pack(side="right", padx=4)
        self._btn(top, "Open Folder", self._open_folder).pack(side="right", padx=4)
        self._btn(top, "Clear",       self._clear_input).pack(side="right", padx=4)

        self.batch_status = tk.Label(p, text="", font=FONT_SMALL,
                                     bg=C["bg"], fg=C["accent_h"])
        self.batch_status.grid(row=0, column=0, sticky="e", padx=10)

        in_frame, self.input_text = self._text_area(p, height=9)
        in_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=2)

        self.wc_label = tk.Label(p, text="", font=FONT_SMALL,
                                  bg=C["bg"], fg=C["text_muted"])
        self.wc_label.grid(row=2, column=0, sticky="e", padx=12)
        self.input_text.bind("<KeyRelease>", self._update_wc)

        ctrl = tk.LabelFrame(p, text="  Normalization Rules  ",
                              font=FONT_BOLD, bg=C["bg"], fg=C["text"],
                              relief="flat", bd=1,
                              highlightbackground=C["border"], highlightthickness=1)
        ctrl.grid(row=3, column=0, sticky="ew", padx=10, pady=4)

        self._opt = {}
        opts = [
            ("numbers",       "Convert numbers to words"),
            ("symbols",       "Convert symbols to words"),
            ("abbreviations", "Expand abbreviations / acronyms"),
            ("diacritics",    "Remove diacritics  (é → e)"),
            ("punctuation",   "Strip punctuation marks"),
            ("spacing",       "Fix spacing"),
        ]
        for i, (key, label) in enumerate(opts):
            v = tk.BooleanVar(value=True)
            self._opt[key] = v
            tk.Checkbutton(ctrl, text=label, variable=v,
                           font=FONT, bg=C["bg"], fg=C["text"],
                           activebackground=C["bg"], selectcolor=C["accent_lt"],
                           relief="flat", cursor="hand2"
                           ).grid(row=i//3, column=i%3, sticky="w", padx=14, pady=3)

        self._opt["expand_currency"] = tk.BooleanVar(value=False)
        tk.Checkbutton(ctrl, text="Expand currency codes (INR → Indian Rupees)",
                       variable=self._opt["expand_currency"],
                       font=FONT, bg=C["bg"], fg=C["text"],
                       activebackground=C["bg"], selectcolor=C["accent_lt"],
                       relief="flat", cursor="hand2"
                       ).grid(row=2, column=0, columnspan=2, sticky="w", padx=14, pady=3)

        loc_frame = tk.Frame(ctrl, bg=C["bg"])
        loc_frame.grid(row=2, column=2, sticky="e", padx=14, pady=3)
        self._label(loc_frame, "Locale:", muted=True).pack(side="left", padx=(0,4))
        self._locale_var = tk.StringVar(value="India")
        ttk.Combobox(loc_frame, textvariable=self._locale_var,
                     values=["India","US","UK"],
                     state="readonly", width=8, font=FONT_SMALL).pack(side="left")

        bot = tk.Frame(p, bg=C["bg"])
        bot.grid(row=4, column=0, sticky="ew", padx=10, pady=4)

        px_frame = tk.LabelFrame(p, text="  Partial Expansion  ",
                                  font=FONT_BOLD, bg=C["bg"], fg=C["text"],
                                  relief="flat", bd=1,
                                  highlightbackground=C["border"], highlightthickness=1)
        px_frame.grid(row=4, column=0, sticky="ew", padx=10, pady=(0,4))
        px_frame.columnconfigure(0, weight=1)

        _PX_DEFAULTS = [
            ("GB",   "gigabytes",              "G B"),
            ("KB",   "kilobytes",              "K B"),
            ("MB",   "megabytes",              "M B"),
            ("Gbps", "gigabits per second",    "Gbps"),
            ("Mbps", "megabits per second",    "Mbps"),
            ("Kbps", "kilobits per second",    "Kbps"),
            ("KG",   "kilograms",              "K G"),
        ]
        self._px_vars   = {}   # token → BooleanVar (enabled)
        self._px_thresh = {}   # token → IntVar (threshold %)

        tok_row = tk.Frame(px_frame, bg=C["bg"])
        tok_row.pack(fill="x", padx=8, pady=(4,2))

        for i, (tok, expand, letters) in enumerate(_PX_DEFAULTS):
            col_frame = tk.Frame(tok_row, bg=C["bg"])
            col_frame.grid(row=0, column=i, padx=6, pady=2, sticky="w")

            en_var = tk.BooleanVar(value=False)
            th_var = tk.IntVar(value=50)
            self._px_vars[tok]   = (en_var, expand, letters)
            self._px_thresh[tok] = th_var

            tk.Checkbutton(col_frame, text=tok, variable=en_var,
                           font=FONT_BOLD, bg=C["bg"], fg=C["text"],
                           activebackground=C["bg"], selectcolor=C["accent_lt"],
                           relief="flat", cursor="hand2").pack(anchor="w")
            th_frame = tk.Frame(col_frame, bg=C["bg"])
            th_frame.pack(anchor="w")
            tk.Label(th_frame, text="Expand %:", font=FONT_SMALL,
                     bg=C["bg"], fg=C["text_muted"]).pack(side="left")
            tk.Spinbox(th_frame, from_=0, to=100, textvariable=th_var,
                       width=4, font=FONT_SMALL).pack(side="left")

        cust_row = tk.Frame(px_frame, bg=C["bg"])
        cust_row.pack(fill="x", padx=8, pady=(2,6))
        tk.Label(cust_row, text="Custom token:", font=FONT_SMALL,
                 bg=C["bg"], fg=C["text_muted"]).pack(side="left")
        self._px_custom_tok  = tk.StringVar()
        self._px_custom_exp  = tk.StringVar()
        self._px_custom_pct  = tk.IntVar(value=50)
        tk.Entry(cust_row, textvariable=self._px_custom_tok,
                 font=FONT_SMALL, width=10,
                 bg=C["bg"], fg=C["text"], relief="flat",
                 highlightbackground=C["border"],
                 highlightthickness=1).pack(side="left", padx=4)
        tk.Label(cust_row, text="Expand to:", font=FONT_SMALL,
                 bg=C["bg"], fg=C["text_muted"]).pack(side="left")
        tk.Entry(cust_row, textvariable=self._px_custom_exp,
                 font=FONT_SMALL, width=18,
                 bg=C["bg"], fg=C["text"], relief="flat",
                 highlightbackground=C["border"],
                 highlightthickness=1).pack(side="left", padx=4)
        tk.Label(cust_row, text="% expand:", font=FONT_SMALL,
                 bg=C["bg"], fg=C["text_muted"]).pack(side="left")
        tk.Spinbox(cust_row, from_=0, to=100, textvariable=self._px_custom_pct,
                   width=4, font=FONT_SMALL).pack(side="left", padx=2)

        self._spell_check_var = tk.BooleanVar(value=False)
        tk.Checkbutton(cust_row, text="Enable spell-check",
                       variable=self._spell_check_var,
                       font=FONT_SMALL, bg=C["bg"], fg=C["text"],
                       activebackground=C["bg"], selectcolor=C["accent_lt"],
                       relief="flat", cursor="hand2").pack(side="right", padx=8)


        bot.grid(row=5, column=0, sticky="ew", padx=10, pady=4)
        self._label(bot, "Output Case:", bold=True).pack(side="left")
        self._case_var = tk.StringVar(value="UPPERCASE")
        for mode in ["UPPERCASE","lowercase","Sentence case","Title Case","As-is"]:
            tk.Radiobutton(bot, text=mode, variable=self._case_var, value=mode,
                           font=FONT_SMALL, bg=C["bg"], fg=C["text"],
                           activebackground=C["bg"], selectcolor=C["accent_lt"],
                           cursor="hand2").pack(side="left", padx=8)
        self._btn(bot, "⚡  Normalize",     self._run_normalize, accent=True).pack(side="right", padx=4)

        out_lbl = tk.Frame(p, bg=C["bg"])
        out_lbl.grid(row=6, column=0, sticky="w", padx=10)
        self._label(out_lbl, "Output", bold=True).pack(side="left")

        out_frame, self.output_text = self._text_area(p, height=9, readonly=True)
        out_frame.grid(row=7, column=0, sticky="nsew", padx=10, pady=2)

        btn_row = tk.Frame(p, bg=C["bg"])
        btn_row.grid(row=8, column=0, sticky="e", padx=10, pady=(4,8))
        self._btn(btn_row, "Copy to Clipboard", self._copy_output).pack(side="right", padx=4)
        self._btn(btn_row, "Save Output",       self._save_output).pack(side="right", padx=4)


    def _build_qc_tab(self):
        p = self.tab_qc
        p.columnconfigure(0, weight=1)
        p.rowconfigure(3, weight=1)

        hdr = tk.Frame(p, bg=C["bg"])
        hdr.grid(row=0, column=0, sticky="ew", padx=12, pady=(10,4))
        self._label(hdr, "QC Tool", bold=True).pack(side="left")
        self._btn(hdr, "Load File",   self._qc_load_file).pack(side="right", padx=4)
        self._btn(hdr, "Load Folder", self._qc_load_folder).pack(side="right", padx=4)

        self.qc_status = tk.Label(p, text="No files loaded.", font=FONT_SMALL,
                                   bg=C["bg"], fg=C["text_muted"])
        self.qc_status.grid(row=0, column=0, sticky="e", padx=12)

        opt_row = tk.LabelFrame(p, text="  QC Options  ",
                                font=FONT_BOLD, bg=C["bg"], fg=C["text"],
                                relief="flat", bd=1,
                                highlightbackground=C["border"], highlightthickness=1)
        opt_row.grid(row=1, column=0, sticky="ew", padx=12, pady=4)

        self._qc_opts = {}
        for i, (key, label) in enumerate([
            ("numbers","Lingering numbers"),("symbols","Lingering symbols"),
            ("diacritics","Lingering diacritics"),("spacing","Spacing issues"),
            ("abbreviations","Unexpanded abbreviations"),
            ("format_errors","Format errors"),
            ("consistency","Cross-row consistency checks"),
            ("spell_check","Spell-check"),
        ]):
            v = tk.BooleanVar(value=True if key != "spell_check" else False)
            self._qc_opts[key] = v
            tk.Checkbutton(opt_row, text=label, variable=v,
                           font=FONT, bg=C["bg"], fg=C["text"],
                           activebackground=C["bg"], selectcolor=C["accent_lt"],
                           relief="flat", cursor="hand2"
                           ).grid(row=i//3, column=i%3, sticky="w", padx=12, pady=4)

        case_row = tk.Frame(p, bg=C["bg"])
        case_row.grid(row=2, column=0, sticky="ew", padx=12, pady=(2,4))
        self._qc_check_case = tk.BooleanVar(value=False)
        tk.Checkbutton(case_row, text="Check casing consistency",
                       variable=self._qc_check_case,
                       font=FONT, bg=C["bg"], fg=C["text"],
                       activebackground=C["bg"], selectcolor=C["accent_lt"],
                       relief="flat", cursor="hand2").pack(side="left", padx=4)
        self._label(case_row, "Expected case:", muted=True).pack(side="left", padx=(16,4))
        self._qc_case_var = tk.StringVar(value="UPPERCASE")
        ttk.Combobox(case_row, textvariable=self._qc_case_var,
                     values=["UPPERCASE","lowercase","Sentence case","Title Case"],
                     state="readonly", width=16, font=FONT).pack(side="left", padx=4)

        run_row = tk.Frame(p, bg=C["bg"])
        run_row.grid(row=2, column=0, sticky="e", padx=12)
        self._btn(run_row, "▶  Run QC", self._run_qc, accent=True).pack(side="right", padx=4)

        rf, self.qc_result_text = self._text_area(p, height=22, readonly=True)
        rf.grid(row=3, column=0, sticky="nsew", padx=12, pady=4)
        self.qc_result_text.config(state="normal")
        self.qc_result_text.insert("1.0", "Load a file or folder and click Run QC.")
        self.qc_result_text.config(state="disabled")

        exp_row = tk.Frame(p, bg=C["bg"])
        exp_row.grid(row=4, column=0, sticky="e", padx=12, pady=(4,8))
        self._btn(exp_row, "Export QC Report (Excel)", self._export_qc_excel, accent=True).pack(side="right", padx=4)

        self._qc_files: list = []
        self._qc_report_text = ""


    def _build_dict_tab(self):
        p = self.tab_dict
        p.columnconfigure(0, weight=1)
        p.rowconfigure(1, weight=1)

        hdr = tk.Frame(p, bg=C["bg"])
        hdr.grid(row=0, column=0, sticky="ew", padx=12, pady=(10,4))
        self._label(hdr, "Custom Dictionary", bold=True).pack(side="left")
        self._btn(hdr, "+ Add Entry", self._add_dict_entry, accent=True).pack(side="right")
        tk.Label(hdr, text="Entries override standard rules when confirmed in the conflict dialog.",
                 font=FONT_SMALL, bg=C["bg"], fg=C["text_muted"]).pack(side="left", padx=14)

        outer = tk.Frame(p, bg=C["border"], bd=1)
        outer.grid(row=1, column=0, sticky="nsew", padx=12, pady=4)

        canvas = tk.Canvas(outer, bg=C["bg"], bd=0, highlightthickness=0)
        vsb = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        self.dict_inner = tk.Frame(canvas, bg=C["bg"])
        self.dict_window = canvas.create_window((0,0), window=self.dict_inner, anchor="nw")
        self.dict_inner.bind("<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>",
            lambda e: canvas.itemconfig(self.dict_window, width=e.width))
        self.dict_canvas = canvas

        hrow = tk.Frame(self.dict_inner, bg=C["bg2"])
        hrow.pack(fill="x", padx=4, pady=(4,0))
        tk.Label(hrow, text="Original Token", font=FONT_BOLD,
                 bg=C["bg2"], fg=C["text"], width=22, anchor="w").pack(side="left", padx=8)
        tk.Label(hrow, text="→", font=FONT_BOLD,
                 bg=C["bg2"], fg=C["accent"], width=3).pack(side="left")
        tk.Label(hrow, text="Normalized Form", font=FONT_BOLD,
                 bg=C["bg2"], fg=C["text"], width=28, anchor="w").pack(side="left", padx=8)
        tk.Label(hrow, text="Actions", font=FONT_BOLD,
                 bg=C["bg2"], fg=C["text"]).pack(side="left", padx=20)

        self._dict_rows = []
        self._refresh_dict_rows()
        self._btn(p, "💾  Save Dictionary", self._save_dict_from_ui,
                  accent=True).grid(row=2, column=0, sticky="e", padx=12, pady=8)

    def _refresh_dict_rows(self):
        for w in self._dict_rows: w[0].destroy()
        self._dict_rows.clear()
        for key, val in sorted(self.custom_dict.items()):
            self._add_dict_row(key, val)

    def _add_dict_row(self, key="", val=""):
        i = len(self._dict_rows)
        bg = C["row_even"] if i%2==0 else C["row_odd"]
        row = tk.Frame(self.dict_inner, bg=bg)
        row.pack(fill="x", padx=4, pady=1)
        kv, vv = tk.StringVar(value=key), tk.StringVar(value=val)
        for sv, w in [(kv,22),(vv,28)]:
            tk.Entry(row, textvariable=sv, font=FONT, bg=C["white"], fg=C["text"],
                     relief="flat", insertbackground=C["accent"], width=w,
                     highlightbackground=C["border"],
                     highlightthickness=1).pack(side="left", padx=8, pady=4)
            if sv == kv:
                tk.Label(row, text="→", font=FONT_BOLD,
                         bg=bg, fg=C["accent"], width=3).pack(side="left")
        def _del(r=row, k=kv):
            self.custom_dict.pop(k.get().upper(), None)
            self._save_custom_dict()
            r.destroy()
            self._dict_rows = [x for x in self._dict_rows if x[0].winfo_exists()]
        tk.Button(row, text="Delete", font=FONT_SMALL,
                  bg=C["bg2"], fg=C["red"], relief="flat",
                  activebackground=C["border"], cursor="hand2",
                  command=_del).pack(side="left", padx=4)
        self._dict_rows.append((row, kv, vv))

    def _add_dict_entry(self): self._add_dict_row()

    def _save_dict_from_ui(self):
        self.custom_dict = {}
        for row, kv, vv in self._dict_rows:
            k, v = kv.get().strip().upper(), vv.get().strip()
            if k and v: self.custom_dict[k] = v
        self._save_custom_dict()
        messagebox.showinfo("Saved", "Custom dictionary saved.")
        self._refresh_dict_rows()



    def _build_about_tab(self):
        p = self.tab_about
        p.columnconfigure(0, weight=1)
        p.rowconfigure(0, weight=1)

        outer = tk.Frame(p, bg=C["bg"])
        outer.grid(row=0, column=0, sticky="nsew")
        outer.columnconfigure(0, weight=1)
        outer.rowconfigure(0, weight=1)

        canvas = tk.Canvas(outer, bg=C["bg"], bd=0, highlightthickness=0)
        vsb = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")

        content = tk.Frame(canvas, bg=C["bg"])
        content_id = canvas.create_window(0, 0, window=content, anchor="n")

        def _recentre(event=None):
            canvas.update_idletasks()
            cw = canvas.winfo_width()
            canvas.coords(content_id, cw // 2, 0)
            canvas.configure(scrollregion=canvas.bbox("all"))

        self._about_recentre = _recentre
        content.bind("<Configure>", _recentre)
        canvas.bind("<Configure>", _recentre)
        self.after(100, _recentre)

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        tk.Label(content, text="Chichingpoipoi", font=("Segoe UI", 22, "bold"),
                 bg=C["bg"], fg=C["accent"]).pack(pady=(30, 2))
        tk.Label(content, text="v2.0  ·  Text Normalization and QC Tool",
                 font=FONT, bg=C["bg"], fg=C["text_muted"]).pack()

        tk.Frame(content, height=1, bg=C["border"]).pack(fill="x", padx=60, pady=16)

        guide = (
            "Normalizer Tab:\n"
            "  Paste text or load a file / folder. Toggle rules, pick locale and output case,\n"
            "  click Normalize. Folders output to a 'normalized/' subfolder.\n\n"
            "Normalization pipeline:\n"
            "  IP addresses → Ratios → Code numbers → Flight numbers → Product codes\n"
            "  → Time → Temperature → Units → Custom Dictionary → Abbreviations\n"
            "  → Symbols → Numbers → Diacritics → Punctuation → Spacing → Case\n\n"
            "QC Tool Tab:\n"
            "  Checks normalized files for lingering issues.\n\n"
            "Custom Dictionary:\n"
            "  Saved to custom_dict.json next to this executable.\n\n"
            "UK spellings used throughout (kilometre, litre, centre …)."
        )
        tk.Label(content, text=guide, font=FONT, bg=C["bg"], fg=C["text"],
                 justify="left").pack(padx=60, anchor="w")

        tk.Frame(content, height=1, bg=C["border"]).pack(fill="x", padx=60, pady=16)

        THANKS_NAMES = [
            "Kiran Mukhiya", "Joe Jacintha", "Liz Grace John",
            "Harika", "Shushanth P.", "Sandesh Koirala", "Yathaarth M.P.",
        ]
        tk.Label(content, text="Special Thanks To These Guys and Gals",
                 font=FONT_BOLD, bg=C["bg"], fg=C["text"]).pack()
        for name in THANKS_NAMES:
            tk.Label(content, text=name, font=FONT, bg=C["bg"],
                     fg=C["text_muted"]).pack()

        tk.Frame(content, height=1, bg=C["border"]).pack(fill="x", padx=60, pady=16)

        tk.Label(content,
                 text="Created with Love and Designed with Care\n"
                      "All Rights Reserved © Priyangshu Swarnakar",
                 font=FONT_SMALL, bg=C["bg"], fg=C["text_muted"],
                 justify="center").pack(pady=(0, 30))


    def _update_wc(self, _=None):
        txt = self.input_text.get("1.0","end-1c")
        self.wc_label.config(text=f"{len(txt.split())} words  ·  {len(txt)} chars")
        self._preload_scan_done = False

    def _open_file(self):
        path = filedialog.askopenfilename(
            filetypes=[("Supported files","*.txt *.xlsx"),
                       ("Text files","*.txt"),
                       ("Excel files","*.xlsx"),
                       ("All files","*.*")])
        if not path:
            return
        path = Path(path)
        self.batch_files = []
        self.batch_status.config(text="")

        if path.suffix.lower() == ".xlsx":
            dlg = ColumnPickerDialog(self, path)
            self.wait_window(dlg)
            if not dlg.result:
                return
            sheet_name, column_name = dlg.result
            rows, col_idx, header = read_excel_column(path, sheet_name, column_name)
            content = "\n".join(text for _, text in rows)
            self._excel_source = {
                "path": path, "sheet": sheet_name, "column": column_name,
                "rows": rows,
            }
            self.input_text.config(state="normal")
            self.input_text.delete("1.0","end")
            self.input_text.insert("1.0", content)
            self._update_wc()
            self._run_chem_detection(content, row_label="Row")
        else:
            self._excel_source = None
            content = path.read_text(encoding="utf-8", errors="replace")
            self.input_text.config(state="normal")
            self.input_text.delete("1.0","end")
            self.input_text.insert("1.0", content)
            self._update_wc()
            self._run_chem_detection(content, row_label="Line")

    def _open_folder(self):
        folder = filedialog.askdirectory()
        if not folder:
            return
        all_files = sorted(Path(folder).glob("*.txt")) + sorted(Path(folder).glob("*.xlsx"))
        if not all_files:
            messagebox.showwarning("No files","No .txt or .xlsx files found.")
            return

        dlg = MultiFileSelectDialog(self, all_files)
        self.wait_window(dlg)
        if dlg.result is None or not dlg.result:
            return
        selected = dlg.result

        batch_entries = []
        header_cache = {}   # tuple(header) -> (sheet_name, column_name)
        for fp in selected:
            if fp.suffix.lower() == ".txt":
                batch_entries.append({"path": fp, "type": "txt"})
                continue
            import openpyxl
            wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
            sheet_name = wb.sheetnames[0]
            ws = wb[sheet_name]
            header = tuple(str(c) if c is not None else f"Column {i+1}"
                           for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())))
            wb.close()

            if header in header_cache:
                sheet_used, col_used = header_cache[header]
            else:
                pdlg = ColumnPickerDialog(self, fp)
                self.wait_window(pdlg)
                if not pdlg.result:
                    continue   # skip this file if cancelled
                sheet_used, col_used = pdlg.result
                header_cache[header] = (sheet_used, col_used)

            batch_entries.append({
                "path": fp, "type": "xlsx",
                "sheet": sheet_used, "column": col_used,
            })

        if not batch_entries:
            return

        self.batch_files = batch_entries
        n_txt  = sum(1 for e in batch_entries if e["type"] == "txt")
        n_xlsx = sum(1 for e in batch_entries if e["type"] == "xlsx")
        self.batch_status.config(
            text=f"{n_txt} .txt + {n_xlsx} .xlsx file(s) loaded")
        self.input_text.config(state="normal")
        self.input_text.delete("1.0","end")
        self.input_text.insert("1.0", "\n".join(e["path"].name for e in batch_entries))
        self.input_text.config(state="disabled")

        combined_parts = []
        for e in batch_entries:
            if e["type"] == "txt":
                combined_parts.append(e["path"].read_text(encoding="utf-8", errors="replace"))
            else:
                rows, _, _ = read_excel_column(e["path"], e["sheet"], e["column"])
                combined_parts.append("\n".join(text for _, text in rows))
        combined = "\n".join(combined_parts)
        row_label = "Row" if n_xlsx else "Line"
        self._run_chem_detection(combined, row_label=row_label)

    def _clear_input(self):
        self.batch_files = []
        self.batch_status.config(text="")
        self._chem_decisions = {}
        self._preload_scan_done = False
        self._excel_source = None
        self.input_text.config(state="normal")
        self.input_text.delete("1.0","end")
        self.output_text.config(state="normal")
        self.output_text.delete("1.0","end")
        self.output_text.config(state="disabled")
        self.wc_label.config(text="")

    def _run_chem_detection(self, text, wait=False, row_label="Line"):
        """
        Run the full pre-load detection scan (chemicals, units, currency,
        flight/reference codes, code/PIN/OTP numbers). If anything is
        found, show PreloadDialog. wait=True: blocks until dialog closed
        (used for manual paste before normalize). row_label switches the
        review-tab header between "Line" (.txt sources) and "Row" (Excel
        sources, where the number reflects the spreadsheet row).
        """
        formulas      = detect_chemicals(text)
        formula_texts = {item["text"] for item in formulas}
        units         = detect_units_in_text(text)
        currency      = detect_currency_in_text(text)
        flight_ref    = detect_flight_ref_ids(text, exclude=formula_texts)
        code_pins     = detect_code_pins(text)
        abbreviations = detect_abbreviations_in_text(text)

        self._preload_scan_done = True

        if not (formulas or units or currency or flight_ref or code_pins or abbreviations):
            self._chem_decisions = {}
            return

        dlg = PreloadDialog(self, formulas=formulas, units=units,
                             currency=currency, flight_ref=flight_ref,
                             code_pins=code_pins, abbreviations=abbreviations,
                             row_label=row_label)
        self.wait_window(dlg)
        self._chem_decisions = dlg.decisions

    def _get_options(self):
        opts = {k: v.get() for k, v in self._opt.items()}
        opts["case"]        = self._case_var.get()
        opts["locale"]      = self._locale_var.get()
        opts["spell_check"] = self._spell_check_var.get()

        px_config = {}
        for tok, (en_var, expand, letters) in self._px_vars.items():
            if en_var.get():
                px_config[tok] = {
                    "expand":        expand,
                    "letters":       letters,
                    "threshold_pct": self._px_thresh[tok].get(),
                }
        ctok = self._px_custom_tok.get().strip()
        cexp = self._px_custom_exp.get().strip()
        if ctok and cexp:
            px_config[ctok] = {
                "expand":        cexp,
                "letters":       " ".join(list(ctok.upper())),
                "threshold_pct": self._px_custom_pct.get(),
            }
        opts["partial_expansion"] = px_config
        return opts

    def _detect_conflicts(self, text):
        conflicts = []
        seen = set()
        for tok in re.findall(r'[A-Za-z]{2,}', text):
            upper = tok.upper()
            if upper in seen: continue
            seen.add(upper)
            if upper in self.custom_dict:
                custom_out = self.custom_dict[upper]
                exp, was, _ = expand_abbreviation(tok, tok)
                if was and exp != custom_out:
                    conflicts.append((tok, exp, custom_out))
        return conflicts

    def _resolve_conflicts(self, text):
        conflicts = self._detect_conflicts(text)
        if not conflicts: return {}
        dlg = ConflictDialog(self, conflicts)
        self.wait_window(dlg)
        return dlg.results

    def _run_normalize(self):
        opts = self._get_options()
        if self.batch_files:
            self._run_batch(opts); return

        raw = self.input_text.get("1.0","end-1c")
        if not raw.strip():
            messagebox.showinfo("Empty","Please enter some text first."); return
        if not self._preload_scan_done:
            self._run_chem_detection(raw, wait=True)
        conflict_choices = self._resolve_conflicts(raw) if opts.get("abbreviations") else {}
        eff_dict = dict(self.custom_dict)
        for token, choice in conflict_choices.items():
            if choice == "standard": eff_dict.pop(token.upper(), None)

        excel_src = getattr(self, "_excel_source", None)

        if excel_src:
            normalized_by_row = {}
            combined_report = {
                "unverified":[], "ambiguous":[], "spelling":[],
                "corrections":[], "chemicals":[], "numbers":[], "symbols":[],
                "time":[], "temperature":[], "units":[], "diacritics":[],
                "abbreviations":[],
            }
            for r, text in excel_src["rows"]:
                result, report = normalize_text(text, opts, eff_dict,
                                                  chem_decisions=self._chem_decisions)
                normalized_by_row[r] = result
                for key, items in report.items():
                    if key not in combined_report:
                        continue
                    for item in items:
                        item = dict(item)
                        item["line"] = r   # remap to actual Excel row
                        combined_report[key].append(item)

            full_text = "\n".join(text for _, text in excel_src["rows"])
            out_dir = excel_src["path"].parent
            default_out = out_dir / f"{excel_src['path'].stem}_normalized.xlsx"
            save_path = filedialog.asksaveasfilename(
                title="Save Normalized Excel File",
                initialfile=default_out.name,
                initialdir=str(out_dir),
                defaultextension=".xlsx",
                filetypes=[("Excel file","*.xlsx")])
            if save_path:
                try:
                    write_excel_normalized(excel_src["path"], excel_src["sheet"],
                                            excel_src["column"], normalized_by_row, save_path)
                except Exception as e:
                    messagebox.showerror("Export error", str(e)); return

            preview = "\n".join(normalized_by_row[r] for r, _ in excel_src["rows"])
            self.output_text.config(state="normal")
            self.output_text.delete("1.0","end")
            self.output_text.insert("1.0", preview)
            self.output_text.config(state="disabled")

            report = combined_report
            filename_for_report = excel_src["path"].name
        else:
            result, report = normalize_text(raw, opts, eff_dict,
                                             chem_decisions=self._chem_decisions)
            self.output_text.config(state="normal")
            self.output_text.delete("1.0","end")
            self.output_text.insert("1.0", result)
            self.output_text.config(state="disabled")
            filename_for_report = "pasted text"

        self.last_norm_report = report
        self._last_unverified = [u["token"] for u in report.get("unverified", [])]

        needs_qc = bool(report.get("unverified") or report.get("ambiguous")
                        or report.get("spelling"))
        comments = []
        if report.get("unverified"):
            comments.append(f"{len(report['unverified'])} unverified token(s) found")
        if report.get("ambiguous"):
            comments.append(f"{len(report['ambiguous'])} ambiguous security code(s) — verify reading")
        if report.get("spelling"):
            comments.append(f"{len(report['spelling'])} spelling issue(s) detected")
        fr = {"filename": filename_for_report, "needs_qc": needs_qc,
              "comments": comments,
              "ambiguous": report.get("ambiguous", []),
              "spelling":  report.get("spelling", []),
              "changes":   flatten_report_changes(report)}

        row_label = "Row" if excel_src else "Line"
        default_dir = str(excel_src["path"].parent) if excel_src else str(Path.home())
        save_path2 = filedialog.asksaveasfilename(
            title="Save Normalization Report",
            initialfile="normalization_report.xlsx",
            initialdir=default_dir,
            defaultextension=".xlsx",
            filetypes=[("Excel file","*.xlsx")])
        if save_path2:
            try:
                export_qc_report_excel([fr], save_path2, row_label=row_label)
                messagebox.showinfo("Report saved", f"Report saved to:\n{save_path2}")
            except Exception as e:
                messagebox.showerror("Export error", str(e))

    def _run_batch(self, opts):
        out_folder = self.batch_files[0]["path"].parent / "normalized"
        out_folder.mkdir(exist_ok=True)
        file_reports = []
        processed = []
        any_xlsx = any(e["type"] == "xlsx" for e in self.batch_files)
        row_label = "Row" if any_xlsx else "Line"

        for entry in self.batch_files:
            fp = entry["path"]

            if entry["type"] == "xlsx":
                rows, _, _ = read_excel_column(fp, entry["sheet"], entry["column"])
                normalized_by_row = {}
                combined_report = {
                    "unverified":[], "ambiguous":[], "spelling":[],
                    "corrections":[], "chemicals":[], "numbers":[], "symbols":[],
                    "time":[], "temperature":[], "units":[], "diacritics":[],
                    "abbreviations":[],
                }
                for r, text in rows:
                    conflict_choices = self._resolve_conflicts(text) if opts.get("abbreviations") else {}
                    eff_dict = dict(self.custom_dict)
                    for token, choice in conflict_choices.items():
                        if choice == "standard": eff_dict.pop(token.upper(), None)
                    result, report = normalize_text(text, opts, eff_dict,
                                                      chem_decisions=self._chem_decisions)
                    normalized_by_row[r] = result
                    for key, items in report.items():
                        if key not in combined_report:
                            continue
                        for item in items:
                            item = dict(item)
                            item["line"] = r
                            combined_report[key].append(item)

                out_path = out_folder / f"{fp.stem}_normalized.xlsx"
                write_excel_normalized(fp, entry["sheet"], entry["column"],
                                        normalized_by_row, out_path)
                report = combined_report
            else:
                raw = fp.read_text(encoding="utf-8", errors="replace")
                conflict_choices = self._resolve_conflicts(raw) if opts.get("abbreviations") else {}
                eff_dict = dict(self.custom_dict)
                for token, choice in conflict_choices.items():
                    if choice == "standard": eff_dict.pop(token.upper(), None)
                result, report = normalize_text(raw, opts, eff_dict,
                                                 chem_decisions=self._chem_decisions)
                (out_folder / (fp.stem + "_normalized.txt")).write_text(result, encoding="utf-8")

            processed.append(fp.name)

            needs_qc = bool(report.get("unverified") or report.get("ambiguous")
                            or report.get("spelling"))
            comments = []
            if report.get("unverified"):
                comments.append(f"{len(report['unverified'])} unverified token(s)")
            if report.get("ambiguous"):
                comments.append(f"{len(report['ambiguous'])} ambiguous security code(s) — verify reading")
            if report.get("spelling"):
                comments.append(f"{len(report['spelling'])} spelling issue(s) detected")
            file_reports.append({
                "filename": fp.name,
                "needs_qc": needs_qc,
                "comments": comments,
                "ambiguous": report.get("ambiguous", []),
                "spelling":  report.get("spelling", []),
                "changes":   flatten_report_changes(report),
            })

        self.last_norm_report = {}
        self._last_unverified = []

        default_dir  = str(self.batch_files[0]["path"].parent)
        default_name = "normalization_report.xlsx"
        save_path = filedialog.asksaveasfilename(
            title="Save Normalization Report",
            initialfile=default_name,
            initialdir=default_dir,
            defaultextension=".xlsx",
            filetypes=[("Excel file","*.xlsx")])
        if save_path:
            try:
                export_qc_report_excel(file_reports, save_path, row_label=row_label)
                messagebox.showinfo(
                    "Done",
                    f"Processed {len(processed)} file(s).\n"
                    f"Normalized files → {out_folder}\n"
                    f"Report saved → {save_path}")
            except Exception as e:
                messagebox.showerror("Export error", str(e))
        else:
            messagebox.showinfo("Done",
                f"Processed {len(processed)} file(s).\nSaved to: {out_folder}")

    def _show_norm_report(self, report, batch_files=None):
        pass  # report is now Excel-only; no in-app text area


    def _copy_output(self):
        self.clipboard_clear()
        self.clipboard_append(self.output_text.get("1.0","end-1c"))
        messagebox.showinfo("Copied","Output copied to clipboard.")

    def _save_output(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text file","*.txt"),("All files","*.*")])
        if path:
            Path(path).write_text(self.output_text.get("1.0","end-1c"), encoding="utf-8")
            messagebox.showinfo("Saved", f"Saved to:\n{path}")


    def _qc_load_file(self):
        path = filedialog.askopenfilename(
            filetypes=[("Supported files","*.txt *.xlsx"),
                       ("Text files","*.txt"),
                       ("Excel files","*.xlsx"),
                       ("All files","*.*")])
        if not path:
            return
        path = Path(path)
        if path.suffix.lower() == ".xlsx":
            dlg = ColumnPickerDialog(self, path)
            self.wait_window(dlg)
            if not dlg.result:
                return
            sheet_name, column_name = dlg.result
            self._qc_files = [{"path": path, "type": "xlsx",
                                "sheet": sheet_name, "column": column_name}]
        else:
            self._qc_files = [{"path": path, "type": "txt"}]
        self.qc_status.config(text=f"1 file loaded: {path.name}", fg=C["accent_h"])

    def _qc_load_folder(self):
        folder = filedialog.askdirectory()
        if not folder:
            return
        all_files = sorted(Path(folder).glob("*.txt")) + sorted(Path(folder).glob("*.xlsx"))
        if not all_files:
            messagebox.showwarning("No files","No .txt or .xlsx files found."); return

        dlg = MultiFileSelectDialog(self, all_files)
        self.wait_window(dlg)
        if dlg.result is None or not dlg.result:
            return
        selected = dlg.result

        entries = []
        header_cache = {}
        for fp in selected:
            if fp.suffix.lower() == ".txt":
                entries.append({"path": fp, "type": "txt"})
                continue
            import openpyxl
            wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
            sheet_name = wb.sheetnames[0]
            ws = wb[sheet_name]
            header = tuple(str(c) if c is not None else f"Column {i+1}"
                           for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())))
            wb.close()
            if header in header_cache:
                sheet_used, col_used = header_cache[header]
            else:
                pdlg = ColumnPickerDialog(self, fp)
                self.wait_window(pdlg)
                if not pdlg.result:
                    continue
                sheet_used, col_used = pdlg.result
                header_cache[header] = (sheet_used, col_used)
            entries.append({"path": fp, "type": "xlsx",
                             "sheet": sheet_used, "column": col_used})

        if not entries:
            return
        self._qc_files = entries
        self.qc_status.config(text=f"{len(entries)} file(s) loaded", fg=C["accent_h"])

    def _run_qc(self):
        if not getattr(self, "_qc_files", []):
            messagebox.showinfo("No files","Load a file or folder first."); return
        opts = {k: v.get() for k, v in self._qc_opts.items()}
        opts["check_case"]    = self._qc_check_case.get()
        opts["expected_case"] = self._qc_case_var.get()

        any_xlsx = any(e["type"] == "xlsx" for e in self._qc_files)
        row_label = "Row" if any_xlsx else "Line"

        combined_issues, filenames = [], []
        file_reports = []
        for entry in self._qc_files:
            fp = entry["path"]
            if entry["type"] == "xlsx":
                rows, _, _ = read_excel_column(fp, entry["sheet"], entry["column"])
                text = "\n".join(t for _, t in rows)
                issues = qc_check_text(text, opts)
                row_map = {i+1: r for i, (r, _) in enumerate(rows)}
                for iss in issues:
                    iss["line"] = row_map.get(iss["line"], iss["line"])
                    iss["file"] = fp.name
            else:
                text = fp.read_text(encoding="utf-8", errors="replace")
                issues = qc_check_text(text, opts)
                for iss in issues:
                    iss["file"] = fp.name

            combined_issues.extend(issues)
            filenames.append(fp.name)

            comments = sorted({iss["issue_type"] for iss in issues})
            file_reports.append({
                "filename": fp.name,
                "needs_qc": bool(issues),
                "comments": [f"{len(issues)} issue(s): " + ", ".join(comments)] if issues else [],
                "ambiguous": [],
                "spelling": [{"line": iss["line"], "token": iss["token"], "suggestion": iss["suggestion"]}
                             for iss in issues if iss["issue_type"] == "Possible spelling issue"],
                "changes": [],
                "qc_issues": issues,
            })

        self.last_qc_issues = combined_issues
        self._qc_file_reports = file_reports
        report_text = build_qc_report(combined_issues, filenames=filenames)
        self._qc_report_text = report_text
        self._qc_row_label = row_label

        self.qc_result_text.config(state="normal")
        self.qc_result_text.delete("1.0","end")
        self.qc_result_text.insert("1.0", report_text)
        self.qc_result_text.config(state="disabled")

    def _export_qc_excel(self):
        if not getattr(self, "_qc_file_reports", None):
            messagebox.showinfo("No results","Run QC first."); return
        default_dir = str(self._qc_files[0]["path"].parent) if self._qc_files else str(Path.home())
        save_path = filedialog.asksaveasfilename(
            title="Save QC Report",
            initialfile="qc_report.xlsx",
            initialdir=default_dir,
            defaultextension=".xlsx",
            filetypes=[("Excel file","*.xlsx")])
        if not save_path:
            return
        try:
            export_qc_issues_excel(self._qc_file_reports, save_path,
                                    row_label=getattr(self, "_qc_row_label", "Line"))
            messagebox.showinfo("Saved", f"QC report saved to:\n{save_path}")
        except Exception as e:
            messagebox.showerror("Export error", str(e))


if __name__ == "__main__":
    app = NormalizerApp()
    app.mainloop()
